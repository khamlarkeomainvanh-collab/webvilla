from django.shortcuts import render,redirect,reverse
from django.db.models import Count, Sum, Q
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import uuid
import json
import re
from . import forms,models
from django.http import HttpResponseRedirect,HttpResponse
from django.core.mail import send_mail
from django.contrib.auth.models import Group, User
from django.contrib.auth import login as auth_login
from django.contrib.auth.forms import AuthenticationForm as _AuthForm
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib import messages
from django.conf import settings
from django.shortcuts import render, redirect
from .models import Product 
from django.shortcuts import get_object_or_404
from webpush import send_group_notification, send_user_notification

def home_view(request):
    if request.user.is_authenticated and not is_customer(request.user):
        return HttpResponseRedirect('afterlogin')
    cat_id = request.GET.get('cat')
    sub_id = request.GET.get('sub')
    color  = request.GET.get('color')

    base_qs = models.Product.objects.all()
    if cat_id:
        base_qs = base_qs.filter(category_id=cat_id)
    if sub_id:
        base_qs = base_qs.filter(subcategory_id=sub_id)

    available_colors = list(
        models.ProductColor.objects.filter(product__in=base_qs)
        .values_list('color_name', flat=True).distinct().order_by('color_name')
    )

    products = base_qs
    if color:
        products = products.filter(colors__color_name=color).distinct()
    products = products.prefetch_related('colors', 'extra_images')

    categories = models.Category.objects.prefetch_related('subcategories')
    if request.user.is_authenticated:
        cart = request.session.get('cart', {})
        product_count_in_cart = sum(cart.values())
    elif 'product_ids' in request.COOKIES:
        product_ids = request.COOKIES['product_ids']
        counter = product_ids.split('|')
        product_count_in_cart = len(set(counter))
    else:
        product_count_in_cart = 0
    return render(request, 'ecom/index.html', {
        'products': products,
        'categories': categories,
        'active_cat': cat_id,
        'active_sub': sub_id,
        'active_color': color,
        'available_colors': available_colors,
        'product_count_in_cart': product_count_in_cart,
        'closure_announcement': models.Announcement.objects.filter(kind='closed', is_active=True).order_by('-id').first(),
        'promo_announcements': models.Announcement.objects.filter(kind='promo', is_active=True),
    })




def process_order(request):
    if request.method == "POST":
        # ... ບັນທຶກອໍເດີ ...

        payload = {
            "title": "🥤 ມີອໍເດີໃໝ່ເຂົ້າມາ!",
            "body": f"ອໍເດີຈາກ: {request.user.username} ລາຄາ {total_price} ກີບ",
            "url": "/admin-advance-bookings/" # Link ໄປໜ້າຈັດການ
        }
        
        # ສົ່ງແຈ້ງເຕືອນ
        send_group_notification(group_name="admins", payload=payload, ttl=1000)

        return render(request, 'payment_success.html')


#for showing login button for admin(by sumit)
def adminclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return HttpResponseRedirect('adminlogin')


def admin_login_view(request):
    """Custom admin login (replaces Django's built-in LoginView) — on a
    successful login it also checks for orders that arrived while the admin
    was signed out, and stashes them in the session so the dashboard can pop
    up a "missed while you were away" notification once, every time they log in."""
    from django.contrib.auth import login as _auth_login
    from django.contrib.auth.forms import AuthenticationForm
    import json as _json

    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        previous_last_login = user.last_login  # capture before login() overwrites it
        _auth_login(request, user)

        if previous_last_login and user.is_staff:
            missed = models.Orders.objects.filter(order_date__gt=previous_last_login) \
                .exclude(delivery_type=WALKIN_DELIVERY_TYPE) \
                .select_related('product', 'customer__user').order_by('-order_date')[:20]
            if missed:
                request.session['admin_missed_orders'] = _json.dumps({
                    'count': len(missed),
                    'orders': [
                        {
                            'product':  o.product.name if o.product else (o.note or 'ສັ່ງເມນູຕາມໃຈ'),
                            'customer': o.customer.get_name if o.customer else '—',
                            'amount':   float(o.amount or 0),
                        }
                        for o in missed
                    ],
                })
        return HttpResponseRedirect('afterlogin')
    return render(request, 'ecom/adminlogin.html', {'form': form})


LAO_MOBILE_RE = re.compile(r'^20[259]\d{7}$')


def _normalize_lao_mobile(raw):
    digits = re.sub(r'\D', '', raw or '')
    if digits.startswith('856'):
        digits = digits[3:]
    return digits.lstrip('0')


def _is_valid_lao_mobile(raw):
    return bool(LAO_MOBILE_RE.match(_normalize_lao_mobile(raw)))


def customer_login_view(request):
    login_form = _AuthForm(request, data=request.POST or None)
    if request.method == 'POST':
        if login_form.is_valid():
            auth_login(request, login_form.get_user())
            return HttpResponseRedirect('afterlogin')
    return render(request, 'ecom/auth.html', context={
        'form': login_form,
        'userForm': forms.CustomerUserForm(),
        'customerForm': forms.CustomerForm(),
        'active_tab': 'login',
    })


def customer_signup_view(request):
    userForm = forms.CustomerUserForm()
    customerForm = forms.CustomerForm()
    if request.method == 'POST':
        userForm = forms.CustomerUserForm(request.POST)
        customerForm = forms.CustomerForm(request.POST, request.FILES)
        verify_method = request.POST.get('verify_method', 'camera')
        otp_ok = request.session.get('otp_verified', False)
        otp_mobile = request.session.get('otp_mobile', '')
        if userForm.is_valid() and customerForm.is_valid():
            if verify_method == 'otp' and not otp_ok:
                mydict = {'userForm': userForm, 'customerForm': customerForm, 'form': _AuthForm(), 'error': 'ກາລຸນາຢືນຢັນ OTP ກ່ອນ', 'active_tab': 'signup'}
                return render(request, 'ecom/auth.html', context=mydict)
            mobile_raw = request.POST.get('mobile', otp_mobile).strip()
            if not _is_valid_lao_mobile(mobile_raw):
                mydict = {
                    'userForm': userForm, 'customerForm': customerForm, 'form': _AuthForm(),
                    'error': 'ກະລຸນາໃສ່ເບີໂທທີ່ຖືກຕ້ອງ (ຮູບແບບ 20XXXXXXXX ຂອງລາວ)', 'active_tab': 'signup',
                }
                return render(request, 'ecom/auth.html', context=mydict)
            target_mobile = _normalize_lao_mobile(mobile_raw)
            for existing_cust in models.Customer.objects.exclude(mobile=''):
                if _normalize_lao_mobile(existing_cust.mobile) == target_mobile:
                    mydict = {
                        'userForm': userForm, 'customerForm': customerForm, 'form': _AuthForm(),
                        'error': 'ເບີໂທນີ້ລົງທະບຽນເປັນສະມາຊິກແລ້ວ — ກະລຸນາໄປໜ້າ "ເຂົ້າສູ່ລະບົບ" ແທນ (ໃຊ້ຊື່ ຫຼື ເບີໂທ ເພື່ອເຂົ້າ)', 'active_tab': 'signup',
                    }
                    return render(request, 'ecom/auth.html', context=mydict)
            user = userForm.save(commit=False)
            # auto-generate unique username from mobile number
            mobile_clean = mobile_raw.lstrip('+').replace('856', '', 1).replace(' ', '').lstrip('0') or 'user'
            base_uname = 'u' + mobile_clean
            uname, counter = base_uname, 1
            while User.objects.filter(username=uname).exists():
                uname = base_uname + str(counter); counter += 1
            user.username = uname
            user.set_password(user.password)
            user.save()
            customer = customerForm.save(commit=False)
            customer.user = user
            if verify_method == 'otp' and otp_mobile:
                customer.mobile = otp_mobile
            customer.save()
            Group.objects.get_or_create(name='CUSTOMER')[0].user_set.add(user)
            request.session.pop('otp_verified', None)
            request.session.pop('otp_mobile', None)
            request.session.pop('otp_code', None)
            user.backend = 'django.contrib.auth.backends.ModelBackend'
            auth_login(request, user)
            return HttpResponseRedirect('afterlogin')
        mydict = {'userForm': userForm, 'customerForm': customerForm, 'form': _AuthForm(), 'error': 'ກາລຸນາກວດຂໍ້ມູນທີ່ປ້ອນໃຫ້ຖືກຕ້ອງ', 'active_tab': 'signup'}
        return render(request, 'ecom/auth.html', context=mydict)
    return render(request, 'ecom/auth.html', context={'userForm': userForm, 'customerForm': customerForm, 'form': _AuthForm(), 'active_tab': 'signup'})


def _find_customer_by_mobile(mobile_raw):
    target = _normalize_lao_mobile(mobile_raw)
    if not target:
        return None
    for c in models.Customer.objects.exclude(mobile='').select_related('user'):
        if _normalize_lao_mobile(c.mobile) == target:
            return c
    return None


def forgot_password_view(request):
    """Self-service password reset — verified by phone number match only
    (no SMS/email infra is configured yet). Low-stakes account (no payment
    info stored), so this is an acceptable tradeoff for a small shop."""
    stage = 'phone'
    error = None
    mobile_value = ''

    if request.method == 'POST':
        mobile_raw = request.POST.get('mobile', '').strip()
        mobile_value = mobile_raw
        cust = _find_customer_by_mobile(mobile_raw)

        if 'new_password' in request.POST:
            stage = 'password'
            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')
            if not cust:
                error = 'ບໍ່ພົບບັນຊີທີ່ໃຊ້ເບີໂທນີ້ໃນລະບົບ'
                stage = 'phone'
            elif len(new_password) < 6:
                error = 'ລະຫັດຜ່ານໃໝ່ຕ້ອງມີຢ່າງໜ້ອຍ 6 ໂຕ'
            elif new_password != confirm_password:
                error = 'ລະຫັດຜ່ານທັງສອງຊ່ອງບໍ່ຄືກັນ'
            else:
                cust.user.set_password(new_password)
                cust.user.save()
                return render(request, 'ecom/forgot_password.html', {'stage': 'done'})
        else:
            if not cust:
                error = 'ບໍ່ພົບບັນຊີທີ່ໃຊ້ເບີໂທນີ້ໃນລະບົບ'
            else:
                stage = 'password'

    return render(request, 'ecom/forgot_password.html', {
        'stage': stage, 'error': error, 'mobile_value': mobile_value,
    })


def ajax_send_otp(request):
    import random as _rnd
    from django.conf import settings as _cfg
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'msg': 'POST only'})
    mobile = request.POST.get('mobile', '').strip()
    if not mobile or len(mobile) < 8:
        return JsonResponse({'ok': False, 'msg': 'ກາລຸນາໃສ່ເບີໂທທີ່ຖືກຕ້ອງ'})

    otp = str(_rnd.randint(100000, 999999))
    request.session['otp_code']     = otp
    request.session['otp_mobile']   = mobile
    request.session['otp_verified'] = False

    sid   = getattr(_cfg, 'TWILIO_ACCOUNT_SID', '')
    token = getattr(_cfg, 'TWILIO_AUTH_TOKEN', '')
    frm   = getattr(_cfg, 'TWILIO_FROM_NUMBER', '')

    # ຍັງບໍ່ໄດ້ຕັ້ງຄ່າ → dev mode
    if not sid or sid.startswith('ACxxx') or not token or token.startswith('xxx'):
        print(f'[OTP DEV] Mobile={mobile} Code={otp}')
        return JsonResponse({'ok': True, 'msg': f'[DEV] OTP: {otp}', 'dev_otp': otp})

    try:
        from twilio.rest import Client as _TwilioClient
        client = _TwilioClient(sid, token)
        client.messages.create(
            body=f'EX ມໍເຕີ້\nລະຫັດ OTP: {otp}\n(ໃຊ້ໄດ້ 5 ນາທີ - ຢ່າໃຫ້ໃຜ)',
            from_=frm,
            to=mobile,
        )
        return JsonResponse({'ok': True, 'msg': f'ສົ່ງ OTP ໄປ {mobile} ແລ້ວ ✅'})
    except ImportError:
        print(f'[OTP DEV - twilio not installed] Mobile={mobile} Code={otp}')
        return JsonResponse({'ok': True, 'msg': f'[DEV] OTP: {otp}', 'dev_otp': otp})
    except Exception as _e:
        err = str(_e)
        print(f'[OTP ERROR] {err}')
        if 'unverified' in err.lower():
            return JsonResponse({'ok': False, 'msg': 'ເບີ ' + mobile + ' ຍັງບໍ່ໄດ້ verify ໃນ Twilio trial — ໄປ console.twilio.com/verified-caller-ids ເພີ່ມເບີກ່ອນ'})
        if 'invalid' in err.lower() or 'not a valid' in err.lower():
            return JsonResponse({'ok': False, 'msg': 'ຮູບແບບເບີໂທບໍ່ຖືກ — ຕ້ອງເປັນ +856XXXXXXXXX'})
        return JsonResponse({'ok': False, 'msg': f'ສົ່ງ OTP ບໍ່ສຳເລັດ: {err[:120]}'})


def ajax_verify_otp(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'msg': 'POST only'})
    code = request.POST.get('code', '').strip()
    stored = request.session.get('otp_code', '')
    if not stored:
        return JsonResponse({'ok': False, 'msg': 'ກາລຸນາຂໍ OTP ກ່ອນ'})
    if code == stored:
        request.session['otp_verified'] = True
        return JsonResponse({'ok': True, 'mobile': request.session.get('otp_mobile', '')})
    return JsonResponse({'ok': False, 'msg': 'OTP ບໍ່ຖືກຕ້ອງ — ກວດສອບອີກຄັ້ງ'})

#-----------for checking user iscustomer
def is_customer(user):
    return user.groups.filter(name='CUSTOMER').exists()



#---------AFTER ENTERING CREDENTIALS WE CHECK WHETHER USERNAME AND PASSWORD IS OF ADMIN,CUSTOMER
def afterlogin_view(request):
    if is_customer(request.user):
        return redirect('/')
    else:
        return redirect('admin-dashboard')

#---------------------------------------------------------------------------------
#------------------------ ADMIN RELATED VIEWS START ------------------------------
#---------------------------------------------------------------------------------
@login_required(login_url='adminlogin')
def admin_dashboard_view(request):
    from datetime import date as _dc, datetime as _dtt, timedelta as _td
    from django.utils import timezone as _tz
    from datetime import date as date_cls, timedelta, timezone as _dtz2
    from datetime import datetime as dt_cls2

    _real_today = _tz.localdate()

    # ── ເລືອກວັນທີ (GET ?date=) ──
    date_str = request.GET.get('date', '')
    try:
        selected_date = date_cls.fromisoformat(date_str)
    except ValueError:
        selected_date = _real_today

    prev_date = selected_date - timedelta(days=1)
    next_date = selected_date + timedelta(days=1)
    is_today  = (selected_date == _real_today)

    # ─ ຊ່ວງເວລາຂອງ selected_date ─
    _ts = _tz.make_aware(_dtt(selected_date.year, selected_date.month, selected_date.day, 0, 0, 0))
    _te = _tz.make_aware(_dtt(selected_date.year, selected_date.month, selected_date.day, 23, 59, 59))

    # Cards (all-time)
    customercount   = models.Customer.objects.all().count()
    productcount    = models.Product.objects.all().count()
    ordercount      = models.Orders.objects.all().count()
    # Status donut counts vehicle QUANTITY (not row-count), so it always sums
    # to the same "ອໍເດີ້ທັງໝົດ" total shown in the all-time summary card.
    def _units_for(status):
        return models.Orders.objects.filter(status=status).aggregate(q=Sum('quantity'))['q'] or 0
    delivered_count = _units_for('Delivered')
    pending_count   = _units_for('Pending')
    confirmed_count = _units_for('Confirmed')
    processing_count= _units_for('Processing')
    cancelled_count = _units_for('Cancelled')
    other_count     = ordercount - (delivered_count + pending_count)

    # ── All-time summary cards: real revenue/deposit/outstanding totals,
    # not scoped to the selected day ──
    _active_orders_qs = models.Orders.objects.exclude(status='Cancelled')
    # "ອໍເດີ້ທັງໝົດ" counts actual vehicle QUANTITY, not line-item rows — a
    # customer booking 4 vehicles across 2 different models is 4 orders, not 2.
    total_units_alltime = _active_orders_qs.aggregate(q=Sum('quantity'))['q'] or 0
    total_revenue_alltime = float(_active_orders_qs.aggregate(t=Sum('amount'))['t'] or 0)
    # Deposit income only counts orders an admin has actually verified as
    # paid (see the "ຢືນຢັນມັດຈຳ" step on admin-advance-bookings) — a real
    # collected-cash figure, not just a theoretical 20% of everything booked.
    total_deposit_alltime = round(
        float(_active_orders_qs.filter(deposit_verified=True).aggregate(t=Sum('amount'))['t'] or 0) * 0.20
    )
    # Outstanding balance — the 80% still owed on bookings that haven't been
    # collected yet (walk-in sales are paid in full on the spot, so excluded).
    _outstanding_qs = models.Orders.objects.filter(
        status__in=['Pending', 'Confirmed', 'Processing']
    ).exclude(delivery_type=WALKIN_DELIVERY_TYPE)
    total_outstanding_alltime = round(
        float(_outstanding_qs.aggregate(t=Sum('amount'))['t'] or 0) * 0.80
    )

    # Which vehicle models have sold today, and how many of each.
    today_model_sales = list(
        models.Orders.objects.filter(order_date__gte=_ts, order_date__lte=_te)
        .exclude(status='Cancelled')
        .exclude(product__isnull=True)
        .values('product__name')
        .annotate(qty=Sum('quantity'))
        .order_by('-qty')
    )
    today_model_sales_total = sum(m['qty'] for m in today_model_sales)

    # Stats ຂອງ selected_date — advance bookings only count once actually
    # fulfilled (see _revenue_orders_qs), not on the day they were booked.
    today_orders_qs = _revenue_orders_qs(_ts, _te).select_related('product')

    # ── Source filter (?source=) — ອອນລາຍ / ໜ້າຮ້ານ, ຄືກັບ admin-view-booking.
    # ຄ່າເລີ່ມຕົ້ນ "all" ຄືເກົ່າ (ລວມທຸກປະເພດ), ຕົວກອງນີ້ນຳໃຊ້ກັບທັງ card ແລະ
    # ຕາຕະລາງລຸ່ມນຳກັນ ເພື່ອບໍ່ໃຫ້ຕົວເລກຄາດເຄື່ອນກັນ.
    source = request.GET.get('source', 'all')
    if source not in ('all', 'online', 'walkin'):
        source = 'all'
    if source == 'online':
        today_orders_qs = today_orders_qs.exclude(delivery_type=WALKIN_DELIVERY_TYPE)
    elif source == 'walkin':
        today_orders_qs = today_orders_qs.filter(delivery_type=WALKIN_DELIVERY_TYPE)

    today_order_count     = today_orders_qs.count()

    def _amt(o):
        if o.amount:
            return float(o.amount)
        if o.product and o.product.price:
            return float(o.product.price) * int(o.quantity or 1)
        return 0.0

    today_revenue         = sum(_amt(o) for o in today_orders_qs)
    today_pending_count   = today_orders_qs.filter(status='Pending').count()
    today_confirmed_count = today_orders_qs.filter(status='Confirmed').count()
    today_processing_count= today_orders_qs.filter(status='Processing').count()
    today_delivered_count = today_orders_qs.filter(status='Delivered').count()
    today_cancelled_count = today_orders_qs.filter(status='Cancelled').count()
    current_year  = _real_today.year
    current_month = _real_today.strftime('%Y-%m')
    current_date  = selected_date.strftime('%Y-%m-%d')
    today_str     = _real_today.strftime('%Y-%m-%d')

    sales_data = [10, 25, 15, 30, 45, 35, 55]

    # ── ຕາຕະລາງ Orders grouped by order_group — ໃຊ້ set ດຽວກັນກັບການ໋ໍນັບຍອດ
    # ຂ້າງເທິງ (_revenue_orders_qs) ເພື່ອໃຫ້ ຂາຍໜ້າຮ້ານ + ຈອງອອນລາຍ + ຈອງລ່ວງໜ້າ
    # (ທີ່ຮັບແລ້ວມື້ນີ້) ທັງໝົດສະແດງຢູ່ນີ້ ແລະ ຍອດຂາຍຕົງກັນທຸກບ່ອນ. ──
    from django.utils import timezone as _tz3
    all_day_orders = today_orders_qs.select_related('customer__user').order_by('id')

    day_revenue = today_revenue

    _LAO_TZ2 = _dtz2(timedelta(hours=7))

    # Build order groups
    _all_groups = {}
    _group_order = []
    for _o in all_day_orders:
        _key = _o.order_group or str(_o.id)
        if _key not in _all_groups:
            # ຈອງລ່ວງໜ້າ — ໃຊ້ເວລາ "ຮັບແທ້ຈິງ" (fulfilled_at) ແທນເວລາ "ຈອງ" (order_date)
            _time_src = _o.order_date
            _all_groups[_key] = {
                'key': _key,
                'canonical': _o,
                'orders': [],
                'customer': _o.customer,
                'local_time': _time_src.astimezone(_LAO_TZ2).strftime("%H:%M") if _time_src else "--:--",
                'is_advance': bool(_o.pickup_date),
                'is_walkin':  _o.delivery_type == WALKIN_DELIVERY_TYPE,
            }
            _group_order.append(_key)
        _all_groups[_key]['orders'].append(_o)

    for _i, _key in enumerate(_group_order):
        _g = _all_groups[_key]
        _g['queue_num'] = _i + 1
        _g['item_count'] = len(_g['orders'])
        _g['total_amount'] = sum(float(_ord.amount or 0) for _ord in _g['orders'])
        _g['order_ids'] = [_ord.id for _ord in _g['orders']]

    status_counts = {
        'All':       len(_group_order),
        'Pending':    sum(1 for _k in _group_order if (_all_groups[_k]['canonical'].status or 'Pending') == 'Pending'),
        'Confirmed':  sum(1 for _k in _group_order if (_all_groups[_k]['canonical'].status or 'Pending') == 'Confirmed'),
        'Processing': sum(1 for _k in _group_order if (_all_groups[_k]['canonical'].status or 'Pending') == 'Processing'),
        'Delivered':  sum(1 for _k in _group_order if (_all_groups[_k]['canonical'].status or 'Pending') == 'Delivered'),
        'Cancelled':  sum(1 for _k in _group_order if (_all_groups[_k]['canonical'].status or 'Pending') == 'Cancelled'),
    }
    day_order_count = len(_group_order)

    active_status = request.GET.get('status', 'Pending')
    if active_status not in ['Pending', 'Confirmed', 'Processing', 'Delivered', 'Cancelled', 'All']:
        active_status = 'Pending'

    if active_status == 'All':
        data = list(reversed([_all_groups[_k] for _k in _group_order]))
    else:
        data = [_all_groups[_k] for _k in _group_order
                if (_all_groups[_k]['canonical'].status or 'Pending') == active_status]
        if active_status != 'Pending':
            data = list(reversed(data))

    # ຂໍ້ມູນຍອດຂາຍລາຍວັນ 30 ວັນຫຼ້າສຸດ — group ດ້ວຍ Python ຫຼີກ SQLite functions
    from datetime import datetime as dt_dash
    today = date_cls.today()
    range_start = _tz.make_aware(dt_dash(today.year, today.month, today.day, 0, 0, 0)) - timedelta(days=29)
    range_end   = _tz.make_aware(dt_dash(today.year, today.month, today.day, 23, 59, 59))
    recent_rows = _revenue_orders_qs(range_start, range_end).values_list('order_date', 'amount')

    date_map = {(today - timedelta(days=i)).isoformat(): {'orders': 0, 'amount': 0.0} for i in range(29, -1, -1)}
    for odt, amt in recent_rows:
        rev_dt = odt
        if rev_dt is None:
            continue
        day_key = rev_dt.date().isoformat() if hasattr(rev_dt, 'date') else str(rev_dt)[:10]
        if day_key in date_map:
            date_map[day_key]['orders'] += 1
            date_map[day_key]['amount'] += float(amt or 0)

    daily_labels = json.dumps([k[5:] for k in date_map.keys()])   # MM-DD
    daily_full_dates = json.dumps(list(date_map.keys()))           # YYYY-MM-DD
    daily_orders = json.dumps([v['orders'] for v in date_map.values()])
    daily_amounts = json.dumps([v['amount'] for v in date_map.values()])

    latest_order = models.Orders.objects.order_by('-id').first()
    last_order_id = latest_order.id if latest_order else 0

    mydict = {
        'customercount':    customercount,
        'productcount':     productcount,
        'ordercount':       ordercount,
        'data':             data,
        'delivered_count':  delivered_count,
        'pending_count':    pending_count,
        'processing_count': processing_count,
        'confirmed_count':  confirmed_count,
        'cancelled_count':  cancelled_count,
        'other_count':      other_count,
        'total_units_alltime':       total_units_alltime,
        'total_revenue_alltime':     total_revenue_alltime,
        'total_deposit_alltime':     total_deposit_alltime,
        'total_outstanding_alltime': total_outstanding_alltime,
        'today_model_sales':         today_model_sales,
        'today_model_sales_total':   today_model_sales_total,
        'today_order_count':       today_order_count,
        'today_revenue':           today_revenue,
        'today_pending_count':     today_pending_count,
        'today_confirmed_count':   today_confirmed_count,
        'today_processing_count':  today_processing_count,
        'today_delivered_count':   today_delivered_count,
        'today_cancelled_count':   today_cancelled_count,
        'current_year':    current_year,
        'current_month':   current_month,
        'current_date':    current_date,
        'today_str':       today_str,
        'day_revenue':      day_revenue,
        'sales_data':       sales_data,
        'daily_labels':      daily_labels,
        'daily_full_dates':  daily_full_dates,
        'daily_orders':      daily_orders,
        'daily_amounts':     daily_amounts,
        'last_order_id':    last_order_id,
        'selected_date':    selected_date,
        'prev_date':        prev_date,
        'next_date':        next_date,
        'is_today':         is_today,
        'day_order_count':  day_order_count,
        'active_status':    active_status,
        'status_counts':    status_counts,
        'source':           source,
    }
    return render(request, 'ecom/admin_dashboard.html', context=mydict)


def _revenue_orders_qs(start_dt=None, end_dt=None):
    """Orders that count as revenue (optionally within [start_dt, end_dt]) —
    every order counts on the day it was placed (order_date), including
    advance ("ຈອງລ່ວງໜ້າ") bookings that haven't been collected yet. Pass no
    bounds for an all-time queryset."""
    qs = models.Orders.objects.all()
    if start_dt is not None and end_dt is not None:
        qs = qs.filter(order_date__gte=start_dt, order_date__lte=end_dt)
    return qs


def _revenue_date(order_dict):
    """Given a dict with an order_date key (e.g. from .values()), return the
    date the row's revenue is attributed to — always the booking date."""
    return order_dict['order_date']


@login_required(login_url='adminlogin')
def admin_finance_view(request):
    from django.db.models import Sum, Count
    import datetime as _dt
    import calendar as _cal
    import json
    from decimal import Decimal
    from datetime import timezone as _dtz, timedelta as _td

    _lao_tz  = _dtz(_td(hours=7))
    now_lao  = _dt.datetime.now(_lao_tz)
    real_today = now_lao.date()

    profit_percent = float(models.FinanceSettings.get_solo().profit_percent)

    # ── Single date param controls all stats ──
    sel_date_param = request.GET.get('sel_date', str(real_today))
    try:
        sel_date_obj = _dt.date.fromisoformat(sel_date_param)
    except ValueError:
        sel_date_obj = real_today
    if sel_date_obj > real_today:
        sel_date_obj = real_today

    sel_year  = sel_date_obj.year
    sel_month = sel_date_obj.month
    sel_day   = sel_date_obj.day
    is_sel_today = (sel_date_obj == real_today)
    today_str    = real_today.strftime('%Y-%m-%d')
    sel_date_str = sel_date_obj.strftime('%Y-%m-%d')

    # ── Selected date stats ──
    sd_start = _dt.datetime(sel_year, sel_month, sel_day, 0, 0, 0, tzinfo=_lao_tz)
    sd_end   = sd_start + _td(days=1) - _td(seconds=1)
    sd_orders = _revenue_orders_qs(sd_start, sd_end).select_related('product')
    today_revenue = float(sd_orders.aggregate(t=Sum('amount'))['t'] or 0)
    today_count   = sd_orders.count()
    today_profit  = round(today_revenue * profit_percent / 100, 2)
    today_status  = {
        'Pending':    sd_orders.filter(status='Pending').count(),
        'Processing': sd_orders.filter(status='Processing').count(),
        'Delivered':  sd_orders.filter(status='Delivered').count(),
        'Cancelled':  sd_orders.filter(status='Cancelled').count(),
    }

    class _Dec(json.JSONEncoder):
        def default(self, o):
            return float(o) if isinstance(o, Decimal) else super().default(o)

    sel_orders_list = list(sd_orders.values('id', 'product__name', 'quantity', 'amount', 'status', 'delivery_type', 'deposit_verified'))
    sel_orders_json = json.dumps(sel_orders_list, cls=_Dec)

    # ── Month stats (selected year/month) ──
    month_start = _dt.datetime(sel_year, sel_month, 1, tzinfo=_lao_tz)
    if sel_month == 12:
        month_end = _dt.datetime(sel_year + 1, 1, 1, tzinfo=_lao_tz) - _td(seconds=1)
    else:
        month_end = _dt.datetime(sel_year, sel_month + 1, 1, tzinfo=_lao_tz) - _td(seconds=1)
    month_orders  = _revenue_orders_qs(month_start, month_end)
    month_revenue = float(month_orders.aggregate(t=Sum('amount'))['t'] or 0)
    month_count   = month_orders.count()
    month_start_d = month_start.date()
    month_end_d   = month_end.date()
    month_profit  = round(month_revenue * profit_percent / 100, 2)

    # ── Year stats ──
    year_start = _dt.datetime(sel_year, 1, 1, tzinfo=_lao_tz)
    year_end   = _dt.datetime(sel_year, 12, 31, 23, 59, 59, tzinfo=_lao_tz)
    year_orders  = _revenue_orders_qs(year_start, year_end)
    year_revenue = float(year_orders.aggregate(t=Sum('amount'))['t'] or 0)
    year_count   = year_orders.count()
    year_profit  = round(year_revenue * profit_percent / 100, 2)

    # ── Monthly breakdown ──
    month_labels = ['ມ.ກ','ກ.ພ','ມີ.ນ','ເມ.ສ','ພ.ພ','ມິ.ຖ','ກ.ລ','ສ.ຫ','ກ.ຍ','ຕ.ລ','ພ.ຈ','ທ.ວ']
    monthly_revenue = [0.0] * 12
    monthly_count   = [0]   * 12
    for o in year_orders.values('order_date', 'amount', 'pickup_date', 'fulfilled_at'):
        # advance bookings count on the day they were fulfilled, not booked
        revenue_dt = o['order_date']
        m = revenue_dt.astimezone(_lao_tz).month - 1
        monthly_revenue[m] += float(o['amount'] or 0)
        monthly_count[m]   += 1
    monthly_profit = [round(monthly_revenue[i] * profit_percent / 100, 2) for i in range(12)]

    # ── Status all time ──
    status_all = {s: models.Orders.objects.filter(status=s).count()
                  for s in ['Pending', 'Confirmed', 'Processing', 'Delivered', 'Cancelled']}

    # ── Daily breakdown for selected month ──
    _, days_in_month = _cal.monthrange(sel_year, sel_month)
    daily_rev  = [0.0] * days_in_month
    for o in month_orders.values('order_date', 'amount', 'pickup_date', 'fulfilled_at'):
        revenue_dt = o['order_date']
        d_idx = (revenue_dt.astimezone(_lao_tz).date() - month_start_d).days
        if 0 <= d_idx < days_in_month:
            daily_rev[d_idx] += float(o['amount'] or 0)
    daily_profit = [round(daily_rev[i] * profit_percent / 100, 2) for i in range(days_in_month)]

    available_years  = list(range(real_today.year, real_today.year - 5, -1))
    total_orders_all = sum(status_all.values())

    mydict = {
        'today_revenue':    today_revenue,
        'today_count':      today_count,
        'today_profit':     today_profit,
        'sel_orders_json':  sel_orders_json,
        'today_status':     today_status,
        'is_sel_today':     is_sel_today,
        'sel_date_str':     sel_date_str,
        'today_str':        today_str,
        'today':            sel_date_obj,
        'month_revenue':    month_revenue,
        'month_count':      month_count,
        'month_profit':     month_profit,
        'year_revenue':     year_revenue,
        'year_count':       year_count,
        'year_profit':      year_profit,
        'sel_year':         sel_year,
        'sel_month':        sel_month,
        'month_labels':     month_labels,
        'monthly_revenue':  monthly_revenue,
        'monthly_count':    monthly_count,
        'monthly_profit':   monthly_profit,
        'status_all':       status_all,
        'daily_rev':        daily_rev,
        'daily_profit':     daily_profit,
        'available_years':  available_years,
        'total_orders_all': total_orders_all,
        'profit_percent':   profit_percent,
    }
    return render(request, 'ecom/admin_finance.html', context=mydict)


@login_required(login_url='adminlogin')
def set_profit_percent_view(request):
    from decimal import Decimal, InvalidOperation
    if request.method == 'POST':
        try:
            pct = Decimal(request.POST.get('profit_percent', '30'))
            if pct < 0:
                pct = Decimal('0')
            if pct > 100:
                pct = Decimal('100')
        except (InvalidOperation, ValueError, TypeError):
            pct = Decimal('30')
        settings_obj = models.FinanceSettings.get_solo()
        settings_obj.profit_percent = pct
        settings_obj.save(update_fields=['profit_percent'])
    return redirect(request.META.get('HTTP_REFERER', 'admin-finance'))


@login_required(login_url='adminlogin')
def add_expense_view(request):
    if request.method == 'POST':
        f = forms.ExpenseForm(request.POST)
        if f.is_valid():
            f.save()
    return redirect('admin-finance')


@login_required(login_url='adminlogin')
def delete_expense_view(request, pk):
    try:
        models.Expense.objects.get(id=pk).delete()
    except models.Expense.DoesNotExist:
        pass
    return redirect('admin-finance')


WALKIN_DELIVERY_TYPE = 'Walkin'
_DEPOSIT_GATE_STATUSES = {'Confirmed', 'Processing', 'Delivered'}


@login_required(login_url='adminlogin')
def admin_walkin_sale_view(request):
    from django.db.models import Sum
    from django.utils import timezone as _tz

    if request.method == 'POST':
        import json as _json

        payment  = request.POST.get('payment_method', 'Cash')
        note_in  = request.POST.get('note', '').strip()
        note     = 'ຂາຍໜ້າຮ້ານ' + (f' — {note_in}' if note_in else '')

        try:
            items = _json.loads(request.POST.get('items', '') or '[]')
        except (ValueError, TypeError):
            items = []

        if items:
            group_id = str(uuid.uuid4())
            for it in items:
                try:
                    qty = max(1, int(it.get('quantity', 1) or 1))
                except (ValueError, TypeError):
                    continue

                product = None
                pid = it.get('product')
                if pid:
                    try:
                        product = models.Product.objects.get(id=pid)
                    except models.Product.DoesNotExist:
                        product = None

                price_raw = str(it.get('unit_price', '')).strip()
                try:
                    unit_price = int(price_raw) if price_raw else (product.price if product else 0)
                except (ValueError, TypeError):
                    continue

                if product:
                    item_note = note
                else:
                    # Off-menu item typed in by hand — no matching Product row, so the
                    # typed name becomes the note (shown wherever product.name normally would be).
                    custom_name = str(it.get('name', '')).strip()[:80]
                    if not custom_name:
                        continue
                    item_note = custom_name
                    if note_in:
                        item_note += f' ({note_in})'

                models.Orders.objects.create(
                    customer=None,
                    product=product,
                    quantity=qty,
                    amount=unit_price * qty,
                    status='Delivered',
                    order_group=group_id,
                    email='',
                    mobile='',
                    address='ໜ້າຮ້ານ',
                    delivery_type=WALKIN_DELIVERY_TYPE,
                    payment_method=payment,
                    note=item_note,
                )
        return redirect('admin-walkin-sale')

    from datetime import datetime as _dtt, timedelta as _td, date as _dtdate
    today_local = _tz.localdate()

    view_day_param = request.GET.get('day', '')
    try:
        view_day = _dtdate.fromisoformat(view_day_param) if view_day_param else today_local
    except (ValueError, TypeError):
        view_day = today_local

    day_start = _tz.make_aware(_dtt(view_day.year, view_day.month, view_day.day))
    day_end   = day_start + _td(days=1)
    sales = models.Orders.objects.filter(
        delivery_type=WALKIN_DELIVERY_TYPE, order_date__gte=day_start, order_date__lt=day_end
    ).select_related('product').order_by('-id')
    sales_agg   = sales.aggregate(t=Sum('amount'), q=Sum('quantity'))
    total_today = sales_agg['t'] or 0
    qty_today   = sales_agg['q'] or 0
    cash_today     = sales.filter(payment_method='Cash').aggregate(t=Sum('amount'))['t'] or 0
    transfer_today = sales.filter(payment_method='Transfer').aggregate(t=Sum('amount'))['t'] or 0

    return render(request, 'ecom/admin_walkin_sale.html', {
        'products':       models.Product.objects.filter(is_available=True).order_by('name'),
        'sales':          sales,
        'total_today':    total_today,
        'qty_today':      qty_today,
        'cash_today':     cash_today,
        'transfer_today': transfer_today,
        'today_local':  view_day,
        'is_today':     view_day == today_local,
        'view_day_str': view_day.strftime('%Y-%m-%d'),
        'prev_day_str': (view_day - _td(days=1)).strftime('%Y-%m-%d'),
        'next_day_str': (view_day + _td(days=1)).strftime('%Y-%m-%d'),
        'today_str':    today_local.strftime('%Y-%m-%d'),
    })


@login_required(login_url='adminlogin')
def delete_walkin_sale_view(request, pk):
    try:
        models.Orders.objects.get(id=pk, delivery_type=WALKIN_DELIVERY_TYPE).delete()
    except models.Orders.DoesNotExist:
        pass
    day_param = request.POST.get('day') or request.GET.get('day')
    return redirect(f"/admin-walkin-sale?day={day_param}" if day_param else 'admin-walkin-sale')


@login_required(login_url='adminlogin')
def edit_walkin_sale_view(request, pk):
    if request.method == 'POST':
        try:
            sale = models.Orders.objects.get(id=pk, delivery_type=WALKIN_DELIVERY_TYPE)
            qty = max(1, int(request.POST.get('quantity', 1) or 1))
            price_raw = re.sub(r'\D', '', request.POST.get('unit_price', ''))
            unit_price = int(price_raw) if price_raw else 0
            payment = request.POST.get('payment_method', sale.payment_method)
            sale.quantity = qty
            sale.amount = unit_price * qty
            sale.payment_method = payment
            sale.save(update_fields=['quantity', 'amount', 'payment_method'])
        except (models.Orders.DoesNotExist, ValueError, TypeError):
            pass
    day_param = request.POST.get('day') or request.GET.get('day')
    return redirect(f"/admin-walkin-sale?day={day_param}" if day_param else 'admin-walkin-sale')


def _walkin_sales_for_period(request):
    """Shared query: walk-in Orders within the period resolved by _resolve_invoice_period()."""
    from django.utils import timezone as _tz
    from datetime import datetime as _dtt, timedelta as _td
    start_d, end_d, period_label, invoice_no = _resolve_invoice_period(request)
    _lao_tz_start = _tz.make_aware(_dtt(start_d.year, start_d.month, start_d.day))
    _lao_tz_end   = _tz.make_aware(_dtt(end_d.year, end_d.month, end_d.day)) + _td(days=1)
    sales = models.Orders.objects.filter(
        delivery_type=WALKIN_DELIVERY_TYPE, order_date__gte=_lao_tz_start, order_date__lt=_lao_tz_end
    ).select_related('product').order_by('order_date')
    return sales, period_label, invoice_no


@login_required(login_url='adminlogin')
def export_walkin_excel_view(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.utils import timezone as _tz
    from django.db.models import Sum

    sales, period_label, invoice_no = _walkin_sales_for_period(request)
    total = sales.aggregate(t=Sum('amount'))['t'] or 0
    today = _tz.localdate()

    LAO_FONT = "Phetsarath OT"
    def hfont(size=12): return Font(name=LAO_FONT, bold=True, color="FFFFFF", size=size)
    def dfont(bold=False, size=11, color="000000"): return Font(name=LAO_FONT, bold=bold, size=size, color=color)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    date_al = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def dotfmt(n):
        if n is None: return '—'
        return f"{int(round(float(n))):,}".replace(',', '.') + ' ກີບ'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ຍອດຂາຍໜ້າຮ້ານ"[:31]
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]; c.value = f"ຍອດຂາຍໜ້າຮ້ານ — {period_label}"
    c.font = hfont(size=14); c.fill = fill("7C3AED"); c.alignment = center
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:F2")
    s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
    s.font = dfont(color="94A3B8", size=10); s.fill = fill("1E293B"); s.alignment = center
    ws.row_dimensions[2].height = 22

    for ci, h in enumerate(["ລຳດັບ","ວັນທີ","ສິນຄ້າ","ຈຳນວນ","ຊຳລະ","ຍອດ (ກີບ)"], 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hfont(size=11); cell.fill = fill("1E3A5F"); cell.alignment = center; cell.border = border
    ws.row_dimensions[3].height = 28

    PAY_LAO = {'Cash': 'ເງິນສົດ', 'Transfer': 'ໂອນເງິນ'}
    for i, o in enumerate(sales):
        amt = float(o.amount or 0)
        r = i + 4; rf = fill("F5F3FF") if i % 2 == 0 else fill("FFFFFF")
        odate = o.order_date.strftime('%d/%m/%Y %H:%M') if o.order_date else '—'
        for ci, (v, a) in enumerate(zip(
            [i+1, odate, o.product.name if o.product else '—', o.quantity, PAY_LAO.get(o.payment_method, o.payment_method), dotfmt(amt)],
            [center, date_al, left, center, center, right]
        ), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = dfont(color="7C3AED" if ci == 6 else "374151", bold=(ci == 6))
            cell.fill = rf; cell.alignment = a; cell.border = border
        ws.row_dimensions[r].height = 22

    tr = sales.count() + 4
    for ci, (v, a) in enumerate(zip(
        ["", f"ລວມ {sales.count()} ລາຍການ", "", "", "", dotfmt(total)],
        [center, center, center, center, center, right]
    ), 1):
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.font = hfont(size=12)
        cell.fill = fill("7C3AED") if ci == 6 else fill("1E3A5F")
        cell.alignment = a; cell.border = border
    ws.row_dimensions[tr].height = 28
    for col, w in zip("ABCDEF", [8, 19, 26, 10, 12, 16]):
        ws.column_dimensions[col].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="EX_ຍອດຂາຍໜ້າຮ້ານ_{invoice_no}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='adminlogin')
def walkin_invoice_view(request):
    from django.utils import timezone as _tz

    sales, period_label, invoice_no = _walkin_sales_for_period(request)
    PAY_LAO = {'Cash': 'ເງິນສົດ', 'Transfer': 'ໂອນເງິນ'}
    items = []
    total = 0.0
    for o in sales:
        amt = float(o.amount or 0)
        total += amt
        items.append({
            'name': o.product.name if o.product else '—',
            'quantity': o.quantity,
            'subtotal': amt,
            'payment': PAY_LAO.get(o.payment_method, o.payment_method),
            'date': o.order_date,
        })

    context = {
        'items': items,
        'total': total,
        'period_label': period_label,
        'invoice_no': invoice_no,
        'generated_at': _tz.localtime(),
        'back_day': request.GET.get('date', ''),
    }
    return render(request, 'ecom/walkin_invoice.html', context)


@login_required(login_url='adminlogin')
def finance_month_daily_data(request):
    """Return daily revenue/profit for a given year+month as JSON."""
    import datetime as _dt
    import calendar as _cal
    from datetime import timezone as _dtz, timedelta as _td
    from django.http import JsonResponse

    _lao_tz = _dtz(_td(hours=7))
    now_lao = _dt.datetime.now(_lao_tz)

    try:
        year  = int(request.GET.get('year',  now_lao.year))
        month = int(request.GET.get('month', now_lao.month))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'invalid params'}, status=400)

    profit_percent = float(models.FinanceSettings.get_solo().profit_percent)
    days_in_month = _cal.monthrange(year, month)[1]

    daily_rev    = [0.0] * days_in_month
    daily_orders = [0]   * days_in_month

    month_start = _dt.datetime(year, month, 1,  0, 0, 0, tzinfo=_lao_tz)
    month_end   = _dt.datetime(year, month, days_in_month, 23, 59, 59, tzinfo=_lao_tz)

    for o in _revenue_orders_qs(month_start, month_end).values('order_date', 'amount', 'pickup_date', 'fulfilled_at'):
        revenue_dt = o['order_date']
        local_day = revenue_dt.astimezone(_lao_tz).day
        idx = local_day - 1
        if 0 <= idx < days_in_month:
            daily_rev[idx]    += float(o['amount'] or 0)
            daily_orders[idx] += 1

    daily_profit = [round(daily_rev[i] * profit_percent / 100, 2) for i in range(days_in_month)]

    return JsonResponse({
        'year':         year,
        'month':        month,
        'days':         days_in_month,
        'daily_rev':    daily_rev,
        'daily_profit': daily_profit,
        'daily_orders': daily_orders,
    })


# admin view customer table
@login_required(login_url='adminlogin')
def view_customer_view(request):
    customers=models.Customer.objects.all()
    return render(request,'ecom/view_customer.html',{'customers':customers})

# admin delete customer
@login_required(login_url='adminlogin')
def delete_customer_view(request,pk):
    customer=models.Customer.objects.get(id=pk)
    user=models.User.objects.get(id=customer.user_id)
    user.delete()
    customer.delete()
    return redirect('view-customer')


@login_required(login_url='adminlogin')
def update_customer_view(request,pk):
    customer=models.Customer.objects.get(id=pk)
    user=models.User.objects.get(id=customer.user_id)
    userForm=forms.CustomerUserForm(instance=user)
    customerForm=forms.CustomerForm(request.FILES,instance=customer)
    mydict={'userForm':userForm,'customerForm':customerForm}
    if request.method=='POST':
        userForm=forms.CustomerUserForm(request.POST,instance=user)
        customerForm=forms.CustomerForm(request.POST,instance=customer)
        if userForm.is_valid() and customerForm.is_valid():
            user=userForm.save()
            user.set_password(user.password)
            user.save()
            customerForm.save()
            return redirect('view-customer')
    return render(request,'ecom/admin_update_customer.html',context=mydict)

# admin view the product
@login_required(login_url='adminlogin')
def admin_products_view(request):
    categories = models.Category.objects.all()
    cat_id = request.GET.get('cat')
    sub_id = request.GET.get('sub')
    products = models.Product.objects.all()
    if cat_id:
        products = products.filter(category_id=cat_id)
    if sub_id:
        products = products.filter(subcategory_id=sub_id)
    return render(request, 'ecom/admin_products.html', {
        'products': products,
        'categories': categories,
        'selected_cat': int(cat_id) if cat_id else None,
        'selected_sub': int(sub_id) if sub_id else None,
    })


def _sync_product_colors(request, product):
    """Create/update/delete this product's ProductColor rows from the
    submitted color_id[]/color_name[]/color_stock[]/color_sold[] arrays.
    Rows with an empty color_id are new; existing rows whose id is not
    resubmitted are removed (matches whatever the admin left in the form)."""
    names  = request.POST.getlist('color_name[]')
    stocks = request.POST.getlist('color_stock[]')
    solds  = request.POST.getlist('color_sold[]')
    ids    = request.POST.getlist('color_id[]')
    while len(ids) < len(names):
        ids.append('')
    while len(solds) < len(names):
        solds.append('0')

    kept_ids = set()
    for idx, name in enumerate(names):
        name = (name or '').strip()
        if not name:
            continue
        try:
            stock = max(int(stocks[idx] or 0), 0)
        except (ValueError, IndexError):
            stock = 0
        try:
            sold = max(int(solds[idx] or 0), 0)
        except (ValueError, IndexError):
            sold = 0
        color_id = ids[idx].strip() if idx < len(ids) else ''
        pc = None
        if color_id:
            try:
                pc = models.ProductColor.objects.get(id=int(color_id), product=product)
                pc.color_name = name
                pc.stock_qty = stock
                pc.sold_qty = sold
                pc.save()
            except (models.ProductColor.DoesNotExist, ValueError):
                pc = None
        if pc is None:
            pc = models.ProductColor.objects.create(product=product, color_name=name, stock_qty=stock, sold_qty=sold)
        kept_ids.add(pc.id)

    product.colors.exclude(id__in=kept_ids).delete()


def _sync_product_gallery(request, product):
    """Add newly-uploaded gallery photos (up to 5 total including the primary
    product_image) and delete any the admin marked for removal."""
    remove_ids = request.POST.getlist('remove_image[]')
    if remove_ids:
        product.extra_images.filter(id__in=remove_ids).delete()
    new_files = request.FILES.getlist('extra_images')
    if new_files:
        room = max(5 - (1 + product.extra_images.count()), 0)
        for f in new_files[:room]:
            models.ProductImage.objects.create(product=product, image=f)


def _subcategories_by_category_json():
    """{category_id: [{id, name}, ...]} for the category→subcategory JS picker."""
    data = {}
    for sub in models.SubCategory.objects.select_related('category'):
        data.setdefault(str(sub.category_id), []).append({'id': sub.id, 'name': sub.name})
    return json.dumps(data)


# admin add product by clicking on floating button
@login_required(login_url='adminlogin')
def admin_add_product_view(request):
    productForm=forms.ProductForm()
    if request.method=='POST':
        productForm=forms.ProductForm(request.POST, request.FILES)
        if productForm.is_valid():
            product = productForm.save()
            _sync_product_colors(request, product)
            _sync_product_gallery(request, product)
        return HttpResponseRedirect('admin-products')
    return render(request,'ecom/admin_add_products.html',{
        'productForm': productForm,
        'subcategories_json': _subcategories_by_category_json(),
    })


@login_required(login_url='adminlogin')
def delete_product_view(request,pk):
    product=models.Product.objects.get(id=pk)
    product.delete()
    return redirect('admin-products')


@login_required(login_url='adminlogin')
def update_product_view(request,pk):
    product=models.Product.objects.get(id=pk)
    productForm=forms.ProductForm(instance=product)
    if request.method=='POST':
        productForm=forms.ProductForm(request.POST,request.FILES,instance=product)
        if productForm.is_valid():
            productForm.save()
            _sync_product_colors(request, product)
            _sync_product_gallery(request, product)
            return redirect('admin-products')
    return render(request,'ecom/admin_update_product.html',{
        'productForm': productForm,
        'colors': product.colors.all(),
        'gallery': product.extra_images.all(),
        'subcategories_json': _subcategories_by_category_json(),
    })


@login_required(login_url='adminlogin')
def toggle_product_stock_view(request, pk):
    try:
        product = models.Product.objects.get(id=pk)
    except models.Product.DoesNotExist:
        return JsonResponse({'ok': False}, status=404)
    product.is_available = not product.is_available
    product.save(update_fields=['is_available'])
    return JsonResponse({'ok': True, 'is_available': product.is_available})


@login_required(login_url='adminlogin')
def admin_advance_bookings_view(request):
    # Unified order-management page — every order (scheduled store-pickup
    # bookings, courier-delivery bookings, and admin-entered walk-in sales)
    # lives here now; there used to be a separate "ລາຍການຈອງຄິວ" page for
    # non-scheduled orders, merged into this one so admins have one place.
    orders = models.Orders.objects.all() \
        .select_related('product', 'customer__user').order_by('pickup_date', 'pickup_time', 'order_date')

    groups = {}
    group_order = []
    for order in orders:
        key = order.order_group or str(order.id)
        if key not in groups:
            groups[key] = {
                'key':          key,
                'canonical':    order,
                'orders':       [],
                'customer':     order.customer,
                'pickup_date':  order.pickup_date,
                'pickup_time':  order.pickup_time,
            }
            group_order.append(key)
        groups[key]['orders'].append(order)

    status_counts = {'All': 0, 'Pending': 0, 'Confirmed': 0, 'Processing': 0, 'Delivered': 0, 'Cancelled': 0}
    kind_counts = {'All': 0, 'pickup': 0, 'delivery': 0, 'walkin': 0}
    for key in group_order:
        g = groups[key]
        g['item_count']   = len(g['orders'])
        g['total_amount'] = sum(float(o.amount or 0) for o in g['orders'])
        g['order_ids']    = [o.id for o in g['orders']]
        non_cancelled = [o for o in g['orders'] if o.status != 'Cancelled']
        g['canonical']    = non_cancelled[0] if non_cancelled else g['orders'][0]
        g['is_walkin']    = g['canonical'].delivery_type == WALKIN_DELIVERY_TYPE
        g['deposit_verified'] = g['is_walkin'] or any(o.deposit_verified for o in g['orders'])
        if g['canonical'].pickup_date:
            g['kind'] = 'pickup'
        elif g['is_walkin']:
            g['kind'] = 'walkin'
        else:
            g['kind'] = 'delivery'
        s = g['canonical'].status or 'Pending'
        status_counts['All'] += 1
        if s in status_counts:
            status_counts[s] += 1
        kind_counts['All'] += 1
        kind_counts[g['kind']] += 1

    active_status = request.GET.get('status', 'All')
    if active_status not in ['Pending', 'Confirmed', 'Processing', 'Delivered', 'Cancelled', 'All']:
        active_status = 'All'

    kind_filter = request.GET.get('kind', 'All')
    if kind_filter not in ('All', 'pickup', 'delivery', 'walkin'):
        kind_filter = 'All'

    # Quick-link filters from the dashboard's all-time cards — "still owe
    # money" (unpaid) and "deposit confirmed by admin" (verified).
    unpaid_filter   = request.GET.get('unpaid')   == '1'
    verified_filter = request.GET.get('verified') == '1'

    data = [groups[key] for key in group_order]
    if active_status != 'All':
        data = [g for g in data if (g['canonical'].status or 'Pending') == active_status]
    if kind_filter != 'All':
        data = [g for g in data if g['kind'] == kind_filter]
    if unpaid_filter:
        data = [g for g in data if not g['is_walkin'] and (g['canonical'].status or 'Pending') in ('Pending', 'Confirmed', 'Processing')]
    if verified_filter:
        data = [g for g in data if g['deposit_verified']]

    return render(request, 'ecom/admin_advance_bookings.html', {
        'data': data,
        'active_status': active_status,
        'status_counts': status_counts,
        'kind_filter': kind_filter,
        'kind_counts': kind_counts,
        'unpaid_filter': unpaid_filter,
        'verified_filter': verified_filter,
    })


# ―― Shared by both the Excel export and the printable invoice below: rebuilds
# the same grouped/filtered advance-booking list admin_advance_bookings_view
# shows, so "ທັງໝົດ" (All) always includes every past booking too. Also honours
# an optional ?pickup_date=YYYY-MM-DD so admins can export/print just one
# specific pickup day instead of everything. ――
def _advance_bookings_for_export(request):
    orders = models.Orders.objects.all() \
        .select_related('product', 'customer__user').order_by('pickup_date', 'pickup_time', 'order_date')

    pickup_date_str = request.GET.get('pickup_date', '')
    pickup_date_obj = None
    if pickup_date_str:
        try:
            from datetime import date as _d
            pickup_date_obj = _d.fromisoformat(pickup_date_str)
            orders = orders.filter(pickup_date=pickup_date_obj)
        except ValueError:
            pickup_date_obj = None

    groups = {}
    group_order = []
    for order in orders:
        key = order.order_group or str(order.id)
        if key not in groups:
            groups[key] = {'key': key, 'canonical': order, 'orders': [], 'customer': order.customer,
                            'pickup_date': order.pickup_date, 'pickup_time': order.pickup_time}
            group_order.append(key)
        groups[key]['orders'].append(order)

    for key in group_order:
        g = groups[key]
        non_cancelled = [o for o in g['orders'] if o.status != 'Cancelled']
        g['canonical']    = non_cancelled[0] if non_cancelled else g['orders'][0]
        g['total_amount'] = sum(float(o.amount or 0) for o in g['orders'])
        g['item_names']   = ', '.join(o.product.name if o.product else (o.note or '—') for o in g['orders'])
        g['item_qty']     = sum(o.quantity for o in g['orders'])
        g['deposit_amount'] = round(g['total_amount'] * 0.20)
        g['is_walkin']    = g['canonical'].delivery_type == WALKIN_DELIVERY_TYPE
        if g['canonical'].pickup_date:
            g['kind'] = 'pickup'
        elif g['is_walkin']:
            g['kind'] = 'walkin'
        else:
            g['kind'] = 'delivery'

    active_status = request.GET.get('status', 'All')
    if active_status not in ['Pending', 'Confirmed', 'Processing', 'Delivered', 'Cancelled', 'All']:
        active_status = 'All'

    if active_status == 'All':
        data = [groups[key] for key in group_order]
    else:
        data = [groups[key] for key in group_order if (groups[key]['canonical'].status or 'Pending') == active_status]
    return data, active_status, pickup_date_obj


@login_required(login_url='adminlogin')
def export_advance_bookings_excel(request):
    from django.utils import timezone as _tz

    data, active_status, pickup_date_obj = _advance_bookings_for_export(request)
    today = _tz.localdate()

    LAO_FONT = "Phetsarath OT"
    def hfont(size=12): return Font(name=LAO_FONT, bold=True, color="FFFFFF", size=size)
    def dfont(bold=False, size=11, color="000000"): return Font(name=LAO_FONT, bold=bold, size=size, color=color)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def dotfmt(n):
        if n is None: return '—'
        return f"{int(round(float(n))):,}".replace(',', '.') + ' ກີບ'

    STATUS_LAO = {'Pending': 'ລໍຖ້າມາຮັບ', 'Confirmed': 'ຢືນຢັນແລ້ວ', 'Processing': 'ກຳລັງກຽມ', 'Delivered': 'ຮັບແລ້ວ', 'Cancelled': 'ຍົກເລີກ'}
    PAYMENT_LAO = {'COD': 'ຈ່າຍປາຍທາງ', 'InStore': 'ຈ່າຍໜ້າຮ້ານ', 'Cash': 'ເງິນສົດ', 'Transfer': 'ໂອນເງິນ'}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ຈອງລ່ວງໜ້າ"[:31]
    ws.sheet_view.showGridLines = False

    date_label = pickup_date_obj.strftime('%d/%m/%Y') if pickup_date_obj else STATUS_LAO.get(active_status, 'ທັງໝົດ')
    ws.merge_cells("A1:K1")
    c = ws["A1"]; c.value = f"ລາຍການຈອງລ່ວງໜ້າ — {date_label}"
    c.font = hfont(size=14); c.fill = fill("7C3AED"); c.alignment = center
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:K2")
    s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
    s.font = dfont(color="94A3B8", size=10); s.fill = fill("1E293B"); s.alignment = center
    ws.row_dimensions[2].height = 22

    headers = ["ລຳດັບ", "ວັນທີມາຮັບລົດ", "ໂມງມາຮັບ", "ລູກຄ້າ", "ເບີໂທ", "ຊື່ສິນຄ້າ",
               "ຈຳນວນລົດ", "ລາຄາລວມເຕັມຂອງລົດ", "ໂອນມັດຈຳເທົ່າໃດ", "ສະຖານະ", "ການຈ່າຍເງິນ"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = hfont(size=11); cell.fill = fill("1E3A5F"); cell.alignment = center; cell.border = border
    ws.row_dimensions[3].height = 28

    total = 0.0
    total_deposit = 0.0
    for i, g in enumerate(data):
        amt = g['total_amount']
        total += amt
        total_deposit += g['deposit_amount']
        r = i + 4; rf = fill("F5F3FF") if i % 2 == 0 else fill("FFFFFF")
        row_vals = [
            i + 1,
            g['pickup_date'].strftime('%d/%m/%Y') if g['pickup_date'] else '—',
            g['pickup_time'].strftime('%H:%M') if g['pickup_time'] else '—',
            g['customer'].get_name if g['customer'] else '—',
            g['customer'].mobile if g['customer'] else '—',
            g['item_names'],
            g['item_qty'],
            dotfmt(amt),
            dotfmt(g['deposit_amount']),
            STATUS_LAO.get(g['canonical'].status or 'Pending', g['canonical'].status),
            PAYMENT_LAO.get(g['canonical'].payment_method, g['canonical'].payment_method or '—'),
        ]
        aligns = [center, center, center, left, center, left, center, right, right, center, center]
        for ci, (v, a) in enumerate(zip(row_vals, aligns), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = dfont(color="7C3AED" if ci == 8 else "374151", bold=(ci in (8, 9)))
            cell.fill = rf; cell.alignment = a; cell.border = border
        ws.row_dimensions[r].height = 22

    tr = len(data) + 4
    tot_vals = ["", "", "", "", "", f"ລວມ {len(data)} ລາຍການ", "", dotfmt(total), dotfmt(total_deposit), "", ""]
    for ci, (v, a) in enumerate(zip(tot_vals, [center]*5 + [left] + [center] + [right]*2 + [center]*2), 1):
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.font = hfont(size=12)
        cell.fill = fill("7C3AED") if ci in (8, 9) else fill("1E3A5F")
        cell.alignment = a; cell.border = border
    ws.row_dimensions[tr].height = 28
    for col, w in zip("ABCDEFGHIJK", [8, 14, 10, 18, 15, 26, 10, 17, 16, 14, 14]):
        ws.column_dimensions[col].width = w

    fname_suffix = pickup_date_obj.strftime('%Y-%m-%d') if pickup_date_obj else 'ທັງໝົດ'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="EX_ຈອງລ່ວງໜ້າ_{fname_suffix}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='adminlogin')
def advance_bookings_invoice_view(request):
    from django.utils import timezone as _tz

    data, active_status, pickup_date_obj = _advance_bookings_for_export(request)
    total = sum(g['total_amount'] for g in data)
    STATUS_LAO = {'Pending': 'ລໍຖ້າມາຮັບ', 'Confirmed': 'ຢືນຢັນແລ້ວ', 'Processing': 'ກຳລັງກຽມ', 'Delivered': 'ຮັບແລ້ວ', 'Cancelled': 'ຍົກເລີກ', 'All': 'ທັງໝົດ'}
    status_label = pickup_date_obj.strftime('ວັນທີ %d/%m/%Y') if pickup_date_obj else STATUS_LAO.get(active_status, 'ທັງໝົດ')

    return render(request, 'ecom/advance_bookings_invoice.html', {
        'data': data,
        'total': total,
        'active_status': active_status,
        'status_label': status_label,
        'generated_at': _tz.localtime(),
    })


@login_required(login_url='adminlogin')
def delete_order_view(request,pk):
    order=models.Orders.objects.select_related('product').get(id=pk)
    if order.status != 'Cancelled' and order.color and order.product:
        _adjust_color_stock(order.product, order.color, -order.quantity)
    order.delete()
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('admin-advance-bookings')

# for changing status of order (pending,delivered...)
@login_required(login_url='adminlogin')
def update_order_view(request,pk):
    order=models.Orders.objects.get(id=pk)
    orderForm=forms.OrderForm(instance=order)
    if request.method=='POST':
        orderForm=forms.OrderForm(request.POST,instance=order)
        if orderForm.is_valid():
            orderForm.save()
            return redirect('admin-advance-bookings')
    return render(request,'ecom/update_order.html',{'orderForm':orderForm})


# admin view the feedback
@login_required(login_url='adminlogin')
def view_feedback_view(request):
    feedbacks = models.Feedback.objects.all().order_by('-id')
    return render(request, 'ecom/view_feedback.html', {'feedbacks': feedbacks})


# delete feedback
@login_required(login_url='adminlogin')
def delete_feedback_view(request, pk):
    models.Feedback.objects.filter(id=pk).delete()
    return redirect('view-feedback')


# AJAX: check new feedback
@login_required(login_url='adminlogin')
def ajax_check_new_feedback(request):
    try:
        last_id = int(request.GET.get('last_id', -1))
    except ValueError:
        last_id = -1
    if last_id == -1:
        latest = models.Feedback.objects.order_by('-id').first()
        return JsonResponse({'init': True, 'latest_id': latest.id if latest else 0})
    new_fb = models.Feedback.objects.filter(id__gt=last_id).order_by('-id')
    latest = new_fb.first()
    items = [{'id': f.id, 'name': f.name, 'preview': f.feedback[:70]} for f in new_fb[:4]]
    return JsonResponse({
        'new_count': new_fb.count(),
        'latest_id': latest.id if latest else last_id,
        'items': items,
    })


# ── Custom order requests (admin) ──
@login_required(login_url='adminlogin')
def admin_custom_orders_view(request):
    requests_qs = models.CustomOrderRequest.objects.all().select_related('customer', 'customer__user')
    return render(request, 'ecom/admin_custom_orders.html', {'requests': requests_qs})


@login_required(login_url='adminlogin')
def delete_custom_order_view(request, pk):
    models.CustomOrderRequest.objects.filter(id=pk).delete()
    return redirect('admin-custom-orders')


@login_required(login_url='adminlogin')
def done_custom_order_view(request, pk):
    try:
        req = models.CustomOrderRequest.objects.get(id=pk)
        req.is_done = not req.is_done
        req.save(update_fields=['is_done'])
    except models.CustomOrderRequest.DoesNotExist:
        pass
    return redirect('admin-custom-orders')


# AJAX: check new (undone) custom order requests
@login_required(login_url='adminlogin')
def ajax_check_new_custom_orders(request):
    try:
        last_id = int(request.GET.get('last_id', -1))
    except ValueError:
        last_id = -1
    if last_id == -1:
        latest = models.CustomOrderRequest.objects.order_by('-id').first()
        return JsonResponse({'init': True, 'latest_id': latest.id if latest else 0})
    new_reqs = models.CustomOrderRequest.objects.filter(id__gt=last_id).order_by('-id')
    latest = new_reqs.first()
    items = []
    for r in new_reqs[:4]:
        try:
            cname = r.customer.get_name if r.customer else 'ລູກຄ້າ'
        except Exception:
            cname = 'ລູກຄ້າ'
        items.append({'id': r.id, 'name': cname, 'preview': r.message[:70]})
    return JsonResponse({
        'new_count': new_reqs.count(),
        'latest_id': latest.id if latest else last_id,
        'items': items,
    })



#---------------------------------------------------------------------------------
#------------------------ PUBLIC CUSTOMER RELATED VIEWS START ---------------------
#---------------------------------------------------------------------------------


def search_view(request):
    # whatever user write in search box we get in query
    query = request.GET['query']
    products=models.Product.objects.all().filter(name__icontains=query).prefetch_related('colors', 'extra_images')
    if request.user.is_authenticated:
        cart = request.session.get('cart', {})
        product_count_in_cart = sum(cart.values())
    elif 'product_ids' in request.COOKIES:
        product_ids = request.COOKIES['product_ids']
        counter=product_ids.split('|')
        product_count_in_cart=len(set(counter))
    else:
        product_count_in_cart=0

    # word variable will be shown in html when user click on search button
    word="ຜົນການຄົ້ນຫາສີນຄ້າ :"

    categories = models.Category.objects.all()
    return render(request, 'ecom/index.html', {'products': products, 'word': word, 'categories': categories, 'product_count_in_cart': product_count_in_cart})




# 💡 1. ຟັງຊັນເພີ່ມຈຳນວນ (+)
def add_qty(request, pk):
    cart = request.session.get('cart', {})
    product_id = str(pk)
    
    if product_id in cart:
        cart[product_id] += 1
    
    request.session['cart'] = cart
    # 💡 ຕ້ອງມີບັນທັດນີ້ເພື່ອບັງຄັບໃຫ້ Save ລົງ Database/Session
    request.session.modified = True 
    
    return redirect(request.META.get('HTTP_REFERER', '/'))

# 💡 2. ຟັງຊັນລົບຈຳນວນ (-)
def remove_qty(request, pk):
    cart = request.session.get('cart', {})
    product_id = str(pk)
    
    if product_id in cart:
        if cart[product_id] > 1:
            cart[product_id] -= 1  # ຫຼຸດລົງ 1
        else:
            del cart[product_id]  # ຖ້າເຫຼືອ 1 ແລ້ວກົດລົບ ໃຫ້ລຶບອອກເລີຍ
            
    request.session['cart'] = cart
    request.session.modified = True
    return redirect(request.META.get('HTTP_REFERER', 'cart'))


# 💡 2. ຟັງຊັນລົບຈຳນວນ (-) ຈຳນວນຫຼາຍໆກວ່າ1ຂື້ນໄປ
def remove_qty_more(request, pk):
      # ດຶງຂໍ້ມູນ cart ຈາກ session
    cart = request.session.get('cart', {})
    product_id = str(pk)
    
    if product_id in cart:
        # ລຶບ product_id ອອກຈາກ dictionary ທັນທີ
        del cart[product_id]
        
        # ບັນທຶກການປ່ຽນແປງລົງໃນ session
        request.session['cart'] = cart
        request.session.modified = True
        
    return redirect(request.META.get('HTTP_REFERER', 'cart'))





# --- ຟັງຊັນຊ່ວຍຈັດການ Cookie ---
def get_response_with_cookie(request, product_id_list):
    response = HttpResponseRedirect(request.META.get('HTTP_REFERER', '/cart/'))
    if not product_id_list:
        response.delete_cookie('product_ids', path='/')
    else:
        new_value = "|".join(product_id_list)
        response.set_cookie('product_ids', new_value, path='/', max_age=3600*24*7)
    return response

# 1. ປຸ່ມບວກ (+) - ເພີ່ມຈຳນວນ
def add_qty_view(request, pk):
    product_ids = request.COOKIES.get('product_ids', "")
    product_id_list = [i for i in product_ids.split('|') if i != ""]
    product_id_list.append(str(pk))
    return get_response_with_cookie(request, product_id_list)

    

# 2. ປຸ່ມລົບ (-) - ລົດຈຳນວນລົງ 1
def remove_from_cart_view(request, pk):
    # 1. ດຶງຂໍ້ມູນ Cookie
    product_ids = request.COOKIES.get('product_ids', "")
    str_pk = str(pk)
    
    if product_ids:
        # 2. 💡 ໃຊ້ List Comprehension ເພື່ອ "ເອົາທຸກຕົວອອກ" ທີ່ເປັນ ID ນີ້
        # (ຖ້າມີ ID '5' ຢູ່ 10 ໂຕ ມັນຈະຖືກລຶບອອກທັງ 10 ໂຕເລີຍ)
        product_id_list = [i for i in product_ids.split('|') if i != str_pk and i != ""]
        
        # 3. ກຽມ Response (ກັບໄປໜ້າເດີມ)
        response = HttpResponseRedirect(request.META.get('HTTP_REFERER', '/cart/'))
        
        # 4. 💡 ສິ່ງສຳຄັນ: ຕ້ອງໃສ່ path='/' ທັງຕອນຕັ້ງ ແລະ ຕອນລຶບ
        if not product_id_list:
            response.delete_cookie('product_ids', path='/')
        else:
            new_value = "|".join(product_id_list)
            response.set_cookie('product_ids', new_value, path='/', max_age=3600*24*7)
            
        return response
    return redirect('cart')



# any one can add product to cart, no need of signin
def add_to_cart_view(request, pk):
    # 1. ດຶງຈຳນວນຈາກ URL (?quantity=5) ຖ້າບໍ່ມີໃຫ້ເປັນ 1
    quantity = int(request.GET.get('quantity', 1))

    # 2. ດຶງຂໍ້ມູນກະຕ່າຈາກ Session (Dictionary)
    cart = request.session.get('cart', {})

    # 3. ບວກຈຳນວນໃໝ່ໃສ່ຈຳນວນເກົ່າ — ສີທີ່ເລືອກ (ຖ້າມີ) ຝັງໄວ້ໃນ line_key
    # ("id-color") ເພື່ອໃຫ້ສີດຽວກັນລວມກັນ ແລະສີຕ່າງກັນແຍກເປັນຄົນລະແຖວໃນກະຕ່າ.
    color = request.GET.get('color', '').strip()
    product_id = f"{pk}-{color}" if color else str(pk)
    if product_id in cart:
        cart[product_id] += quantity
    else:
        cart[product_id] = quantity

    # 4. ບັນທຶກ ແລະ ນັບຈຳນວນທັງໝົດ
    request.session['cart'] = cart
    request.session.modified = True
    product_count_in_cart = sum(cart.values())

    return redirect(request.META.get('HTTP_REFERER', '/'))

    

def update_cart_qty(request, p_id, delta):
    cart = request.session.get('cart', {})
    p_id_str = str(p_id)
    
    if p_id_str in cart:
        cart[p_id_str] += int(delta)
        if cart[p_id_str] < 1:
            del cart[p_id_str]
            
    request.session['cart'] = cart
    request.session.modified = True

    # --- ຄິດໄລ່ລາຄາໃໝ່ທັງໝົດ ---
    grand_total = 0
    current_subtotal = 0
    for id, qty in cart.items():
        product = Product.objects.get(id=id)
        sub = product.price * qty
        grand_total += sub
        if id == p_id_str:
            current_subtotal = sub

    return JsonResponse({
        'status': 'success',
        'product_qty': cart.get(p_id_str, 0),
        'subtotal': "{:,}".format(current_subtotal),
        'grand_total': "{:,}".format(grand_total),
        'total_items': sum(cart.values())
    })

def _cart_lines(cart_dict):
    """Yield (line_key, product, qty, unit_price, subtotal) for every cart line."""
    if not cart_dict:
        return
    base_ids = {key.split('-')[0] for key in cart_dict.keys()}
    products_by_id = {str(p.id): p for p in models.Product.objects.filter(id__in=base_ids).select_related('category').prefetch_related('colors')}
    for line_key, qty in cart_dict.items():
        pid = line_key.split('-')[0]
        product = products_by_id.get(pid)
        if not product:
            continue
        unit_price = product.price
        subtotal = unit_price * qty
        yield line_key, product, qty, unit_price, subtotal


def _line_key_color(line_key):
    """Extracts the color name from a cart line_key of the form 'id-color',
    or None if this line has no color (product has no color variants)."""
    if '-' in line_key:
        return line_key.split('-', 1)[1] or None
    return None


def _adjust_color_stock(product, color_name, delta):
    """Moves a ProductColor's sold_qty by delta (positive = a booking just
    consumed `delta` units, negative = a cancelled booking gives them back).
    No-op if the product has no matching color row."""
    if not color_name:
        return
    try:
        pc = models.ProductColor.objects.get(product=product, color_name=color_name)
    except models.ProductColor.DoesNotExist:
        return
    pc.sold_qty = max(0, pc.sold_qty + delta)
    pc.save(update_fields=['sold_qty'])


# for checkout of cart

def cart_view(request):
    cart = request.session.get('cart', {})
    products_list = [] # 💡 ສ້າງ List ໃໝ່ເພື່ອເກັບຂໍ້ມູນຈາກ Session
    total = 0

    for line_key, product, item_qty, unit_price, subtotal in _cart_lines(cart):
        total += subtotal
        chosen_color = _line_key_color(line_key)
        if chosen_color:
            color_display = chosen_color
        else:
            color_display = ', '.join(product.colors.values_list('color_name', flat=True))
        products_list.append({
            'id':           product.id,
            'line_key':     line_key,
            'name':         product.name,
            'price':        product.price,
            'unit_price':   unit_price,
            'qty':          item_qty,
            'subtotal':     subtotal,
            'product_image': product.product_image,
            'category_name': product.category.name if product.category else '',
            'color_names':  color_display,
        })

    custom_cart_items = request.session.get('custom_cart_items', [])
    for c in custom_cart_items:
        c.setdefault('qty', 1)
        c['subtotal'] = c.get('unit_price', 0) * c['qty']
    total += sum(c['subtotal'] for c in custom_cart_items)

    # ຄິວຕໍ່ມື້ — ນັບສະເພາະ Pending orders ທີ່ຍັງຄ້າງຢູ່ + 1
    from datetime import date as _date, datetime as _dt
    from django.utils import timezone as _tz_cart
    _today = _date.today()
    _start = _tz_cart.make_aware(_dt(_today.year, _today.month, _today.day, 0, 0, 0))
    _end   = _tz_cart.make_aware(_dt(_today.year, _today.month, _today.day, 23, 59, 59))
    pending_count = models.Orders.objects.filter(
        order_date__gte=_start, order_date__lte=_end,
        status='Pending'
    ).values('order_group').distinct().count()
    queue_number = pending_count + 1

    customer_name = ''
    customer_mobile = ''
    if request.user.is_authenticated:
        try:
            _cust = models.Customer.objects.get(user=request.user)
            customer_name = _cust.get_name
            customer_mobile = _cust.mobile
        except models.Customer.DoesNotExist:
            pass

    return render(request, 'ecom/cart.html', {
        'products':          products_list,
        'total':             total,
        'deposit_amount':    round(total * 0.20),
        'queue_number':      queue_number,
        'customer_name':     customer_name,
        'customer_mobile':   customer_mobile,
        'custom_cart_items': custom_cart_items,
        'cart_total_items':  sum(cart.values()) + sum(c['qty'] for c in custom_cart_items),
        'closed_error':      _shop_closed_message() if not _is_shop_open() else None,
    })



def remove_from_cart_view(request, pk):
    # 1. ກວດເບິ່ງ Cookies ວ່າສິນຄ້າມີຢູ່ບໍ່
    if 'product_ids' in request.COOKIES:
        product_ids = request.COOKIES['product_ids']
        product_id_list = product_ids.split('|')
        
        # 2. Logic ການລຶບ: ລຶບ ID ທີ່ກົງກັບ pk ອອກພຽງ 1 ຕົວ (ເພື່ອໃຫ້ເຫຼືອຕົວຊ້ຳໄວ້)
        if str(pk) in product_id_list:
            product_id_list.remove(str(pk))
        
        # 3. ສ້າງສາຍ String ໃໝ່ເພື່ອເກັບລົງ Cookies
        value = "|".join(product_id_list)
        
        # 4. ຄິດໄລ່ສິນຄ້າທີ່ເຫຼືອເພື່ອສະແດງຜົນ
        products = models.Product.objects.filter(id__in=product_id_list)
        
        # ຄິດໄລ່ລາຄາລວມ (Total) ຕາມຈຳນວນທີ່ເຫຼືອແທ້
        total = 0
        for p in products:
            # ນັບວ່າ ID ນີ້ເຫຼືອຈັກອັນໃນ list ແລ້ວຄູນລາຄາ
            count = product_id_list.count(str(p.id))
            total += (p.price * count)
        
        # ນັບຈຳນວນສິນຄ້າທັງໝົດໃນກະຕ່າ (Counter)
        product_count_in_cart = len(product_id_list)
        
        # 5. ສ້າງ Response ແລະ ອັບເດດ Cookies ໃໝ່
        response = render(request, 'ecom/cart.html', {
            'products': products,
            'total': total,
            'product_count_in_cart': product_count_in_cart
        })
        
        if value == "":
            response.delete_cookie('product_ids')
        else:
            response.set_cookie('product_ids', value)
            
        return response
    
    return redirect('cart') # ຖ້າບໍ່ມີ Cookies ໃຫ້ກັບໄປໜ້າ Cart
    
    

def send_feedback_view(request):
    feedbackForm=forms.FeedbackForm()
    if request.method == 'POST':
        feedbackForm = forms.FeedbackForm(request.POST)
        if feedbackForm.is_valid():
            feedbackForm.save()
            return render(request, 'ecom/feedback_sent.html')
    return render(request, 'ecom/send_feedback.html', {'feedbackForm':feedbackForm})


@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def custom_order_request_view(request):
    """Add a custom ("ສັ່ງເມນູຕາມໃຈ") request to the session cart — it then goes
    through the exact same cart → address/checkout → payment-success flow as a
    normal product, and only becomes a real Orders row once checkout completes."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    message = request.POST.get('message', '').strip()
    if not message:
        return JsonResponse({'ok': False, 'error': 'ກະລຸນາພິມລາຍລະອຽດຄຳຂໍພິເສດ'})
    message = message[:300]

    # Price is quoted by admin afterward (via ajax_set_order_price) — there's
    # no formula to auto-price a one-off special request.
    custom_items = request.session.get('custom_cart_items', [])
    custom_items.append({'id': str(uuid.uuid4()), 'message': message, 'unit_price': 0, 'qty': 1})
    request.session['custom_cart_items'] = custom_items
    request.session.modified = True

    cart = request.session.get('cart', {})
    cart_total_items = sum(cart.values()) + sum(c.get('qty', 1) for c in custom_items)
    return JsonResponse({'ok': True, 'cart_total_items': cart_total_items, 'unit_price': 0})


@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def add_qty_custom_view(request, custom_id):
    custom_items = request.session.get('custom_cart_items', [])
    for c in custom_items:
        if c.get('id') == custom_id:
            c['qty'] = c.get('qty', 1) + 1
            break
    request.session['custom_cart_items'] = custom_items
    request.session.modified = True
    return redirect(request.META.get('HTTP_REFERER', 'cart'))


@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def remove_qty_custom_view(request, custom_id):
    custom_items = request.session.get('custom_cart_items', [])
    for c in custom_items:
        if c.get('id') == custom_id:
            if c.get('qty', 1) > 1:
                c['qty'] -= 1
            else:
                custom_items = [x for x in custom_items if x.get('id') != custom_id]
            break
    request.session['custom_cart_items'] = custom_items
    request.session.modified = True
    return redirect(request.META.get('HTTP_REFERER', 'cart'))


@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def remove_custom_cart_item_view(request, custom_id):
    custom_items = request.session.get('custom_cart_items', [])
    custom_items = [c for c in custom_items if c.get('id') != custom_id]
    request.session['custom_cart_items'] = custom_items
    request.session.modified = True
    return redirect('cart')


#---------------------------------------------------------------------------------
#------------------------ CUSTOMER RELATED VIEWS START ------------------------------
#---------------------------------------------------------------------------------
def customer_home_view(request):
    """The customer shop now lives entirely on the main site ('/') for both
    guests and logged-in customers — this route just bounces old links there."""
    cat_id = request.GET.get('cat')
    return redirect('/?cat=' + cat_id if cat_id else '/')



SHOP_CLOSED_MSG = 'ຂໍອະໄພ ຮ້ານຂອງເຮົາໄດ້ປິດແລ້ວ ເລີ່ມເປີດການຈອງຄິວໃໝ່ຂອງມື້ອື່ນໃນເວລາ 09:00 ຂໍຂອບໃຈ'


def _active_manual_closure():
    """The admin's manual 'ປິດຮ້ານ' override (from admin-announcements), if one is currently active."""
    return models.Announcement.objects.filter(kind='closed', is_active=True).order_by('-id').first()


def _is_shop_open():
    from django.utils import timezone as _tz
    from datetime import time as _time
    if _active_manual_closure():
        return False
    now_t = _tz.localtime().time()
    return _time(9, 0) <= now_t < _time(18, 0)


def _shop_closed_message():
    """The reason to show customers — the admin's custom message if manually closed, else the default hours message."""
    manual = _active_manual_closure()
    if manual and manual.message:
        return manual.message
    return SHOP_CLOSED_MSG


# ── Queue-wait warning: shop can comfortably work on BATCH_CAPACITY cups at once
# (~MIN_PER_CUP minutes each). Past that, new orders are still accepted (never blocked) —
# customers just get a heads-up estimate of how long they'll wait, based on the REAL
# number of cups currently active (Pending/Confirmed/Processing) placed TODAY.
# As soon as staff mark an order Delivered/Cancelled it drops out of this count, and the
# whole count resets fresh at midnight each day — it never carries over from a previous day.
BATCH_CAPACITY = 10
MIN_PER_CUP = 3


def _active_batch_qty():
    """Cups currently 'in the kitchen' — active status (not yet Delivered/Cancelled)
    AND placed today (real order_date, shop-local calendar day). Resets to 0 every
    new day; orders from a previous day never inflate today's count."""
    from django.db.models import Sum
    from django.utils import timezone as _tz
    from datetime import datetime as _dtt, timedelta as _td
    today_local = _tz.localdate()
    day_start = _tz.make_aware(_dtt(today_local.year, today_local.month, today_local.day))
    day_end   = day_start + _td(days=1)
    total = models.Orders.objects.filter(
        status__in=['Pending', 'Confirmed', 'Processing'],
        order_date__gte=day_start, order_date__lt=day_end,
    ).aggregate(s=Sum('quantity'))['s']
    return total or 0


def _queue_warning_message(active_qty):
    """Non-blocking heads-up shown when the queue is already at/over capacity. Returns None when it's fine to order with no extra wait."""
    if active_qty < BATCH_CAPACITY:
        return None
    est_low  = active_qty * MIN_PER_CUP
    est_high = est_low + 5
    return (
        f'ຂໍອະໄພ ຕອນນີ້ການຈອງມີປະມານ {active_qty} ຄິວ — '
        f'ຖ້າສັ່ງຕອນນີ້ຕ້ອງລໍອີກປະມານ {active_qty} ຄິວ ຫຼື ໃຊ້ເວລາປະມານ {est_low}-{est_high} ນາທີ. '
        f'ທ່ານຍັງກົດ "ຢືນຢັນການສັ່ງຊື້" ໄດ້ຕາມປົກກະຕິ'
    )


# shipment address before placing order
@login_required(login_url='customerlogin')
def customer_address_view(request):
    # Check session cart (new system uses session, not cookies)
    cart = request.session.get('cart', {})
    custom_cart_items = request.session.get('custom_cart_items', [])
    for c in custom_cart_items:
        c.setdefault('qty', 1)
    product_in_cart = len(cart) > 0 or len(custom_cart_items) > 0
    product_count_in_cart = sum(cart.values()) + sum(c['qty'] for c in custom_cart_items)

    from urllib.parse import unquote as _uq
    # Pre-fill from cookies (previous order) or customer profile
    try:
        customer = models.Customer.objects.get(user_id=request.user.id)
        prefill = {
            'Mobile':  _uq(request.COOKIES['mobile'])  if 'mobile'  in request.COOKIES else (customer.mobile  or ''),
            'Address': _uq(request.COOKIES['address']) if 'address' in request.COOKIES else (customer.address or ''),
        }
    except Exception:
        customer = None
        prefill = {
            'Mobile':  _uq(request.COOKIES.get('mobile',  '')),
            'Address': _uq(request.COOKIES.get('address', '')),
        }

    addressForm = forms.AddressForm(initial=prefill)
    closed_error = _shop_closed_message() if not _is_shop_open() else None

    if request.method == 'POST' and not closed_error:
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        print("=== CUSTOMER ADDRESS POST ===")
        print("POST data:", dict(request.POST))
        print("Cart:", cart)

        delivery_type  = request.POST.get('delivery_type',  'Delivery')
        payment_method = request.POST.get('payment_method', 'COD')
        delivery_km    = request.POST.get('delivery_km',    '0') or '0'
        delivery_fee   = request.POST.get('delivery_fee',   '0') or '0'

        # Pickup day+time only applies to "ຮັບໜ້າຮ້ານ" (store pickup) — falls
        # back to today if missing/invalid. Courier deliveries ship out
        # whenever ready, so they carry no scheduled pickup slot at all.
        pickup_date_str = ''
        pickup_time_str = ''
        from datetime import datetime as _dt_parse
        if delivery_type == 'Pickup':
            try:
                _pd = _dt_parse.strptime(request.POST.get('pickup_date', ''), '%Y-%m-%d').date()
                if _pd < timezone.localdate():
                    _pd = timezone.localdate()
                pickup_date_str = _pd.isoformat()
            except (ValueError, TypeError):
                pickup_date_str = timezone.localdate().isoformat()
            try:
                _pt = _dt_parse.strptime(request.POST.get('pickup_time', ''), '%H:%M').time()
                pickup_time_str = _pt.strftime('%H:%M')
            except (ValueError, TypeError):
                pickup_time_str = ''

        # Get mobile — fallback to customer profile
        mobile = request.POST.get('Mobile', '').strip()
        if not mobile and customer:
            mobile = customer.mobile or ''

        # Get address — for Pickup always use store name; for Delivery it's
        # composed from the chosen courier company + branch the customer
        # will collect the package from (no more door-to-door zone delivery).
        if delivery_type == 'Pickup':
            address = 'ຮັບໜ້າຮ້ານ'
        else:
            courier_name = request.POST.get('courier_name', '').strip()
            if courier_name == 'ອື່ນໆ':
                courier_name = request.POST.get('courier_other', '').strip() or 'ອື່ນໆ'
            courier_branch = request.POST.get('courier_branch', '').strip()
            address = f"ຂົນສົ່ງ: {courier_name} | ສາຂາ: {courier_branch}" if courier_name or courier_branch else ''
            if not address and customer:
                address = customer.address or ''

        email = (customer.user.email if customer else '') or ''

        from urllib.parse import quote as _q
        print(f"  → redirecting to payment-success | delivery={delivery_type} | mobile={mobile} | address={address}")

        response = HttpResponseRedirect(reverse('payment-success'))
        response.set_cookie('email',          _q(email))
        response.set_cookie('mobile',         _q(mobile))
        response.set_cookie('address',        _q(address))
        response.set_cookie('delivery_type',  _q(delivery_type))
        response.set_cookie('delivery_km',    _q(str(delivery_km)))
        response.set_cookie('delivery_fee',   _q(str(delivery_fee)))
        response.set_cookie('payment_method', _q(payment_method))
        response.set_cookie('pickup_date',    _q(pickup_date_str))
        response.set_cookie('pickup_time',    _q(pickup_time_str))
        return response

    # Build cart items for display
    cart_items = []
    subtotal = 0
    for line_key, p, qty, unit_price, item_total in _cart_lines(cart):
        subtotal += item_total
        cart_items.append({'product': p, 'line_key': line_key, 'qty': qty, 'total': item_total, 'unit_price': unit_price})

    for c in custom_cart_items:
        c['subtotal'] = c.get('unit_price', 0) * c['qty']
    subtotal += sum(c['subtotal'] for c in custom_cart_items)

    return render(request, 'ecom/customer_address.html', {
        'addressForm': addressForm,
        'product_in_cart': product_in_cart,
        'product_count_in_cart': product_count_in_cart,
        'prefill': prefill,
        'cart_items': cart_items,
        'custom_cart_items': custom_cart_items,
        'subtotal': subtotal,
        'deposit_amount': round(subtotal * 0.20),
        'remaining_amount': subtotal - round(subtotal * 0.20),
        'closed_error': closed_error,
        'queue_warning': _queue_warning_message(_active_batch_qty()) if not closed_error else None,
    })




# here we are just directing to this view...actually we have to check whther payment is successful or not
#then only this view should be accessed
@login_required(login_url='customerlogin')
def payment_success_view(request):
    from datetime import datetime as _dt, date as _date
    from urllib.parse import unquote as _uq
    print("=== PAYMENT SUCCESS VIEW ===")
    try:
        customer = models.Customer.objects.get(user_id=request.user.id)
        print(f"  customer: {customer}")
    except Exception as e:
        print(f"  ERROR getting customer: {e}")
        return redirect('customerlogin')

    if not _is_shop_open():
        return render(request, 'ecom/shop_closed.html', {
            'closed_error': _shop_closed_message(),
            'show_hours': not _active_manual_closure(),
        })

    cart = request.session.get('cart', {})
    custom_cart_items = request.session.get('custom_cart_items', [])
    print(f"  cart: {cart}")
    print(f"  custom_cart_items: {custom_cart_items}")

    group_id = None
    ordered_products = []
    grand_total = 0

    if cart or custom_cart_items:
        group_id = str(uuid.uuid4())

        try:
            d_km  = float(request.COOKIES.get('delivery_km', 0) or 0)
            d_fee = float(request.COOKIES.get('delivery_fee', 0) or 0)
        except ValueError:
            d_km, d_fee = 0, 0

        _order_email    = _uq(request.COOKIES.get('email', customer.user.email or ''))
        _order_mobile   = _uq(request.COOKIES.get('mobile', customer.mobile or ''))
        _order_address  = _uq(request.COOKIES.get('address', customer.address or ''))
        _order_delivery = _uq(request.COOKIES.get('delivery_type', 'Delivery'))
        _order_payment  = _uq(request.COOKIES.get('payment_method', 'COD'))
        _pickup_date_str = _uq(request.COOKIES.get('pickup_date', ''))
        _pickup_time_str = _uq(request.COOKIES.get('pickup_time', ''))
        try:
            _order_pickup_date = _dt.strptime(_pickup_date_str, '%Y-%m-%d').date() if _pickup_date_str else None
        except ValueError:
            _order_pickup_date = None
        try:
            _order_pickup_time = _dt.strptime(_pickup_time_str, '%H:%M').time() if _pickup_time_str else None
        except ValueError:
            _order_pickup_time = None

    if cart:
        for line_key, product, qty, unit_price, total_amount in _cart_lines(cart):
            grand_total += total_amount
            note_str = ''
            chosen_color = _line_key_color(line_key)

            ordered_products.append({'product': product, 'qty': qty, 'subtotal': total_amount, 'unit_price': unit_price, 'note': note_str})

            try:
                models.Orders.objects.create(
                    customer=customer,
                    product=product,
                    quantity=qty,
                    amount=total_amount,
                    status='Pending',
                    order_group=group_id,
                    email=_order_email,
                    mobile=_order_mobile,
                    address=_order_address,
                    delivery_type=_order_delivery,
                    delivery_km=d_km or None,
                    delivery_fee=d_fee or None,
                    payment_method=_order_payment,
                    note=note_str,
                    pickup_date=_order_pickup_date,
                    pickup_time=_order_pickup_time,
                    color=chosen_color,
                )
                # ຈອງແລ້ວ = ຫລົດສະຕອກສີນັ້ນທັນທີ (ບໍ່ລໍຖ້າຢືນຢັນ) — ຄືນຄືນເມື່ອອໍເດີ້ຖືກຍົກເລີກ.
                if chosen_color:
                    _adjust_color_stock(product, chosen_color, qty)
                print(f"  ✓ Created order for product {product.id}")
            except Exception as e:
                print(f"  ✗ ERROR creating order for product {product.id}: {e}")

    if custom_cart_items:
        for c in custom_cart_items:
            message    = c.get('message', '')[:300]
            unit_price = c.get('unit_price', 0)
            qty        = max(1, c.get('qty', 1))
            item_total = unit_price * qty
            grand_total += item_total
            note_str = message

            ordered_products.append({
                'product': None, 'qty': qty, 'subtotal': item_total, 'unit_price': unit_price,
                'note': note_str, 'is_custom': True, 'custom_message': message,
            })
            try:
                order = models.Orders.objects.create(
                    customer=customer,
                    product=None,
                    quantity=qty,
                    amount=item_total,
                    status='Pending',
                    order_group=group_id,
                    email=_order_email,
                    mobile=_order_mobile,
                    address=_order_address,
                    delivery_type=_order_delivery,
                    delivery_km=d_km or None,
                    delivery_fee=d_fee or None,
                    payment_method=_order_payment,
                    note=note_str,
                    pickup_date=_order_pickup_date,
                    pickup_time=_order_pickup_time,
                )
                models.CustomOrderRequest.objects.create(customer=customer, message=message, order_group=group_id)
                print(f"  ✓ Created custom order {order.id}")
            except Exception as e:
                print(f"  ✗ ERROR creating custom order: {e}")

    # ລາຄາລວມທັງໝົດ ຕ້ອງບວກຄ່າຈັດສົ່ງເຂົ້ານຳ
    subtotal = grand_total
    delivery_fee_amount = d_fee if (cart or custom_cart_items) and _order_delivery != 'Pickup' else 0
    grand_total = subtotal + delivery_fee_amount

    if cart or custom_cart_items:
        request.session['cart'] = {}
        request.session['custom_cart_items'] = []
        request.session.modified = True

    # ຄິວ = ຕຳແໜ່ງໃນ Pending orders ວັນນີ້ (ລວມ order ໃໝ່ທີ່ຫາກ່ຽວ place ໄປ)
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate()
        day_start = _tz.make_aware(_dt(today.year, today.month, today.day, 0, 0, 0))
        day_end   = _tz.make_aware(_dt(today.year, today.month, today.day, 23, 59, 59))
        queue_number = models.Orders.objects.filter(
            order_date__gte=day_start, order_date__lte=day_end,
            status='Pending'
        ).values('order_group').distinct().count()
    except Exception:
        queue_number = models.Orders.objects.filter(status='Pending').values('order_group').distinct().count()

    from urllib.parse import unquote as _uq
    delivery_type  = _uq(request.COOKIES.get('delivery_type', 'Delivery'))
    payment_method = _uq(request.COOKIES.get('payment_method', 'COD'))
    address        = _uq(request.COOKIES.get('address', ''))
    mobile         = _uq(request.COOKIES.get('mobile', ''))

    deposit_amount = round(float(grand_total) * 0.20)
    remaining_amount = float(grand_total) - deposit_amount

    return render(request, 'ecom/payment_success.html', {
        'queue_number':    queue_number,
        'grand_total':     grand_total,
        'subtotal':        subtotal,
        'delivery_fee_amount': delivery_fee_amount,
        'ordered_products': ordered_products,
        'delivery_type':   delivery_type,
        'payment_method':  payment_method,
        'address':         address,
        'mobile':          mobile,
        'deposit_amount':  deposit_amount,
        'remaining_amount': remaining_amount,
    })



# AJAX: fired from every customer page on load — pops up any unread order-status
# notifications immediately as a toast, no push opt-in required, then marks them
# read so they don't pop up again on the next page.
@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def ajax_customer_notifications(request):
    try:
        customer = models.Customer.objects.get(user_id=request.user.id)
    except models.Customer.DoesNotExist:
        return JsonResponse({'ok': False}, status=404)

    unread = list(customer.notifications.filter(is_read=False).order_by('-created_at')[:10])
    items = [{'title': n.title, 'body': n.body} for n in unread]
    if unread:
        models.CustomerNotification.objects.filter(id__in=[n.id for n in unread]).update(is_read=True)
    return JsonResponse({'ok': True, 'notifications': items})


@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def my_order_view(request):
    from datetime import datetime as _dt
    from collections import OrderedDict
    from django.utils import timezone as _tz
    from django.db.models import Min as _MinQ

    customer = models.Customer.objects.get(user_id=request.user.id)
    today = _tz.localdate()
    day_start = _tz.make_aware(_dt(today.year, today.month, today.day, 0, 0, 0))
    day_end   = _tz.make_aware(_dt(today.year, today.month, today.day, 23, 59, 59))

    # Today's orders only — new day = fresh start
    orders = models.Orders.objects.filter(
        customer_id=customer,
        order_date__gte=day_start,
        order_date__lte=day_end,
    ).order_by('id')

    # Queue map for today's pending orders — advance ("ຈອງລ່ວງໜ້າ") bookings
    # aren't being prepared today, so they don't take a spot in today's queue.
    _grp_pending = list(
        models.Orders.objects.filter(
            order_date__gte=day_start, order_date__lte=day_end,
            status='Pending', order_group__isnull=False, pickup_date__isnull=True,
        ).values('order_group').annotate(_fid=_MinQ('id')).order_by('_fid')
    )
    _sgl_pending = list(
        models.Orders.objects.filter(
            order_date__gte=day_start, order_date__lte=day_end,
            status='Pending', order_group__isnull=True, pickup_date__isnull=True,
        ).order_by('id').values_list('id', flat=True)
    )
    _slots = [(str(g['order_group']), g['_fid']) for g in _grp_pending]
    _slots += [(str(sid), sid) for sid in _sgl_pending]
    _slots.sort(key=lambda x: x[1])
    queue_map     = {key: i + 1 for i, (key, _) in enumerate(_slots)}
    total_pending = len(_slots)

    # Group by (local_date, order_group) — each date gets its own counter starting at 1
    date_groups = OrderedDict()   # {date: OrderedDict{key: group_data}}

    for order in orders:
        order_local_date = _tz.localdate(order.order_date) if order.order_date else today
        key = str(order.order_group) if order.order_group else str(order.id)
        is_today = (order_local_date == today)

        if order_local_date not in date_groups:
            date_groups[order_local_date] = OrderedDict()

        if key not in date_groups[order_local_date]:
            date_groups[order_local_date][key] = {
                'orders':        [],
                'items':         [],
                'first_order':   order,
                'queue':         queue_map.get(key, None) if is_today else None,
                'total_pending': total_pending if is_today else 0,
                'order_date':    order_local_date,
                'is_today':      is_today,
                'pickup_date':   order.pickup_date,
                'pickup_time':   order.pickup_time,
                'is_advance':    bool(order.pickup_date),
            }
        date_groups[order_local_date][key]['orders'].append(order)
        date_groups[order_local_date][key]['items'].append({
            'id':       order.id,
            'product':  order.product,
            'note':     order.note,
            'status':   order.status,
            'amount':   order.amount,
            'quantity': order.quantity,
        })

    # Flatten newest-date-first; compute per-day sequence numbers
    groups = []
    for date in sorted(date_groups.keys(), reverse=True):
        for seq_num, g in enumerate(date_groups[date].values(), start=1):
            non_cancelled = [o for o in g['orders'] if o.status != 'Cancelled']
            g['group_status'] = non_cancelled[0].status if non_cancelled else 'Cancelled'
            g['rep_order']    = non_cancelled[0] if non_cancelled else g['first_order']
            status_counts = {}
            for o in g['orders']:
                status_counts[o.status] = status_counts.get(o.status, 0) + 1
            g['status_counts'] = status_counts
            g['is_mixed']  = len(set(o.status for o in g['orders'])) > 1
            g['item_ids']  = ','.join(str(o.id) for o in g['orders'])
            g['daily_seq'] = seq_num   # resets to 1 every new day
            groups.append(g)

    # Keep today's regular queue and advance ("ຈອງລ່ວງໜ້າ") bookings in two
    # separate, non-interleaved blocks in the slide sequence — advance bookings
    # never share a queue position with today's orders, so they shouldn't be
    # interleaved with them here either. Advance block: soonest-pickup-first,
    # numbered on its own (independent of the regular block's daily_seq).
    regular_groups = [g for g in groups if not g['is_advance']]
    advance_groups = sorted(
        [g for g in groups if g['is_advance']],
        key=lambda g: (g['pickup_date'], g['pickup_time'] or _dt.min.time())
    )
    for adv_seq, g in enumerate(advance_groups, start=1):
        g['daily_seq'] = adv_seq
    groups = regular_groups + advance_groups

    # In-site notification inbox — show recent status updates, mark them read
    # once the customer has viewed this page.
    notifications = list(customer.notifications.all()[:20])
    unread_ids = [n.id for n in notifications if not n.is_read]
    if unread_ids:
        models.CustomerNotification.objects.filter(id__in=unread_ids).update(is_read=True)

    return render(request, 'ecom/my_order.html', {
        'groups': groups,
        'today': today,
        'notifications': notifications,
        'unread_notif_count': len(unread_ids),
    })




#--------------for discharge patient bill (pdf) download and printing
import io
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.template import Context
from django.http import HttpResponse


def render_to_pdf(template_src, context_dict):
    template = get_template(template_src)
    html  = template.render(context_dict)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return

@login_required(login_url='adminlogin')
def download_group_invoice_view(request, orderID):
    from django.db.models import Min as _Min
    from django.utils import timezone as _tz

    any_order = models.Orders.objects.get(id=orderID)
    group_id = any_order.order_group

    if group_id:
        orders = models.Orders.objects.filter(order_group=group_id).order_by('id')
    else:
        orders = models.Orders.objects.filter(id=orderID).order_by('id')

    canonical = orders.first()

    # Daily sequential invoice number: count distinct purchase events on the same
    # local date, so numbering resets to 1 each new day.
    # Use explicit UTC range instead of __date lookup to avoid SQLite timezone UDF errors.
    import datetime as _dt
    local_dt = _tz.localtime(canonical.order_date)
    local_date = local_dt.date()
    tz_obj = _tz.get_current_timezone()
    day_start = _tz.make_aware(_dt.datetime.combine(local_date, _dt.time.min), tz_obj)
    day_end   = _tz.make_aware(_dt.datetime.combine(local_date, _dt.time.max), tz_obj)

    grouped_before = (
        models.Orders.objects
        .filter(order_group__isnull=False, order_date__range=(day_start, day_end))
        .values('order_group')
        .annotate(_first=_Min('id'))
        .filter(_first__lte=canonical.id)
        .count()
    )
    single_before = models.Orders.objects.filter(
        order_group__isnull=True,
        order_date__range=(day_start, day_end),
        id__lte=canonical.id,
    ).count()
    invoice_number = grouped_before + single_before

    items = []
    subtotal_sum = 0
    for o in orders:
        if o.amount:
            subtotal = o.amount
        elif o.product:
            subtotal = o.product.price * (o.quantity or 1)
        else:
            subtotal = 0
        is_cancelled = (o.status == 'Cancelled')
        if not is_cancelled:
            subtotal_sum += subtotal
        items.append({
            'name': o.product.name if o.product else (o.note or 'ສັ່ງເມນູຕາມໃຈ'),
            'price': o.product.price if o.product else 0,
            'quantity': o.quantity or 1,
            'subtotal': subtotal,
            'cancelled': is_cancelled,
        })

    delivery_fee = canonical.delivery_fee or 0
    grand_total = subtotal_sum + delivery_fee
    has_cancelled = any(i['cancelled'] for i in items)

    # Deposit is always 20% of the grand total (same rate quoted at checkout
    # in cart.html) — shown here so the printed invoice states exactly how
    # much the customer already transferred vs. what's still owed at pickup.
    deposit_amount = round(float(grand_total) * 0.20)
    remaining_amount = float(grand_total) - deposit_amount
    deposit_verified = orders.filter(deposit_verified=True).exists()

    context = {
        'orderDate': _tz.localtime(canonical.order_date),
        'orderID': invoice_number,
        'customerName': canonical.customer.get_name if canonical.customer else request.user.username,
        'customerEmail': canonical.email,
        'customerMobile': canonical.mobile,
        'shipmentAddress': canonical.address,
        'orderStatus': canonical.status,
        'deliveryType': canonical.delivery_type,
        'deliveryFee': delivery_fee,
        'items': items,
        'subtotalSum': subtotal_sum,
        'grandTotal': grand_total,
        'hasCancelled': has_cancelled,
        'depositAmount': deposit_amount,
        'remainingAmount': remaining_amount,
        'depositVerified': deposit_verified,
    }
    return render(request, 'ecom/download_group_invoice.html', context)


@login_required(login_url='adminlogin')
def expense_invoice_view(request):
    import datetime as _dt
    from django.utils import timezone as _tz

    scope      = request.GET.get('scope', '')
    date_param = request.GET.get('date', '')
    month_param = request.GET.get('month')
    year = int(request.GET.get('year', _tz.localdate().year))

    if date_param:
        try:
            d = _dt.date.fromisoformat(date_param)
        except (ValueError, TypeError):
            d = _tz.localdate()
        expenses = models.Expense.objects.filter(date=d).order_by('id')
        period_label = f"ວັນທີ {d.strftime('%d/%m/%Y')}"
        invoice_no = d.strftime('%Y%m%d')
    elif scope == 'year':
        expenses = models.Expense.objects.filter(date__year=year).order_by('date', 'id')
        period_label = f"ປີ {year}"
        invoice_no = f"Y{year}"
    elif month_param:
        month = int(month_param)
        expenses = models.Expense.objects.filter(date__year=year, date__month=month).order_by('date', 'id')
        period_label = f"ເດືອນ {month:02d}/{year}"
        invoice_no = f"M{month:02d}{year}"
    else:
        today = _tz.localdate()
        expenses = models.Expense.objects.filter(date=today).order_by('id')
        period_label = f"ວັນທີ {today.strftime('%d/%m/%Y')}"
        invoice_no = today.strftime('%Y%m%d')

    items = []
    total = 0.0
    for e in expenses:
        amt = float(e.amount or 0)
        total += amt
        items.append({'date': e.date, 'category': e.category, 'description': e.description, 'amount': amt})

    context = {
        'items': items,
        'total': total,
        'period_label': period_label,
        'invoice_no': invoice_no,
        'generated_at': _tz.localtime(),
    }
    return render(request, 'ecom/expense_invoice.html', context)


def _resolve_invoice_period(request):
    """Shared date/month/year → (start_date, end_date, period_label, invoice_no) resolver."""
    import datetime as _dt
    from django.utils import timezone as _tz

    scope       = request.GET.get('scope', '')
    date_param  = request.GET.get('date', '')
    month_param = request.GET.get('month')
    year        = int(request.GET.get('year', _tz.localdate().year))

    if date_param:
        try:
            d = _dt.date.fromisoformat(date_param)
        except (ValueError, TypeError):
            d = _tz.localdate()
        return d, d, f"ວັນທີ {d.strftime('%d/%m/%Y')}", d.strftime('%Y%m%d')
    elif scope == 'year':
        return _dt.date(year, 1, 1), _dt.date(year, 12, 31), f"ປີ {year}", f"Y{year}"
    elif month_param:
        import calendar as _cal
        month = int(month_param)
        _, dim = _cal.monthrange(year, month)
        return _dt.date(year, month, 1), _dt.date(year, month, dim), f"ເດືອນ {month:02d}/{year}", f"M{month:02d}{year}"
    else:
        today = _tz.localdate()
        return today, today, f"ວັນທີ {today.strftime('%d/%m/%Y')}", today.strftime('%Y%m%d')


@login_required(login_url='adminlogin')
def revenue_invoice_view(request):
    from datetime import datetime as _dtt, timezone as _dtz, timedelta as _td
    from django.utils import timezone as _tz

    start_d, end_d, period_label, invoice_no = _resolve_invoice_period(request)
    _lao_tz = _dtz(_td(hours=7))
    r_start = _tz.make_aware(_dtt(start_d.year, start_d.month, start_d.day, 0, 0, 0))
    r_end   = _tz.make_aware(_dtt(end_d.year, end_d.month, end_d.day, 23, 59, 59))

    orders = models.Orders.objects.filter(
        order_date__gte=r_start, order_date__lte=r_end
    ).select_related('product').order_by('order_date')

    STATUS_LAO = {'Delivered':'ສຳເລັດ','Cancelled':'ຍົກເລີກ','Processing':'ກຳລັງຈັດ','Confirmed':'ຢືນຢັນ','Pending':'ລໍຖ້າ'}
    items = []
    total = 0.0
    for o in orders:
        amt = float(o.amount or 0)
        total += amt
        qty = o.quantity or 1
        items.append({
            'name': o.product.name if o.product else '—',
            'unit_price': amt / qty if qty else amt,
            'quantity': qty,
            'subtotal': amt,
            'status': STATUS_LAO.get(o.status, o.status),
            'date': o.order_date.astimezone(_lao_tz) if o.order_date else None,
        })

    context = {
        'items': items,
        'total': total,
        'period_label': period_label,
        'invoice_no': invoice_no,
        'generated_at': _tz.localtime(),
    }
    return render(request, 'ecom/revenue_invoice.html', context)


@login_required(login_url='adminlogin')
def profit_invoice_view(request):
    from datetime import datetime as _dtt, timezone as _dtz, timedelta as _td
    from django.utils import timezone as _tz

    start_d, end_d, period_label, invoice_no = _resolve_invoice_period(request)
    _lao_tz = _dtz(_td(hours=7))
    r_start = _tz.make_aware(_dtt(start_d.year, start_d.month, start_d.day, 0, 0, 0))
    r_end   = _tz.make_aware(_dtt(end_d.year, end_d.month, end_d.day, 23, 59, 59))

    orders = models.Orders.objects.filter(order_date__gte=r_start, order_date__lte=r_end).select_related('product').order_by('order_date')
    profit_percent = float(models.FinanceSettings.get_solo().profit_percent)

    STATUS_LAO = {'Delivered':'ສຳເລັດ','Cancelled':'ຍົກເລີກ','Processing':'ກຳລັງຈັດ','Confirmed':'ຢືນຢັນ','Pending':'ລໍຖ້າ'}
    revenue_items = []
    total_rev = 0.0
    for o in orders:
        amt = float(o.amount or 0)
        total_rev += amt
        revenue_items.append({
            'name': o.product.name if o.product else '—',
            'quantity': o.quantity or 1,
            'subtotal': amt,
            'status': STATUS_LAO.get(o.status, o.status),
            'date': o.order_date.astimezone(_lao_tz) if o.order_date else None,
        })

    context = {
        'revenue_items': revenue_items,
        'total_revenue': total_rev,
        'profit_percent': profit_percent,
        'total_profit': round(total_rev * profit_percent / 100, 2),
        'period_label': period_label,
        'invoice_no': invoice_no,
        'generated_at': _tz.localtime(),
    }
    return render(request, 'ecom/profit_invoice.html', context)


STATUS_NOTIFY_LAO = {
    'Pending':    ('🕐 ຮັບອໍເດີແລ້ວ', 'ຮ້ານໄດ້ຮັບອໍເດີຂອງທ່ານແລ້ວ ກຳລັງລໍຖ້າຢືນຢັນ'),
    'Confirmed':  ('✅ ຮ້ານຮັບອໍເດີແລ້ວ', 'ຮ້ານໄດ້ຮັບ ແລະ ຢືນຢັນອໍເດີຂອງທ່ານແລ້ວ ກະລຸນາລໍຖ້າ ກຳລັງກຽມສິນຄ້າ'),
    'Processing': ('👩‍🍳 ກຳລັງກຽມສິນຄ້າ', 'ຮ້ານກຳລັງກຽມອໍເດີຂອງທ່ານ ຈະແຈ້ງອີກເທື່ອໜຶ່ງເມື່ອພ້ອມ'),
    'Cancelled':  ('❌ ອໍເດີຖືກຍົກເລີກ', 'ອໍເດີຂອງທ່ານຖືກຍົກເລີກແລ້ວ ຫາກມີຂໍ້ສົງໄສ ກະລຸນາຕິດຕໍ່ຮ້ານ'),
}

# "Delivered" (ປຸ່ມ "ຈັດສົ່ງສຳເລັດ") ໃນລະບົບ ໝາຍເຖິງ "ພ້ອມໃຫ້ລູກຄ້າແລ້ວ" — ຂໍ້ຄວາມຈຶ່ງຕ້ອງ
# ແຕກຕ່າງກັນລະຫວ່າງ ຮັບໜ້າຮ້ານ (ມາຮັບໄດ້ເລີຍ) ແລະ ຈັດສົ່ງ (ໄລເດີ້ກຳລັງນຳສົ່ງ)
DELIVERED_NOTIFY_PICKUP   = ('🎉 ສິນຄ້າພ້ອມແລ້ວ', 'ອໍເດີຂອງທ່ານພ້ອມແລ້ວ ມາຮັບໄດ້ເລີຍທີ່ຮ້ານ 🏪')
DELIVERED_NOTIFY_DELIVERY = ('🚴 ໄລເດີ້ກຳລັງຈັດສົ່ງ', 'ອໍເດີຂອງທ່ານພ້ອມແລ້ວ ໄລເດີ້ກຳລັງນຳສົ່ງໃຫ້ທ່ານ')


def _notify_customer_status(customer, status, delivery_type=None):
    """Notify the customer of a status change two ways:
    - always create an in-site inbox entry (CustomerNotification) shown on
      /my-order — works for every customer, no opt-in needed
    - best-effort browser push, only reaches customers who enabled it
    Neither failure mode should ever break the actual status update."""
    if not customer:
        return
    if status == 'Delivered':
        title, body = DELIVERED_NOTIFY_PICKUP if delivery_type == 'Pickup' else DELIVERED_NOTIFY_DELIVERY
    elif status in STATUS_NOTIFY_LAO:
        title, body = STATUS_NOTIFY_LAO[status]
    else:
        return
    try:
        models.CustomerNotification.objects.create(customer=customer, title=title, body=body)
    except Exception:
        pass
    try:
        send_user_notification(user=customer.user, payload={'title': title, 'body': body, 'url': '/my-order'}, ttl=1000)
    except Exception:
        pass


@login_required(login_url='adminlogin')
@require_POST
def ajax_update_group_status(request):
    group_key = request.POST.get('group_key', '').strip()
    status = request.POST.get('status')
    if not group_key or status not in ['Pending', 'Confirmed', 'Processing', 'Delivered', 'Cancelled']:
        return JsonResponse({'ok': False}, status=400)
    group_orders = models.Orders.objects.filter(order_group=group_key)
    if not group_orders.exists():
        try:
            group_orders = models.Orders.objects.filter(id=int(group_key))
        except (ValueError, TypeError):
            return JsonResponse({'ok': False}, status=400)
    if not group_orders.exists():
        return JsonResponse({'ok': False}, status=400)

    customer, delivery_type = group_orders.select_related('customer__user').values_list('customer', 'delivery_type').first() or (None, None)

    # Online customer bookings must have their deposit manually verified
    # (bank amount + sender phone checked against the booking) before moving
    # past Pending — walk-in sales entered by staff in person skip this.
    if status in _DEPOSIT_GATE_STATUSES and delivery_type != WALKIN_DELIVERY_TYPE:
        if not group_orders.filter(deposit_verified=True).exists():
            return JsonResponse({'ok': False, 'error': 'deposit_not_verified'}, status=400)

    # Snapshot old statuses before the bulk update — needed to restore/re-deduct
    # per-color stock exactly on the Pending↔Cancelled transition edge.
    orders_snapshot = list(group_orders.values_list('id', 'status', 'product_id', 'color', 'quantity'))

    updated = group_orders.update(status=status)

    if status == 'Cancelled' or any(s == 'Cancelled' for _, s, _, _, _ in orders_snapshot):
        product_ids = {pid for _, _, pid, _, _ in orders_snapshot if pid}
        products_by_id = {p.id: p for p in models.Product.objects.filter(id__in=product_ids)}
        for oid, old_status, pid, color, qty in orders_snapshot:
            was_cancelled = (old_status == 'Cancelled')
            is_cancelled  = (status == 'Cancelled')
            if was_cancelled == is_cancelled or not color or not pid:
                continue
            product = products_by_id.get(pid)
            if not product:
                continue
            # Cancelling gives the color stock back; un-cancelling re-deducts it.
            _adjust_color_stock(product, color, -qty if is_cancelled else qty)

    # Advance ("ຈອງລ່ວງໜ້າ") bookings track when they were actually collected —
    # revenue for them counts on this day, not the (earlier) booking day.
    if status == 'Delivered':
        group_orders.filter(pickup_date__isnull=False).update(fulfilled_at=timezone.now())
    if customer:
        cust = models.Customer.objects.filter(id=customer).select_related('user').first()
        if cust:
            _notify_customer_status(cust, status, delivery_type=delivery_type)
    return JsonResponse({'ok': True, 'status': status, 'updated': updated})


@login_required(login_url='adminlogin')
@require_POST
def ajax_delete_group(request):
    group_key = request.POST.get('group_key', '').strip()
    if not group_key:
        return JsonResponse({'ok': False}, status=400)
    group_orders = models.Orders.objects.filter(order_group=group_key)
    if not group_orders.exists():
        try:
            group_orders = models.Orders.objects.filter(id=int(group_key))
        except (ValueError, TypeError):
            return JsonResponse({'ok': False}, status=400)

    for o in group_orders.select_related('product').exclude(status='Cancelled').exclude(color__isnull=True).exclude(color=''):
        if o.product:
            _adjust_color_stock(o.product, o.color, -o.quantity)

    deleted, _ = group_orders.delete()
    return JsonResponse({'ok': True, 'deleted': deleted})


@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def download_invoice_view(request,orderID,productID):
    order=models.Orders.objects.get(id=orderID)
    product=models.Product.objects.get(id=productID)
    mydict={
        'orderDate':order.order_date,
        'customerName':request.user,
        'customerEmail':order.email,
        'customerMobile':order.mobile,
        'shipmentAddress':order.address,
        'orderStatus':order.status,

        'productName':product.name,
        'productImage':product.product_image,
        'productPrice':product.price,
        'productDescription':product.description,


    }
    return render_to_pdf('ecom/download_invoice.html',mydict)






@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def my_profile_view(request):
    customer=models.Customer.objects.get(user_id=request.user.id)
    return render(request,'ecom/my_profile.html',{'customer':customer})


@login_required(login_url='customerlogin')
@user_passes_test(is_customer)
def edit_profile_view(request):
    customer=models.Customer.objects.get(user_id=request.user.id)
    user=models.User.objects.get(id=customer.user_id)
    userForm=forms.CustomerUserForm(instance=user)
    customerForm=forms.CustomerForm(request.FILES,instance=customer)
    mydict={'userForm':userForm,'customerForm':customerForm}
    if request.method=='POST':
        userForm=forms.CustomerUserForm(request.POST,instance=user)
        customerForm=forms.CustomerForm(request.POST,instance=customer)
        if userForm.is_valid() and customerForm.is_valid():
            user=userForm.save()
            user.set_password(user.password)
            user.save()
            customerForm.save()
            return HttpResponseRedirect('my-profile')
    return render(request,'ecom/edit_profile.html',context=mydict)



#---------------------------------------------------------------------------------
#------------------------ ABOUT US AND CONTACT US VIEWS START --------------------
#---------------------------------------------------------------------------------
def aboutus_view(request):
    return render(request, 'ecom/aboutus.html', {})


def queue_view(request):
    return render(request, 'ecom/queue.html', {})

from django.shortcuts import render, redirect
from . import forms
import urllib.parse

from django.shortcuts import render, redirect
from . import forms, models

def contactus_view(request):
    sub = forms.ContactusForm()
    
    if request.method == 'POST':
        sub = forms.ContactusForm(request.POST)
        
        if sub.is_valid():
            # --- 1. ບັນທຶກລົງ Database (ແກ້ Error 'save') ---
            if hasattr(sub, 'save'):
                # ໃຊ້ວິທີນີ້ຖ້າທ່ານປ່ຽນເປັນ ModelForm ແລ້ວ
                sub.save()
            else:
                # ໃຊ້ວິທີນີ້ຖ້າທ່ານຍັງໃຊ້ forms.Form (ແບບເກົ່າ)
                name = sub.cleaned_data.get('Name')
                message = sub.cleaned_data.get('Message')
                models.Feedback.objects.create(name=name, feedback=message)
            
            # --- 2. ປ່ຽນລິ້ງ: ສົ່ງລູກຄ້າໄປຫາໜ້າສຳເລັດ (Success Page) ---
            # ໃຫ້ແນ່ໃຈວ່າທ່ານມີໄຟລ໌ ecom/contactussuccess.html ຢູ່ໃນ templates
            return render(request, 'ecom/contactussuccess.html')
            
    return render(request, 'ecom/contactus.html', {'form': sub})

def update_order_status(request, id):
    order = models.Orders.objects.get(id=id)
    form = forms.OrderForm(instance=order)

    if request.method == 'POST':
        form = forms.OrderForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('admin-advance-bookings')

    return render(request, 'ecom/update_status.html', {'orderForm': form})


# ---- AJAX: Update Order Status ----
@login_required(login_url='adminlogin')
@require_POST
def ajax_update_order_status(request, pk):
    order = models.Orders.objects.select_related('customer__user', 'product').get(id=pk)
    status = request.POST.get('status')
    if status in ['Pending', 'Confirmed', 'Processing', 'Delivered', 'Cancelled']:
        if (status in _DEPOSIT_GATE_STATUSES and order.delivery_type != WALKIN_DELIVERY_TYPE
                and not order.deposit_verified):
            return JsonResponse({'ok': False, 'error': 'deposit_not_verified'}, status=400)
        was_cancelled = (order.status == 'Cancelled')
        is_cancelled  = (status == 'Cancelled')
        order.status = status
        if status == 'Delivered' and order.pickup_date:
            order.fulfilled_at = timezone.now()
        order.save()
        if was_cancelled != is_cancelled and order.color and order.product:
            _adjust_color_stock(order.product, order.color, -order.quantity if is_cancelled else order.quantity)
        if order.customer:
            _notify_customer_status(order.customer, status, delivery_type=order.delivery_type)
        return JsonResponse({'ok': True, 'status': status})
    return JsonResponse({'ok': False}, status=400)


# ---- AJAX: admin manually confirms the deposit slip matches this booking
# (bank amount + sender phone checked by hand) — required before an online
# booking's status can move past Pending ----
@login_required(login_url='adminlogin')
@require_POST
def ajax_verify_deposit(request):
    group_key = request.POST.get('group_key', '').strip()
    if not group_key:
        return JsonResponse({'ok': False}, status=400)
    group_orders = models.Orders.objects.filter(order_group=group_key)
    if not group_orders.exists():
        try:
            group_orders = models.Orders.objects.filter(id=int(group_key))
        except (ValueError, TypeError):
            return JsonResponse({'ok': False}, status=400)
    updated = group_orders.update(deposit_verified=True)
    if updated == 0:
        return JsonResponse({'ok': False}, status=400)
    return JsonResponse({'ok': True})


# ---- AJAX: admin quotes a price for a custom ("ສັ່ງເມນູຕາມໃຈ") order ----
@login_required(login_url='adminlogin')
def ajax_set_order_price(request, pk):
    try:
        order = models.Orders.objects.get(id=pk)
    except models.Orders.DoesNotExist:
        return JsonResponse({'ok': False}, status=404)
    try:
        amount = int(request.POST.get('amount', ''))
    except (ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'ລາຄາບໍ່ຖືກຕ້ອງ'}, status=400)
    if amount < 0:
        return JsonResponse({'ok': False, 'error': 'ລາຄາບໍ່ຖືກຕ້ອງ'}, status=400)
    order.amount = amount
    order.save(update_fields=['amount'])
    return JsonResponse({'ok': True, 'amount': amount})


# ---- AJAX: Per-item statuses for customer tracking ----
def ajax_item_statuses(request):
    ids_raw = request.GET.get('ids', '')
    try:
        id_list = [int(x) for x in ids_raw.split(',') if x.strip()]
    except ValueError:
        return JsonResponse({'ok': False}, status=400)
    if not id_list:
        return JsonResponse({'ok': False}, status=400)
    rows = models.Orders.objects.filter(id__in=id_list).values('id', 'status')
    return JsonResponse({'ok': True, 'statuses': {str(r['id']): r['status'] for r in rows}})


# ---- AJAX: Live queue position for customer ----
def ajax_queue_position(request):
    from datetime import datetime as _dqt, date as _dqd
    from django.utils import timezone as _tzq
    group = request.GET.get('group', '')
    if not group:
        return JsonResponse({'ok': False}, status=400)
    _today = _dqd.today()
    d_start = _tzq.make_aware(_dqt(_today.year, _today.month, _today.day, 0, 0, 0))
    d_end   = _tzq.make_aware(_dqt(_today.year, _today.month, _today.day, 23, 59, 59))
    from django.db.models import Min as _MinAQ
    _ag = list(
        models.Orders.objects.filter(
            order_date__gte=d_start, order_date__lte=d_end,
            status='Pending', order_group__isnull=False,
        ).values('order_group').annotate(_fid=_MinAQ('id')).order_by('_fid')
    )
    _as = list(
        models.Orders.objects.filter(
            order_date__gte=d_start, order_date__lte=d_end,
            status='Pending', order_group__isnull=True,
        ).order_by('id').values_list('id', flat=True)
    )
    _aslots = [(str(g['order_group']), g['_fid']) for g in _ag]
    _aslots += [(str(sid), sid) for sid in _as]
    _aslots.sort(key=lambda x: x[1])
    _akeys = [k for k, _ in _aslots]

    if group in _akeys:
        pos   = _akeys.index(group) + 1
        ahead = pos - 1
    else:
        pos   = None
        ahead = 0
    order_status = models.Orders.objects.filter(order_group=group).values_list('status', flat=True).first()
    return JsonResponse({
        'ok': True,
        'pos': pos,
        'ahead': ahead,
        'total_pending': len(_aslots),
        'status': order_status or 'Unknown',
    })


# ---- AJAX: Poll new orders ----
@login_required(login_url='adminlogin')
def ajax_check_new_orders(request):
    try:
        last_id = int(request.GET.get('last_id', 0))
    except ValueError:
        last_id = 0
    new_orders = models.Orders.objects.filter(id__gt=last_id).order_by('-id')
    latest = new_orders.first()
    items = []
    for o in new_orders[:5]:
        try:
            pname = o.product.name if o.product else '-'
        except Exception:
            pname = '-'
        try:
            cname = o.customer.get_name if o.customer else '-'
        except Exception:
            cname = '-'
        items.append({'id': o.id, 'product': pname, 'customer': cname, 'amount': str(o.amount or 0), 'is_advance': bool(o.pickup_date)})
    return JsonResponse({
        'new_count': new_orders.count(),
        'latest_id': latest.id if latest else last_id,
        'orders': items,
    })


# ---- AJAX: Day order detail (for dashboard chart click) ----
@login_required(login_url='adminlogin')
def ajax_day_orders(request):
    from datetime import date as _dc, datetime as _dtt, timezone as _dtz, timedelta as _td
    from django.utils import timezone as _tz
    date_str    = request.GET.get('date', '')
    status_filt = request.GET.get('status', '')
    try:
        sel_date = _dc.fromisoformat(date_str)
    except (ValueError, TypeError):
        sel_date = _tz.localdate()
    d_start = _tz.make_aware(_dtt(sel_date.year, sel_date.month, sel_date.day, 0, 0, 0))
    d_end   = _tz.make_aware(_dtt(sel_date.year, sel_date.month, sel_date.day, 23, 59, 59))
    # ໃຊ້ set ດຽວກັນກັບ _revenue_orders_qs ເພື່ອໃຫ້ຕົງກັບຍອດຂາຍທີ່ສະແດງຢູ່ card
    # (ຂາຍໜ້າຮ້ານ + ຈອງອອນລາຍ + ຈອງລ່ວງໜ້າທີ່ຮັບແລ້ວມື້ນີ້)
    qs = _revenue_orders_qs(d_start, d_end).select_related('product', 'customer__user').order_by('id')
    if status_filt and status_filt not in ('', 'All'):
        qs = qs.filter(status=status_filt)
    _lao_tz = _dtz(_td(hours=7))
    orders_list = []
    try:
        for o in qs:
            try:
                img_url = o.product.product_image.url if (o.product and o.product.product_image) else None
            except Exception:
                img_url = None
            _time_src = o.order_date
            orders_list.append({
                'id':          o.id,
                'product':     o.product.name if o.product else '—',
                'product_img': img_url,
                'customer':    o.customer.get_name if o.customer else '—',
                'mobile':      o.customer.mobile if o.customer else (o.mobile or '—'),
                'amount':      float(o.amount or 0),
                'quantity':    o.quantity or 1,
                'status':      o.status or 'Pending',
                'time':        _time_src.astimezone(_lao_tz).strftime('%H:%M') if _time_src else '--:--',
                'address':     o.address or '',
            })
    except Exception as _ex:
        return JsonResponse({'ok': False, 'error': str(_ex)})
    total = sum(x['amount'] for x in orders_list)
    return JsonResponse({'ok': True, 'date': sel_date.isoformat(), 'orders': orders_list, 'total': total, 'count': len(orders_list)})


# ---- AJAX: Get Order Detail ----
@login_required(login_url='adminlogin')
def ajax_order_detail(request, pk):
    try:
        order = models.Orders.objects.get(id=pk)
        customer = models.Customer.objects.filter(id=order.customer.id).first()
        product = models.Product.objects.filter(id=order.product.id).first()
        return JsonResponse({
            'ok': True,
            'id': order.id,
            'customer': customer.get_name if customer else '-',
            'product': product.name if product else '-',
            'price': str(order.amount or (product.price if product else 0)),
            'status': order.status or '',
            'address': order.address or '-',
            'mobile': order.mobile or '-',
            'date': str(order.order_date) if order.order_date else '-',
        })
    except models.Orders.DoesNotExist:
        return JsonResponse({'ok': False}, status=404)


# ---- Category Management ----
@login_required(login_url='adminlogin')
def admin_categories_view(request):
    if request.method == 'POST':
        if request.POST.get('form_type') == 'subcategory':
            sub_name = request.POST.get('sub_name', '').strip()
            cat_id = request.POST.get('category_id')
            if sub_name and cat_id:
                try:
                    parent = models.Category.objects.get(id=cat_id)
                    models.SubCategory.objects.create(category=parent, name=sub_name)
                except models.Category.DoesNotExist:
                    pass
        else:
            name = request.POST.get('name', '').strip()
            if name:
                models.Category.objects.create(name=name)
    categories = models.Category.objects.annotate(
        product_count=Count('products', distinct=True),
        color_count=Count('products__colors', distinct=True),
    ).prefetch_related('subcategories')
    return render(request, 'ecom/admin_categories.html', {'categories': categories})

@login_required(login_url='adminlogin')
def delete_subcategory_view(request, pk):
    models.SubCategory.objects.filter(id=pk).delete()
    return redirect('admin-categories')

@login_required(login_url='adminlogin')
def delete_category_view(request, pk):
    models.Category.objects.filter(id=pk).delete()
    return redirect('admin-categories')


# ---- Announcements: manual "ปิดร้าน" override + "ໂປໂມຊັ່ນ" banners shown on the customer home page ----
@login_required(login_url='adminlogin')
def admin_announcements_view(request):
    if request.method == 'POST':
        kind    = request.POST.get('kind', 'promo')
        title   = request.POST.get('title', '').strip()
        message = request.POST.get('message', '').strip()
        icon    = request.POST.get('icon', '').strip() or ('🚫' if kind == 'closed' else '📢')
        if title:
            if kind == 'closed':
                # Only one manual closure can be active at a time.
                models.Announcement.objects.filter(kind='closed', is_active=True).update(is_active=False)
            models.Announcement.objects.create(kind=kind, title=title, message=message, icon=icon)
        return redirect('admin-announcements')

    closures = models.Announcement.objects.filter(kind='closed')
    promos   = models.Announcement.objects.filter(kind='promo')
    return render(request, 'ecom/admin_announcements.html', {
        'closures': closures,
        'promos': promos,
        'active_closure': closures.filter(is_active=True).order_by('-id').first(),
    })


@login_required(login_url='adminlogin')
def toggle_announcement_view(request, pk):
    try:
        ann = models.Announcement.objects.get(id=pk)
    except models.Announcement.DoesNotExist:
        return redirect('admin-announcements')
    if not ann.is_active and ann.kind == 'closed':
        # Turning a closure ON — switch off any other active closure first.
        models.Announcement.objects.filter(kind='closed', is_active=True).update(is_active=False)
    ann.is_active = not ann.is_active
    ann.save(update_fields=['is_active'])
    return redirect('admin-announcements')


@login_required(login_url='adminlogin')
def delete_announcement_view(request, pk):
    models.Announcement.objects.filter(id=pk).delete()
    return redirect('admin-announcements')


# ---- Export Products Excel ----
@login_required(login_url='adminlogin')
def export_products_excel(request):
    products = models.Product.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ຄັງສິນຄ້າ"

    _LAO_FONT = "Phetsarath OT"
    header_font = Font(name=_LAO_FONT, bold=True, color="FFFFFF", size=12)
    data_font   = Font(name=_LAO_FONT, size=11)
    header_fill = PatternFill("solid", fgColor="1d6f42")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ລຳດັບ", "ລະຫັດ", "ຊື່ສິນຄ້າ", "ລາຍລະອຽດ", "ລາຄາ (ກີບ)"]
    col_widths = [8, 12, 30, 40, 18]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[1].height = 25

    for row_num, p in enumerate(products, 2):
        row_fill = PatternFill("solid", fgColor="F0FFF4") if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [row_num - 1, f"#PRO-{p.id}", p.name, p.description, p.price]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[row_num].height = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
    wb.save(response)
    return response


# ---- Export Customers Excel ----
@login_required(login_url='adminlogin')
def export_customers_excel(request):
    customers = models.Customer.objects.all().select_related('user')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ຂໍ້ມູນລູກຄ້າ"

    _LAO_FONT = "Phetsarath OT"
    header_font = Font(name=_LAO_FONT, bold=True, color="FFFFFF", size=12)
    data_font   = Font(name=_LAO_FONT, size=11)
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ລຳດັບ", "ລະຫັດ", "ຊື່ລູກຄ້າ", "ເບີໂທ", "ທີ່ຢູ່"]
    col_widths = [8, 12, 25, 18, 35]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[1].height = 25

    for row_num, c in enumerate(customers, 2):
        row_fill = PatternFill("solid", fgColor="F5F8FF") if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [row_num - 1, f"#CUS-{c.id}", c.get_name, c.mobile or "-", c.address or "-"]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[row_num].height = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
    wb.save(response)
    return response


# ---- Export Orders Excel ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

@login_required(login_url='adminlogin')
def export_orders_excel(request):
    from django.db.models import Sum as _Sum
    from datetime import timezone as _dtz, timedelta as _td

    _LAO_TZ = _dtz(_td(hours=7))  # UTC+7 (ລາວ)

    STATUS_LAO = {
        'Pending':    'ຮັບອໍເດີ',
        'Processing': 'ກຳລັງຈັດສົ່ງ',
        'Delivered':  'ຈັດສົ່ງສຳເລັດ',
        'Cancelled':  'ຍົກເລີກ',
        'Completed':  'ສຳເລັດ',
    }

    orders = models.Orders.objects.all().select_related('product', 'customer__user').order_by('id')

    # Optional: filter by date if ?date= provided
    _date_str = request.GET.get('date', '')
    if _date_str:
        try:
            from datetime import date as _d_cls, datetime as _dt_cls
            from django.utils import timezone as _tz_ex
            _fd = _d_cls.fromisoformat(_date_str)
            _ds = _tz_ex.make_aware(_dt_cls(_fd.year, _fd.month, _fd.day, 0, 0, 0))
            _de = _tz_ex.make_aware(_dt_cls(_fd.year, _fd.month, _fd.day, 23, 59, 59))
            orders = orders.filter(order_date__gte=_ds, order_date__lte=_de)
        except ValueError:
            pass

    # Optional: filter by month if ?month=YYYY-MM provided
    _month_str = request.GET.get('month', '')
    if _month_str:
        try:
            import calendar as _cal
            from datetime import datetime as _dt_cls2
            from django.utils import timezone as _tz_mo
            _yr, _mo = map(int, _month_str.split('-'))
            _, _last = _cal.monthrange(_yr, _mo)
            _ms = _tz_mo.make_aware(_dt_cls2(_yr, _mo, 1, 0, 0, 0))
            _me = _tz_mo.make_aware(_dt_cls2(_yr, _mo, _last, 23, 59, 59))
            orders = orders.filter(order_date__gte=_ms, order_date__lte=_me)
        except (ValueError, TypeError):
            pass

    # Optional: filter by year if ?year=YYYY provided
    _year_str = request.GET.get('year', '')
    if _year_str:
        try:
            from datetime import datetime as _dt_cls3
            from django.utils import timezone as _tz_yr
            _yr_int = int(_year_str)
            _ys = _tz_yr.make_aware(_dt_cls3(_yr_int, 1, 1, 0, 0, 0))
            _ye = _tz_yr.make_aware(_dt_cls3(_yr_int, 12, 31, 23, 59, 59))
            orders = orders.filter(order_date__gte=_ys, order_date__lte=_ye)
        except (ValueError, TypeError):
            pass

    # Optional: filter by status if ?status= provided
    _status_str = request.GET.get('status', '')
    if _status_str and _status_str not in ('All', ''):
        orders = orders.filter(status=_status_str)

    # Pre-compute totals before iterating
    agg = orders.aggregate(total_qty=_Sum('quantity'), total_amount=_Sum('amount'))
    total_qty    = int(agg['total_qty']    or 0)
    total_amount = int(agg['total_amount'] or 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ລາຍການສັ່ງຊື້"

    _LAO_FONT = "Phetsarath OT"
    header_font = Font(name=_LAO_FONT, bold=True, color="FFFFFF", size=12)
    data_font   = Font(name=_LAO_FONT, size=11)
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ລຳດັບ", "ລະຫັດ", "ສິນຄ້າ", "ລູກຄ້າ", "ເບີໂທ", "ທີ່ຢູ່", "ຈຳນວນ", "ລາຄາລວມ (ກີບ)", "ວັນທີສັ່ງ", "ສະຖານະ"]
    col_widths = [8, 12, 25, 20, 15, 25, 10, 20, 18, 15]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[1].height = 25

    for row_num, order in enumerate(orders, 2):
        row_fill = PatternFill("solid", fgColor="F5F8FF") if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        # ແປງເວລາເປັນ UTC+7 (ລາວ)
        if order.order_date:
            _local_dt = order.order_date.astimezone(_LAO_TZ)
            _date_str = _local_dt.strftime("%d/%m/%Y %H:%M")
        else:
            _date_str = "-"

        row_data = [
            row_num - 1,
            f"#ORDID-{order.id}",
            order.product.name if order.product else "-",
            order.customer.get_name if order.customer else "-",
            order.mobile or "-",
            order.address or "-",
            order.quantity,
            int(order.amount) if order.amount else 0,
            _date_str,
            STATUS_LAO.get(order.status, order.status or '-'),
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[row_num].height = 20

    # ── ROW ສະຫຼຸບລາຄາລວມ ─────────────────────────────────
    total_row   = ws.max_row + 1
    total_font  = Font(name=_LAO_FONT, bold=True, size=12, color="FFFFFF")
    fill_dark   = PatternFill("solid", fgColor="4c1d95")
    fill_purple = PatternFill("solid", fgColor="7c3aed")

    # Label ລວມທັງໝົດ — merge col 1-6
    ws.merge_cells(f"A{total_row}:F{total_row}")
    lbl = ws.cell(row=total_row, column=1, value="ລວມທັງໝົດ")
    lbl.font      = total_font
    lbl.fill      = fill_dark
    lbl.alignment = Alignment(horizontal="center", vertical="center")
    lbl.border    = border

    # ຈຳນວນລວມ (col 7)
    c_qty = ws.cell(row=total_row, column=7, value=total_qty)
    c_qty.font      = total_font
    c_qty.fill      = fill_purple
    c_qty.alignment = Alignment(horizontal="center", vertical="center")
    c_qty.border    = border

    # ລາຄາລວມ (col 8)
    c_amt = ws.cell(row=total_row, column=8, value=total_amount)
    c_amt.font           = Font(name=_LAO_FONT, bold=True, size=13, color="FFFFFF")
    c_amt.fill           = fill_purple
    c_amt.alignment      = Alignment(horizontal="center", vertical="center")
    c_amt.border         = border
    c_amt.number_format  = '#,##0'

    # Fill col 9-10
    for col in range(9, 11):
        cell = ws.cell(row=total_row, column=col, value="")
        cell.fill   = fill_dark
        cell.border = border

    ws.row_dimensions[total_row].height = 28
    # ─────────────────────────────────────────────────────

    if _date_str:
        _fname = f"orders_{_date_str}.xlsx"
    elif _month_str:
        _fname = f"orders_{_month_str}.xlsx"
    elif _year_str:
        _fname = f"orders_{_year_str}.xlsx"
    else:
        _fname = "orders_all.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_fname}"'
    wb.save(response)
    return response


# ---- Export Feedbacks Excel ----
@login_required(login_url='adminlogin')
def export_feedbacks_excel(request):
    feedbacks = models.Feedback.objects.all().order_by('-id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ຄຳຕິຊົມ"

    _LAO_FONT = "Phetsarath OT"
    header_font = Font(name=_LAO_FONT, bold=True, color="FFFFFF", size=12)
    data_font   = Font(name=_LAO_FONT, size=11)
    header_fill = PatternFill("solid", fgColor="5b21b6")
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["ລຳດັບ", "ຊື່ຜູ້ຕິຊົມ", "ຄຳຕິຊົມ", "ວັນທີ"]
    col_widths = [8, 25, 60, 20]

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.row_dimensions[1].height = 25

    for row_num, fb in enumerate(feedbacks, 2):
        row_fill = PatternFill("solid", fgColor="F5F0FF") if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        date_str = fb.date.strftime("%d/%m/%Y") if fb.date else "-"
        row_data = [
            row_num - 1,
            fb.name or "-",
            fb.feedback or "-",
            date_str,
        ]
        aligns = [center, center, left, center]
        for col_num, (value, align) in enumerate(zip(row_data, aligns), 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font      = data_font
            cell.fill      = row_fill
            cell.alignment = align
            cell.border    = border
        ws.row_dimensions[row_num].height = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="feedbacks.xlsx"'
    wb.save(response)
    return response


# ---- Export CSV ----
@login_required(login_url='adminlogin')
def export_orders_csv(request):
    from datetime import timezone as _dtz2, timedelta as _td2
    _LAO2 = _dtz2(_td2(hours=7))
    STATUS_LAO2 = {
        'Pending':    'ຮັບອໍເດີ',
        'Processing': 'ກຳລັງຈັດສົ່ງ',
        'Delivered':  'ຈັດສົ່ງສຳເລັດ',
        'Cancelled':  'ຍົກເລີກ',
        'Completed':  'ສຳເລັດ',
    }
    orders = models.Orders.objects.all().select_related('product', 'customer__user')

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="orders.csv"'

    writer = csv.writer(response)
    writer.writerow(["ລຳດັບ", "ລະຫັດ", "ສິນຄ້າ", "ລູກຄ້າ", "ເບີໂທ", "ທີ່ຢູ່", "ຈຳນວນ", "ລາຄາລວມ", "ວັນທີສັ່ງ", "ສະຖານະ"])

    for i, order in enumerate(orders, 1):
        if order.order_date:
            _ld = order.order_date.astimezone(_LAO2)
            _ds = _ld.strftime("%d/%m/%Y %H:%M")
        else:
            _ds = "-"
        writer.writerow([
            i,
            f"#ORDID-{order.id}",
            order.product.name if order.product else "-",
            order.customer.get_name if order.customer else "-",
            order.mobile or "-",
            order.address or "-",
            order.quantity,
            order.amount,
            _ds,
            STATUS_LAO2.get(order.status, order.status or '-'),
        ])

    return response


# ---- Export PDF ----
@login_required(login_url='adminlogin')
def export_orders_pdf(request):
    orders = models.Orders.objects.all().select_related('product', 'customer__user')
    rows = []
    for i, order in enumerate(orders, 1):
        rows.append({
            'num': i,
            'id': order.id,
            'product': order.product.name if order.product else "-",
            'customer': order.customer.get_name if order.customer else "-",
            'mobile': order.mobile or "-",
            'address': order.address or "-",
            'quantity': order.quantity,
            'amount': order.amount,
            'date': order.order_date.strftime("%d/%m/%Y %H:%M") if order.order_date else "-",
            'status': order.status,
        })
    template = get_template('ecom/export_orders_pdf.html')
    html = template.render({'rows': rows})
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("utf-8")), result, encoding='utf-8')
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="orders.pdf"'
        return response
    return HttpResponse("ເກີດຂໍ້ຜິດພາດໃນການສ້າງ PDF", status=500)


# ---- Export Finance Excel (type-specific) ----
@login_required(login_url='adminlogin')
def export_finance_excel(request):
    import datetime as _dt
    import calendar as _cal
    from datetime import timezone as _dtz, timedelta as _td
    from django.db.models import Sum as _Sum

    _LAO_TZ = _dtz(_td(hours=7))
    now_lao = _dt.datetime.now(_LAO_TZ)
    today   = now_lao.date()

    # --- params ---
    sel_year  = int(request.GET.get('year',  today.year))
    sel_month = int(request.GET.get('month', today.month))
    sel_type  = request.GET.get('type', 'all')  # 'revenue' | 'expense' | 'profit' | 'all'
    sel_date_param = request.GET.get('date', None)   # YYYY-MM-DD → single-day export
    sel_scope = request.GET.get('scope', None)       # 'year' → whole-year export

    _LAO_FONT = "Phetsarath OT"

    def _hfont(bold=True, color="FFFFFF", size=12):
        return Font(name=_LAO_FONT, bold=bold, color=color, size=size)
    def _dfont(bold=False, size=11, color="000000"):
        return Font(name=_LAO_FONT, bold=bold, size=size, color=color)

    _center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _left    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    _right   = Alignment(horizontal="right",  vertical="center")
    _thin    = Side(style="thin", color="CCCCCC")
    _border  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    PURPLE_FILL  = _fill("5B21B6")
    RED_FILL     = _fill("B91C1C")
    GREEN_FILL   = _fill("065F46")
    HEADER_FILL  = _fill("1E3A5F")
    ODD_FILL     = _fill("F8F7FF")
    EVEN_FILL    = _fill("FFFFFF")
    GREEN_LT     = _fill("D1FAE5")
    RED_LT       = _fill("FEE2E2")
    PROFIT_GREEN = "065F46"
    PROFIT_RED   = "B91C1C"

    # ═══════════════════════════════════════════════════════════════
    # SINGLE-DAY EXPORT — triggered when ?date=YYYY-MM-DD is passed
    # ═══════════════════════════════════════════════════════════════
    if sel_date_param:
        try:
            sel_date_d = _dt.date.fromisoformat(sel_date_param)
        except Exception:
            sel_date_d = today

        _LAO_TZ2  = _dtz(_td(hours=7))
        day_start = _dt.datetime(sel_date_d.year, sel_date_d.month, sel_date_d.day, 0, 0, 0, tzinfo=_LAO_TZ2)
        day_end   = _dt.datetime(sel_date_d.year, sel_date_d.month, sel_date_d.day, 23, 59, 59, tzinfo=_LAO_TZ2)
        LAO_DAYS_S = ['ຈັນ','ອັງຄານ','ພຸດ','ພະຫັດ','ສຸກ','ເສົາ','ອາທິດ']
        day_name  = LAO_DAYS_S[sel_date_d.weekday()]
        date_lbl  = sel_date_d.strftime('%d/%m/%Y')

        def _dotfmt(n):
            if n is None: return '—'
            return f"{int(round(float(n))):,}".replace(',', '.') + ' ກີບ'

        wb_d = openpyxl.Workbook()
        ws   = wb_d.active
        ws.sheet_view.showGridLines = False

        if sel_type == 'revenue':
            orders = list(_revenue_orders_qs(day_start, day_end).values('id', 'product__name', 'quantity', 'amount', 'status'))
            total_rev = sum(float(o['amount'] or 0) for o in orders)

            ws.title = f"ລາຍຮັບ {sel_date_d}"
            ws.merge_cells("A1:E1")
            c = ws["A1"]; c.value = f"ລາຍຮັບ ວັນທີ {date_lbl}  ({day_name})"
            c.font = _hfont(size=14); c.fill = _fill("3B82F6"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:E2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22
            for ci, (h, fc) in enumerate(zip(
                ["ລຳດັບ","ສິນຄ້າ","ຈຳນວນ","ລາຍຮັບ (ກີບ)","ສະຖານະ"],
                ["1E3A5F","1E3A5F","1E3A5F","3B82F6","1E3A5F"]
            ), 1):
                cell = ws.cell(row=3, column=ci, value=h)
                cell.font = _hfont(size=11); cell.fill = _fill(fc); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[3].height = 28
            STATUS_LAO = {'Delivered':'ສຳເລັດ','Cancelled':'ຍົກເລີກ','Processing':'ກຳລັງຈັດ','Confirmed':'ຢືນຢັນ','Pending':'ລໍຖ້າ'}
            for i, o in enumerate(orders):
                amt = float(o['amount'] or 0)
                r = i + 4; rf = _fill("EFF6FF") if i%2==0 else _fill("FFFFFF")
                for ci, (v, a) in enumerate(zip(
                    [i+1, o['product__name'] or '—', o['quantity'] or 1, _dotfmt(amt) if amt else '—', STATUS_LAO.get(o['status'], o['status'])],
                    [_center, _left, _center, _right, _center]
                ), 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.font = _dfont(color="1D4ED8" if ci==4 else "374151", bold=(ci==4))
                    cell.fill = rf; cell.alignment = a; cell.border = _border
                ws.row_dimensions[r].height = 22
            tr = len(orders) + 4
            for ci, (v, a) in enumerate(zip(
                ["", f"ລວມ {len(orders)} ອໍເດີ", "", _dotfmt(total_rev), ""],
                [_center, _center, _center, _right, _center]
            ), 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.font = _hfont(size=12)
                cell.fill = _fill("3B82F6") if ci==4 else _fill("1E3A5F")
                cell.alignment = a; cell.border = _border
            ws.row_dimensions[tr].height = 28
            for col, w in zip("ABCDE", [8, 28, 10, 18, 14]):
                ws.column_dimensions[col].width = w
            filename = f"EX_ລາຍຮັບ_{sel_date_d}.xlsx"

        elif sel_type == 'expense':
            expenses = list(models.Expense.objects.filter(
                date=sel_date_d
            ).values('id', 'category', 'description', 'amount'))
            total_exp = sum(float(e['amount'] or 0) for e in expenses)

            ws.title = f"ລາຍຈ່າຍ {sel_date_d}"
            ws.merge_cells("A1:D1")
            c = ws["A1"]; c.value = f"ລາຍຈ່າຍ ວັນທີ {date_lbl}  ({day_name})"
            c.font = _hfont(size=14); c.fill = _fill("B91C1C"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:D2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22
            for ci, (h, fc) in enumerate(zip(
                ["ລຳດັບ","ໝວດໝູ່","ລາຍລະອຽດ","ຈຳນວນ (ກີບ)"],
                ["1E3A5F","B91C1C","1E3A5F","B91C1C"]
            ), 1):
                cell = ws.cell(row=3, column=ci, value=h)
                cell.font = _hfont(size=11); cell.fill = _fill(fc); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[3].height = 28
            for i, e in enumerate(expenses):
                amt = float(e['amount'] or 0)
                r = i + 4; rf = _fill("FEF2F2") if i%2==0 else _fill("FFFFFF")
                for ci, (v, a) in enumerate(zip(
                    [i+1, e['category'], e['description'] or '—', _dotfmt(amt) if amt else '—'],
                    [_center, _center, _left, _right]
                ), 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.font = _dfont(color="EF4444" if ci==4 else "374151", bold=(ci==4))
                    cell.fill = rf; cell.alignment = a; cell.border = _border
                ws.row_dimensions[r].height = 22
            tr = len(expenses) + 4
            for ci, (v, a) in enumerate(zip(
                ["", f"ລວມ {len(expenses)} ລາຍການ", "", _dotfmt(total_exp)],
                [_center, _center, _center, _right]
            ), 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.font = _hfont(size=12)
                cell.fill = _fill("B91C1C") if ci==4 else _fill("1E3A5F")
                cell.alignment = a; cell.border = _border
            ws.row_dimensions[tr].height = 28
            for col, w in zip("ABCD", [8, 20, 32, 18]):
                ws.column_dimensions[col].width = w
            filename = f"EX_ລາຍຈ່າຍ_{sel_date_d}.xlsx"

        else:  # profit
            orders = list(_revenue_orders_qs(day_start, day_end).values('id', 'product__name', 'quantity', 'amount'))
            profit_percent = float(models.FinanceSettings.get_solo().profit_percent)
            total_rev  = sum(float(o['amount'] or 0) for o in orders)
            total_prof = round(total_rev * profit_percent / 100, 2)
            pcolor     = "065F46" if total_prof >= 0 else "B91C1C"
            pbg        = "F0FDF4" if total_prof >= 0 else "FEF2F2"

            ws.title = f"ກຳໄລ {sel_date_d}"
            ws.merge_cells("A1:C1")
            c = ws["A1"]; c.value = f"ສະຫຼຸບກຳໄລ ວັນທີ {date_lbl}  ({day_name})"
            c.font = _hfont(size=14); c.fill = _fill("065F46"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:C2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22

            for ri, (lbl, val, fc, bg) in enumerate([
                ("📦 ລາຍຮັບ", total_rev, "1D4ED8", "EFF6FF"),
                (f"💰 ກຳໄລສຸດທິ ({profit_percent:g}%)", total_prof, pcolor, pbg),
            ], 3):
                ws.merge_cells(f"A{ri}:B{ri}")
                cl = ws.cell(row=ri, column=1, value=lbl)
                cl.font = _hfont(size=13, color=fc); cl.fill = _fill(bg); cl.alignment = _left; cl.border = _border
                cv = ws.cell(row=ri, column=3, value=_dotfmt(val))
                cv.font = _hfont(size=13, color=fc); cv.fill = _fill(bg); cv.alignment = _right; cv.border = _border
                ws.row_dimensions[ri].height = 32

            # Revenue detail block
            r_start = 6
            ws.merge_cells(f"A{r_start}:C{r_start}")
            hd = ws.cell(row=r_start, column=1, value="ລາຍຮັບ — ລາຍລະອຽດ")
            hd.font = _hfont(size=12, color="1D4ED8"); hd.fill = _fill("DBEAFE"); hd.alignment = _center; hd.border = _border
            ws.row_dimensions[r_start].height = 26
            for ci, h in enumerate(["ສິນຄ້າ","ຈຳນວນ","ລາຍຮັບ (ກີບ)"], 1):
                cell = ws.cell(row=r_start+1, column=ci, value=h)
                cell.font = _hfont(size=10); cell.fill = _fill("1D4ED8"); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[r_start+1].height = 24
            for i, o in enumerate(orders):
                r = r_start + 2 + i; rf = _fill("EFF6FF") if i%2==0 else _fill("FFFFFF")
                amt_o = float(o['amount'] or 0)
                for ci, (v, a) in enumerate(zip(
                    [o['product__name'] or '—', o['quantity'] or 1, _dotfmt(amt_o) if amt_o else '—'],
                    [_left, _center, _right]
                ), 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.font = _dfont(color="1D4ED8" if ci==3 else "374151")
                    cell.fill = rf; cell.alignment = a; cell.border = _border
                ws.row_dimensions[r].height = 20

            for col, w in zip("ABC", [28, 32, 20]):
                ws.column_dimensions[col].width = w
            filename = f"EX_ກຳໄລ_{sel_date_d}.xlsx"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb_d.save(response)
        return response

    # ═══════════════════════════════════════════════════════════════
    # MONTH / YEAR, TYPE-SPECIFIC EXPORT
    # triggered by ?scope=month|year&type=revenue|expense|profit
    # ═══════════════════════════════════════════════════════════════
    if sel_scope in ('month', 'year') and sel_type in ('revenue', 'expense', 'profit'):
        def _dotfmt_s(n):
            if n is None: return '—'
            return f"{int(round(float(n))):,}".replace(',', '.') + ' ກີບ'

        if sel_scope == 'month':
            _, _dim = _cal.monthrange(sel_year, sel_month)
            range_start = _dt.date(sel_year, sel_month, 1)
            range_end   = _dt.date(sel_year, sel_month, _dim)
            period_lbl  = f"ເດືອນ {sel_month:02d}/{sel_year}"
            fname_lbl   = f"ເດືອນ{sel_month:02d}-{sel_year}"
        else:
            range_start = _dt.date(sel_year, 1, 1)
            range_end   = _dt.date(sel_year, 12, 31)
            period_lbl  = f"ປີ {sel_year}"
            fname_lbl   = f"ປີ{sel_year}"

        wb_s = openpyxl.Workbook()
        ws = wb_s.active
        ws.sheet_view.showGridLines = False

        _tz_start = _dt.datetime(range_start.year, range_start.month, range_start.day, tzinfo=_LAO_TZ)
        _tz_end   = _dt.datetime(range_end.year, range_end.month, range_end.day, 23, 59, 59, tzinfo=_LAO_TZ)

        if sel_type == 'revenue':
            orders = list(_revenue_orders_qs(_tz_start, _tz_end).select_related('product').order_by('order_date').values(
                'id', 'order_date', 'product__name', 'quantity', 'amount', 'status', 'pickup_date', 'fulfilled_at'
            ))
            total = sum(float(o['amount'] or 0) for o in orders)
            STATUS_LAO = {'Delivered':'ສຳເລັດ','Cancelled':'ຍົກເລີກ','Processing':'ກຳລັງຈັດ','Confirmed':'ຢືນຢັນ','Pending':'ລໍຖ້າ'}

            ws.title = f"ລາຍຮັບ {fname_lbl}"[:31]
            ws.merge_cells("A1:F1")
            c = ws["A1"]; c.value = f"ລາຍຮັບ — {period_lbl}"
            c.font = _hfont(size=14); c.fill = _fill("1D4ED8"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:F2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22
            for ci, (h, fc) in enumerate(zip(
                ["ລຳດັບ","ວັນທີ","ສິນຄ້າ","ຈຳນວນ","ລາຍຮັບ (ກີບ)","ສະຖານະ"],
                ["1E3A5F","1E3A5F","1E3A5F","1E3A5F","1D4ED8","1E3A5F"]
            ), 1):
                cell = ws.cell(row=3, column=ci, value=h)
                cell.font = _hfont(size=11); cell.fill = _fill(fc); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[3].height = 28
            for i, o in enumerate(orders):
                amt = float(o['amount'] or 0)
                r = i + 4; rf = _fill("EFF6FF") if i % 2 == 0 else _fill("FFFFFF")
                _rev_dt = _revenue_date(o)
                odate = _rev_dt.astimezone(_LAO_TZ).strftime('%d/%m/%Y') if _rev_dt else '—'
                for ci, (v, a) in enumerate(zip(
                    [i+1, odate, o['product__name'] or '—', o['quantity'] or 1, _dotfmt_s(amt) if amt else '—', STATUS_LAO.get(o['status'], o['status'])],
                    [_center, _center, _left, _center, _right, _center]
                ), 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.font = _dfont(color="1D4ED8" if ci == 5 else "374151", bold=(ci == 5))
                    cell.fill = rf; cell.alignment = a; cell.border = _border
                ws.row_dimensions[r].height = 22
            tr = len(orders) + 4
            for ci, (v, a) in enumerate(zip(
                ["", f"ລວມ {len(orders)} ອໍເດີ", "", "", _dotfmt_s(total), ""],
                [_center, _center, _center, _center, _right, _center]
            ), 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.font = _hfont(size=12)
                cell.fill = _fill("1D4ED8") if ci == 5 else _fill("1E3A5F")
                cell.alignment = a; cell.border = _border
            ws.row_dimensions[tr].height = 28
            for col, w in zip("ABCDEF", [8, 13, 26, 10, 16, 14]):
                ws.column_dimensions[col].width = w
            filename = f"EX_ລາຍຮັບ_{fname_lbl}.xlsx"

        elif sel_type == 'expense':
            expenses = list(models.Expense.objects.filter(
                date__gte=range_start, date__lte=range_end
            ).order_by('date').values('date', 'category', 'description', 'amount'))
            total = sum(float(e['amount'] or 0) for e in expenses)

            ws.title = f"ລາຍຈ່າຍ {fname_lbl}"[:31]
            ws.merge_cells("A1:E1")
            c = ws["A1"]; c.value = f"ລາຍຈ່າຍ — {period_lbl}"
            c.font = _hfont(size=14); c.fill = _fill("B91C1C"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:E2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22
            for ci, (h, fc) in enumerate(zip(
                ["ລຳດັບ","ວັນທີ","ໝວດໝູ່","ລາຍລະອຽດ","ຈຳນວນ (ກີບ)"],
                ["1E3A5F","1E3A5F","B91C1C","1E3A5F","B91C1C"]
            ), 1):
                cell = ws.cell(row=3, column=ci, value=h)
                cell.font = _hfont(size=11); cell.fill = _fill(fc); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[3].height = 28
            for i, e in enumerate(expenses):
                amt = float(e['amount'] or 0)
                r = i + 4; rf = _fill("FEF2F2") if i % 2 == 0 else _fill("FFFFFF")
                for ci, (v, a) in enumerate(zip(
                    [i+1, e['date'].strftime('%d/%m/%Y'), e['category'], e['description'] or '—', _dotfmt_s(amt) if amt else '—'],
                    [_center, _center, _center, _left, _right]
                ), 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.font = _dfont(color="EF4444" if ci == 5 else "374151", bold=(ci == 5))
                    cell.fill = rf; cell.alignment = a; cell.border = _border
                ws.row_dimensions[r].height = 22
            tr = len(expenses) + 4
            for ci, (v, a) in enumerate(zip(
                ["", f"ລວມ {len(expenses)} ລາຍການ", "", "", _dotfmt_s(total)],
                [_center, _center, _center, _center, _right]
            ), 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.font = _hfont(size=12)
                cell.fill = _fill("B91C1C") if ci == 5 else _fill("1E3A5F")
                cell.alignment = a; cell.border = _border
            ws.row_dimensions[tr].height = 28
            for col, w in zip("ABCDE", [8, 14, 20, 32, 18]):
                ws.column_dimensions[col].width = w
            filename = f"EX_ລາຍຈ່າຍ_{fname_lbl}.xlsx"

        else:  # profit — aggregated per-day (month scope) or per-month (year scope)
            profit_percent = float(models.FinanceSettings.get_solo().profit_percent)
            rev_map = {}
            for o in _revenue_orders_qs(_tz_start, _tz_end).values('order_date', 'amount', 'pickup_date', 'fulfilled_at'):
                d = _revenue_date(o).astimezone(_LAO_TZ).date()
                rev_map[d] = rev_map.get(d, 0.0) + float(o['amount'] or 0)

            ws.title = f"ກຳໄລ {fname_lbl}"[:31]
            ws.merge_cells("A1:C1")
            c = ws["A1"]; c.value = f"ສະຫຼຸບກຳໄລ — {period_lbl}"
            c.font = _hfont(size=14); c.fill = _fill("065F46"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:C2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້  |  ອັດຕາກຳໄລ {profit_percent:g}%"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22
            LAO_DAYS3 = ['ຈັນ','ອັງຄານ','ພຸດ','ພະຫັດ','ສຸກ','ເສົາ','ອາທິດ']
            LAO_MONTHS3 = ['ມັງກອນ','ກຸມພາ','ມີນາ','ເມສາ','ພຶດສະພາ','ມິຖຸນາ','ກໍລະກົດ','ສິງຫາ','ກັນຍາ','ຕຸລາ','ພະຈິກ','ທັນວາ']
            headers_p = ["ວັນ/ເດືອນ","ລາຍຮັບ (ກີບ)","ກຳໄລ (ກີບ)"]
            for ci, (h, fc) in enumerate(zip(headers_p, ["1E3A5F","1D4ED8","065F46"]), 1):
                cell = ws.cell(row=3, column=ci, value=h)
                cell.font = _hfont(size=11); cell.fill = _fill(fc); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[3].height = 28

            total_r = total_p = 0.0
            if sel_scope == 'month':
                _, dim = _cal.monthrange(sel_year, sel_month)
                for i in range(dim):
                    cur = _dt.date(sel_year, sel_month, i + 1)
                    r = rev_map.get(cur, 0.0); p = round(r * profit_percent / 100, 2)
                    total_r += r; total_p += p
                    row = i + 4; rf = _fill("F8F7FF") if i % 2 == 0 else _fill("FFFFFF")
                    lbl = f"{cur.day:02d} ({LAO_DAYS3[cur.weekday()]})"
                    pcolor = "065F46" if p >= 0 else "B91C1C"
                    for ci, (v, a) in enumerate(zip(
                        [lbl, _dotfmt_s(r) if r else '—', (('+' if p >= 0 else '') + _dotfmt_s(p)) if r else '—'],
                        [_left, _right, _right]
                    ), 1):
                        cell = ws.cell(row=row, column=ci, value=v)
                        cell.font = _dfont(color=pcolor if ci == 3 else "374151", bold=(ci == 3))
                        cell.fill = rf; cell.alignment = a; cell.border = _border
                    ws.row_dimensions[row].height = 22
                tr = dim + 4
            else:
                for m in range(1, 13):
                    _, dim = _cal.monthrange(sel_year, m)
                    r = sum(rev_map.get(_dt.date(sel_year, m, dd), 0.0) for dd in range(1, dim + 1))
                    p = round(r * profit_percent / 100, 2)
                    total_r += r; total_p += p
                    row = m + 3; rf = _fill("F8F7FF") if m % 2 == 0 else _fill("FFFFFF")
                    pcolor = "065F46" if p >= 0 else "B91C1C"
                    for ci, (v, a) in enumerate(zip(
                        [LAO_MONTHS3[m-1], _dotfmt_s(r) if r else '—', (('+' if p >= 0 else '') + _dotfmt_s(p)) if r else '—'],
                        [_left, _right, _right]
                    ), 1):
                        cell = ws.cell(row=row, column=ci, value=v)
                        cell.font = _dfont(color=pcolor if ci == 3 else "374151", bold=(ci == 3))
                        cell.fill = rf; cell.alignment = a; cell.border = _border
                    ws.row_dimensions[row].height = 22
                tr = 16

            total_pcolor = "065F46" if total_p >= 0 else "B91C1C"
            for ci, (v, a) in enumerate(zip(
                [f"ລວມ {period_lbl}", _dotfmt_s(total_r), ('+' if total_p >= 0 else '') + _dotfmt_s(total_p)],
                [_left, _right, _right]
            ), 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.font = _hfont(size=12, color=total_pcolor if ci == 3 else "FFFFFF")
                cell.fill = _fill(total_pcolor) if ci == 3 else _fill("1E3A5F")
                cell.alignment = a; cell.border = _border
            ws.row_dimensions[tr].height = 28
            for col, w in zip("ABC", [18, 18, 18]):
                ws.column_dimensions[col].width = w
            filename = f"EX_ກຳໄລ_{fname_lbl}.xlsx"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb_s.save(response)
        return response

    # ═══════════════════════════════════════════════════════════════
    # ALL-TIME, TYPE-SPECIFIC EXPORT — triggered by ?scope=all&type=revenue|expense|profit
    # (unbounded — every record ever, no date filtering at all)
    # ═══════════════════════════════════════════════════════════════
    if sel_scope == 'all' and sel_type in ('revenue', 'expense', 'profit'):
        def _dotfmt_all(n):
            if n is None: return '—'
            return f"{int(round(float(n))):,}".replace(',', '.') + ' ກີບ'

        wb_a = openpyxl.Workbook()
        ws = wb_a.active
        ws.sheet_view.showGridLines = False

        if sel_type == 'revenue':
            orders = list(_revenue_orders_qs().select_related('product').order_by('order_date').values(
                'id', 'order_date', 'product__name', 'quantity', 'amount', 'status', 'pickup_date', 'fulfilled_at'
            ))
            total = sum(float(o['amount'] or 0) for o in orders)
            STATUS_LAO = {'Delivered':'ສຳເລັດ','Cancelled':'ຍົກເລີກ','Processing':'ກຳລັງຈັດ','Confirmed':'ຢືນຢັນ','Pending':'ລໍຖ້າ'}

            ws.title = "ລາຍຮັບ ທັງໝົດ"[:31]
            ws.merge_cells("A1:F1")
            c = ws["A1"]; c.value = "ລາຍຮັບ — ທັງໝົດ (ທຸກໄລຍະເວລາ)"
            c.font = _hfont(size=14); c.fill = _fill("1D4ED8"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:F2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22
            for ci, h in enumerate(["ລຳດັບ","ວັນທີ","ສິນຄ້າ","ຈຳນວນ","ລາຍຮັບ (ກີບ)","ສະຖານະ"], 1):
                cell = ws.cell(row=3, column=ci, value=h)
                cell.font = _hfont(size=11); cell.fill = _fill("1E3A5F"); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[3].height = 28
            for i, o in enumerate(orders):
                r = i + 4; rf = _fill("F8F7FF") if i % 2 == 0 else _fill("FFFFFF")
                _rev_dt = _revenue_date(o)
                odate = _rev_dt.astimezone(_LAO_TZ).strftime('%d/%m/%Y') if _rev_dt else '—'
                for ci, (v, a) in enumerate(zip(
                    [i + 1, odate, o['product__name'] or '—', o['quantity'], _dotfmt_all(o['amount']), STATUS_LAO.get(o['status'], o['status'] or '—')],
                    [_center, _center, _left, _center, _right, _center]
                ), 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.font = _dfont(color="1D4ED8" if ci == 5 else "374151", bold=(ci == 5))
                    cell.fill = rf; cell.alignment = a; cell.border = _border
                ws.row_dimensions[r].height = 20
            tr = len(orders) + 4
            for ci, (v, a) in enumerate(zip(
                ["", "", "", f"ລວມ {len(orders)} ລາຍການ", _dotfmt_all(total), ""],
                [_center, _center, _center, _center, _right, _center]
            ), 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.font = _hfont(size=12); cell.fill = _fill("1E3A5F"); cell.alignment = a; cell.border = _border
            ws.row_dimensions[tr].height = 28
            for col, w in zip("ABCDEF", [8, 14, 26, 10, 16, 14]):
                ws.column_dimensions[col].width = w
            filename = "EX_ລາຍຮັບ_ທັງໝົດ.xlsx"

        elif sel_type == 'expense':
            expenses = list(models.Expense.objects.all().order_by('date', 'id').values(
                'id', 'date', 'category', 'description', 'amount'
            ))
            total = sum(float(e['amount'] or 0) for e in expenses)

            ws.title = "ລາຍຈ່າຍ ທັງໝົດ"[:31]
            ws.merge_cells("A1:E1")
            c = ws["A1"]; c.value = "ລາຍຈ່າຍ — ທັງໝົດ (ທຸກໄລຍະເວລາ)"
            c.font = _hfont(size=14); c.fill = _fill("B91C1C"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:E2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22
            for ci, h in enumerate(["ລຳດັບ","ວັນທີ","ໝວດໝູ່","ລາຍລະອຽດ","ຈຳນວນ (ກີບ)"], 1):
                cell = ws.cell(row=3, column=ci, value=h)
                cell.font = _hfont(size=11); cell.fill = _fill("1E3A5F"); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[3].height = 28
            for i, e in enumerate(expenses):
                r = i + 4; rf = _fill("F8F7FF") if i % 2 == 0 else _fill("FFFFFF")
                for ci, (v, a) in enumerate(zip(
                    [i + 1, e['date'].strftime('%d/%m/%Y') if e['date'] else '—', e['category'], e['description'] or '—', _dotfmt_all(e['amount'])],
                    [_center, _center, _center, _left, _right]
                ), 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.font = _dfont(color="B91C1C" if ci == 5 else "374151", bold=(ci == 5))
                    cell.fill = rf; cell.alignment = a; cell.border = _border
                ws.row_dimensions[r].height = 20
            tr = len(expenses) + 4
            for ci, (v, a) in enumerate(zip(
                ["", "", "", f"ລວມ {len(expenses)} ລາຍການ", _dotfmt_all(total)],
                [_center, _center, _center, _center, _right]
            ), 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.font = _hfont(size=12); cell.fill = _fill("1E3A5F"); cell.alignment = a; cell.border = _border
            ws.row_dimensions[tr].height = 28
            for col, w in zip("ABCDE", [8, 14, 18, 32, 16]):
                ws.column_dimensions[col].width = w
            filename = "EX_ລາຍຈ່າຍ_ທັງໝົດ.xlsx"

        else:  # profit — year-by-year summary across every year that has any data
            from django.db.models import Min as _Min, Max as _Max
            profit_percent = float(models.FinanceSettings.get_solo().profit_percent)
            order_bounds = models.Orders.objects.aggregate(mn=_Min('order_date'), mx=_Max('order_date'))
            years = set()
            if order_bounds['mn']: years.update(range(order_bounds['mn'].astimezone(_LAO_TZ).year, order_bounds['mx'].astimezone(_LAO_TZ).year + 1))
            years = sorted(years) or [today.year]

            ws.title = "ກຳໄລ ທັງໝົດ"[:31]
            ws.merge_cells("A1:C1")
            c = ws["A1"]; c.value = "ຜົນກຳໄລ — ທັງໝົດ (ທຸກປີ)"
            c.font = _hfont(size=14); c.fill = _fill("065F46"); c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells("A2:C2")
            s = ws["A2"]; s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້  |  ອັດຕາກຳໄລ {profit_percent:g}%"
            s.font = _hfont(bold=False, color="94A3B8", size=10); s.fill = _fill("1E293B"); s.alignment = _center
            ws.row_dimensions[2].height = 22
            for ci, (h, fc) in enumerate(zip(["ປີ","ລາຍຮັບ (ກີບ)","ກຳໄລ (ກີບ)"], ["1E3A5F","1D4ED8","065F46"]), 1):
                cell = ws.cell(row=3, column=ci, value=h)
                cell.font = _hfont(size=11); cell.fill = _fill(fc); cell.alignment = _center; cell.border = _border
            ws.row_dimensions[3].height = 28

            total_r = total_p = 0.0
            for i, yr in enumerate(years):
                y_start = _dt.datetime(yr, 1, 1, tzinfo=_LAO_TZ)
                y_end   = _dt.datetime(yr, 12, 31, 23, 59, 59, tzinfo=_LAO_TZ)
                r = float(_revenue_orders_qs(y_start, y_end).aggregate(t=_Sum('amount'))['t'] or 0)
                p = round(r * profit_percent / 100, 2)
                total_r += r; total_p += p
                row = i + 4; rf = _fill("F8F7FF") if i % 2 == 0 else _fill("FFFFFF")
                pcolor = "065F46" if p >= 0 else "B91C1C"
                for ci, (v, a) in enumerate(zip(
                    [str(yr), _dotfmt_all(r) if r else '—', (('+' if p >= 0 else '') + _dotfmt_all(p)) if r else '—'],
                    [_left, _right, _right]
                ), 1):
                    cell = ws.cell(row=row, column=ci, value=v)
                    cell.font = _dfont(color=pcolor if ci == 3 else "374151", bold=(ci == 3))
                    cell.fill = rf; cell.alignment = a; cell.border = _border
                ws.row_dimensions[row].height = 22
            tr = len(years) + 4
            total_pcolor = "065F46" if total_p >= 0 else "B91C1C"
            for ci, (v, a) in enumerate(zip(
                ["ລວມທັງໝົດ", _dotfmt_all(total_r), ('+' if total_p >= 0 else '') + _dotfmt_all(total_p)],
                [_left, _right, _right]
            ), 1):
                cell = ws.cell(row=tr, column=ci, value=v)
                cell.font = _hfont(size=12, color=total_pcolor if ci == 3 else "FFFFFF")
                cell.fill = _fill(total_pcolor) if ci == 3 else _fill("1E3A5F")
                cell.alignment = a; cell.border = _border
            ws.row_dimensions[tr].height = 28
            for col, w in zip("ABC", [14, 18, 18]):
                ws.column_dimensions[col].width = w
            filename = "EX_ກຳໄລ_ທັງໝົດ.xlsx"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb_a.save(response)
        return response

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════
    # SHEET 1 — Daily breakdown for selected month
    # ══════════════════════════════════════════════════
    ws_day = wb.active
    ws_day.title = f"ຕໍ່ມື້ {sel_year}-{sel_month:02d}"
    ws_day.sheet_view.showGridLines = False

    _, days_in_month = _cal.monthrange(sel_year, sel_month)

    # Title row
    ws_day.merge_cells("A1:G1")
    title_cell = ws_day["A1"]
    title_cell.value = f"ລາຍຮັບ-ລາຍຈ່າຍ ຕໍ່ມື້  |  ເດືອນ {sel_month:02d}/{sel_year}"
    title_cell.font      = _hfont(size=14)
    title_cell.fill      = PURPLE_FILL
    title_cell.alignment = _center
    ws_day.row_dimensions[1].height = 36

    # Sub-title row
    ws_day.merge_cells("A2:G2")
    sub = ws_day["A2"]
    sub.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
    sub.font  = _hfont(bold=False, color="94A3B8", size=10)
    sub.fill  = _fill("1E293B")
    sub.alignment = _center
    ws_day.row_dimensions[2].height = 22

    # Header row
    headers_d = ["ລຳດັບ", "ວັນທີ", "ວັນ", "ລາຍຮັບ (ກີບ)", "ລາຍຈ່າຍ (ກີບ)", "ກຳໄລ (ກີບ)", "ໝາຍເຫດ"]
    col_fills  = [HEADER_FILL, HEADER_FILL, HEADER_FILL, PURPLE_FILL, RED_FILL, GREEN_FILL, HEADER_FILL]
    for ci, (h, f) in enumerate(zip(headers_d, col_fills), 1):
        cell = ws_day.cell(row=3, column=ci, value=h)
        cell.font      = _hfont(size=11)
        cell.fill      = f
        cell.alignment = _center
        cell.border    = _border
    ws_day.row_dimensions[3].height = 28

    LAO_DAYS = ['ຈັນ','ອັງຄານ','ພຸດ','ພະຫັດ','ສຸກ','ເສົາ','ອາທິດ']

    # Fetch data
    month_start_d = _dt.date(sel_year, sel_month, 1)
    month_end_d   = _dt.date(sel_year, sel_month, days_in_month)

    # Revenue per day
    _tz_start = _dt.datetime(sel_year, sel_month, 1, tzinfo=_LAO_TZ)
    _tz_end   = _dt.datetime(sel_year, sel_month, days_in_month, 23, 59, 59, tzinfo=_LAO_TZ)
    rev_map = {}
    for o in _revenue_orders_qs(_tz_start, _tz_end).values('order_date', 'amount', 'pickup_date', 'fulfilled_at'):
        d = _revenue_date(o).astimezone(_LAO_TZ).date()
        rev_map[d] = rev_map.get(d, 0.0) + float(o['amount'] or 0)

    # Expense per day
    exp_map = {}
    for e in models.Expense.objects.filter(
        date__gte=month_start_d, date__lte=month_end_d
    ).values('date', 'amount'):
        d = e['date']
        exp_map[d] = exp_map.get(d, 0.0) + float(e['amount'] or 0)

    total_rev_d = total_exp_d = total_prof_d = 0.0
    for i in range(days_in_month):
        cur_date = _dt.date(sel_year, sel_month, i + 1)
        rev  = rev_map.get(cur_date, 0.0)
        exp  = exp_map.get(cur_date, 0.0)
        prof = rev - exp
        total_rev_d  += rev
        total_exp_d  += exp
        total_prof_d += prof

        row_num  = i + 4
        row_fill = ODD_FILL if i % 2 == 0 else EVEN_FILL
        remark   = "ມີຂໍ້ມູນ" if (rev > 0 or exp > 0) else "—"
        day_name = LAO_DAYS[cur_date.weekday()]

        row_vals  = [i + 1, str(cur_date), day_name, rev or None, exp or None, prof if (rev > 0 or exp > 0) else None, remark]
        row_aligns = [_center, _center, _center, _right, _right, _right, _center]
        row_colors = [None, None, None,
                      "7C3AED" if rev > 0 else None,
                      "EF4444" if exp > 0 else None,
                      PROFIT_GREEN if prof > 0 else (PROFIT_RED if prof < 0 else None),
                      None]

        for ci, (val, aln, col) in enumerate(zip(row_vals, row_aligns, row_colors), 1):
            cell = ws_day.cell(row=row_num, column=ci, value=val)
            cell.font   = _dfont(color=col or "374151", bold=(ci == 6 and (rev > 0 or exp > 0)))
            cell.fill   = GREEN_LT if prof > 0 and ci == 6 else (RED_LT if prof < 0 and ci == 6 else row_fill)
            cell.alignment = aln
            cell.border = _border
        ws_day.row_dimensions[row_num].height = 22

    # Totals row
    total_row = days_in_month + 4
    total_vals = ["", "ລວມທັງໝົດ", "", total_rev_d, total_exp_d, total_prof_d, ""]
    total_aligns = [_center, _center, _center, _right, _right, _right, _center]
    for ci, (val, aln) in enumerate(zip(total_vals, total_aligns), 1):
        cell = ws_day.cell(row=total_row, column=ci, value=val)
        cell.font   = _hfont(size=12, color="FFFFFF")
        cell.fill   = GREEN_FILL if ci == 6 else (RED_FILL if ci == 5 else (PURPLE_FILL if ci == 4 else HEADER_FILL))
        cell.alignment = aln
        cell.border = _border
    ws_day.row_dimensions[total_row].height = 28

    ws_day.column_dimensions["A"].width = 8
    ws_day.column_dimensions["B"].width = 14
    ws_day.column_dimensions["C"].width = 12
    ws_day.column_dimensions["D"].width = 18
    ws_day.column_dimensions["E"].width = 18
    ws_day.column_dimensions["F"].width = 18
    ws_day.column_dimensions["G"].width = 12

    # ══════════════════════════════════════════════════
    # SHEET 2 — Monthly breakdown for selected year
    # ══════════════════════════════════════════════════
    ws_mon = wb.create_sheet(title=f"ຕໍ່ເດືອນ {sel_year}")
    ws_mon.sheet_view.showGridLines = False

    # Title
    ws_mon.merge_cells("A1:G1")
    t2 = ws_mon["A1"]
    t2.value = f"ລາຍຮັບ-ລາຍຈ່າຍ ຕໍ່ເດືອນ  |  ປີ {sel_year}"
    t2.font      = _hfont(size=14)
    t2.fill      = PURPLE_FILL
    t2.alignment = _center
    ws_mon.row_dimensions[1].height = 36

    ws_mon.merge_cells("A2:G2")
    s2 = ws_mon["A2"]
    s2.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
    s2.font  = _hfont(bold=False, color="94A3B8", size=10)
    s2.fill  = _fill("1E293B")
    s2.alignment = _center
    ws_mon.row_dimensions[2].height = 22

    # Header
    LAO_MONTHS = ['ມັງກອນ','ກຸມພາ','ມີນາ','ເມສາ','ພຶດສະພາ','ມິຖຸນາ','ກໍລະກົດ','ສິງຫາ','ກັນຍາ','ຕຸລາ','ພະຈິກ','ທັນວາ']
    headers_m  = ["ລຳດັບ", "ເດືອນ", "ເດືອນ (ລາວ)", "ລາຍຮັບ (ກີບ)", "ລາຍຈ່າຍ (ກີບ)", "ກຳໄລ (ກີບ)", "% ກຳໄລ"]
    for ci, h in enumerate(headers_m, 1):
        cell = ws_mon.cell(row=3, column=ci, value=h)
        cell.font      = _hfont(size=11)
        cell.fill      = HEADER_FILL
        cell.alignment = _center
        cell.border    = _border
    ws_mon.row_dimensions[3].height = 28

    # Fetch yearly data
    year_tz_s = _dt.datetime(sel_year, 1, 1, tzinfo=_LAO_TZ)
    year_tz_e = _dt.datetime(sel_year, 12, 31, 23, 59, 59, tzinfo=_LAO_TZ)
    rev_month = [0.0] * 12
    for o in _revenue_orders_qs(year_tz_s, year_tz_e).values('order_date', 'amount', 'pickup_date', 'fulfilled_at'):
        m = _revenue_date(o).astimezone(_LAO_TZ).month - 1
        rev_month[m] += float(o['amount'] or 0)

    exp_month = [0.0] * 12
    for e in models.Expense.objects.filter(date__year=sel_year).values('date', 'amount'):
        exp_month[e['date'].month - 1] += float(e['amount'] or 0)

    total_rev_m = total_exp_m = total_prof_m = 0.0
    for i in range(12):
        rev  = rev_month[i]
        exp  = exp_month[i]
        prof = rev - exp
        pct  = round((prof / rev * 100), 1) if rev > 0 else 0.0
        total_rev_m  += rev
        total_exp_m  += exp
        total_prof_m += prof

        row_num  = i + 4
        row_fill = ODD_FILL if i % 2 == 0 else EVEN_FILL
        pct_str  = f"{pct}%" if (rev > 0 or exp > 0) else "—"

        row_vals   = [i + 1, f"{i+1:02d}/{sel_year}", LAO_MONTHS[i], rev or None, exp or None,
                      prof if (rev > 0 or exp > 0) else None, pct_str if (rev > 0 or exp > 0) else "—"]
        row_aligns = [_center, _center, _left, _right, _right, _right, _center]

        for ci, (val, aln) in enumerate(zip(row_vals, row_aligns), 1):
            cell = ws_mon.cell(row=row_num, column=ci, value=val)
            has_data = (rev > 0 or exp > 0)
            if ci == 6 and has_data:
                cell.font = _dfont(bold=True, color=PROFIT_GREEN if prof >= 0 else PROFIT_RED)
                cell.fill = GREEN_LT if prof >= 0 else RED_LT
            elif ci == 4 and rev > 0:
                cell.font = _dfont(color="7C3AED")
                cell.fill = row_fill
            elif ci == 5 and exp > 0:
                cell.font = _dfont(color="EF4444")
                cell.fill = row_fill
            else:
                cell.font = _dfont()
                cell.fill = row_fill
            cell.alignment = aln
            cell.border    = _border
        ws_mon.row_dimensions[row_num].height = 24

    # Totals
    total_pct = round((total_prof_m / total_rev_m * 100), 1) if total_rev_m > 0 else 0.0
    total_row_m = 16
    total_vals_m = ["", "ລວມທັງໝົດ", "", total_rev_m, total_exp_m, total_prof_m, f"{total_pct}%"]
    for ci, (val, aln) in enumerate(zip(total_vals_m, [_center, _center, _center, _right, _right, _right, _center]), 1):
        cell = ws_mon.cell(row=total_row_m, column=ci, value=val)
        cell.font      = _hfont(size=12, color="FFFFFF")
        cell.fill      = GREEN_FILL if ci == 6 else (RED_FILL if ci == 5 else (PURPLE_FILL if ci == 4 else HEADER_FILL))
        cell.alignment = aln
        cell.border    = _border
    ws_mon.row_dimensions[total_row_m].height = 28

    ws_mon.column_dimensions["A"].width = 8
    ws_mon.column_dimensions["B"].width = 14
    ws_mon.column_dimensions["C"].width = 16
    ws_mon.column_dimensions["D"].width = 20
    ws_mon.column_dimensions["E"].width = 20
    ws_mon.column_dimensions["F"].width = 20
    ws_mon.column_dimensions["G"].width = 12

    LAO_MONTHS = ['ມັງກອນ','ກຸມພາ','ມີນາ','ເມສາ','ພຶດສະພາ','ມິຖຸນາ',
                  'ກໍລະກົດ','ສິງຫາ','ກັນຍາ','ຕຸລາ','ພະຈິກ','ທັນວາ']

    # ══════════════════════════════════════════
    # TYPE-SPECIFIC EXPORT — override wb above
    # ══════════════════════════════════════════
    if sel_type in ('revenue', 'expense', 'profit'):
        wb = openpyxl.Workbook()

        _, days_in_month = _cal.monthrange(sel_year, sel_month)
        month_start_d = _dt.date(sel_year, sel_month, 1)
        month_end_d   = _dt.date(sel_year, sel_month, days_in_month)
        _tz_ms = _dt.datetime(sel_year, sel_month, 1, tzinfo=_LAO_TZ)
        _tz_me = _dt.datetime(sel_year, sel_month, days_in_month, 23, 59, 59, tzinfo=_LAO_TZ)
        _tz_ys = _dt.datetime(sel_year, 1, 1, tzinfo=_LAO_TZ)
        _tz_ye = _dt.datetime(sel_year, 12, 31, 23, 59, 59, tzinfo=_LAO_TZ)
        LAO_DAYS = ['ຈັນ','ອັງຄານ','ພຸດ','ພະຫັດ','ສຸກ','ເສົາ','ອາທິດ']

        def _title_row(ws, text, cols, fill_color="5B21B6"):
            ws.merge_cells(f"A1:{chr(64+cols)}1")
            c = ws["A1"]
            c.value = text
            c.font      = _hfont(size=14)
            c.fill      = _fill(fill_color)
            c.alignment = _center
            ws.row_dimensions[1].height = 36
            ws.merge_cells(f"A2:{chr(64+cols)}2")
            s = ws["A2"]
            s.value = f"Export: {today}  |  ຮ້ານ EX ມໍເຕີ້"
            s.font  = _hfont(bold=False, color="94A3B8", size=10)
            s.fill  = _fill("1E293B")
            s.alignment = _center
            ws.row_dimensions[2].height = 22

        def _header_row(ws, headers, fills, row=3):
            for ci, (h, f) in enumerate(zip(headers, fills), 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = _hfont(size=11); c.fill = f
                c.alignment = _center; c.border = _border
            ws.row_dimensions[row].height = 28

        # ── Build rev_map, exp_map, cnt_map for month ──
        rev_map_d = {}; cnt_map_d = {}
        for o in _revenue_orders_qs(_tz_ms, _tz_me).values('order_date', 'amount', 'pickup_date', 'fulfilled_at'):
            d = _revenue_date(o).astimezone(_LAO_TZ).date()
            rev_map_d[d] = rev_map_d.get(d, 0.0) + float(o['amount'] or 0)
            cnt_map_d[d] = cnt_map_d.get(d, 0) + 1

        exp_map_d = {}
        for e in models.Expense.objects.filter(
            date__gte=month_start_d, date__lte=month_end_d
        ).values('date', 'amount'):
            d = e['date']
            exp_map_d[d] = exp_map_d.get(d, 0.0) + float(e['amount'] or 0)

        # ── Build yearly summaries ──
        rev_map_m = [0.0]*12; cnt_map_m = [0]*12
        for o in _revenue_orders_qs(_tz_ys, _tz_ye).values('order_date', 'amount', 'pickup_date', 'fulfilled_at'):
            m = _revenue_date(o).astimezone(_LAO_TZ).month - 1
            rev_map_m[m] += float(o['amount'] or 0)
            cnt_map_m[m] += 1
        exp_map_m = [0.0]*12
        for e in models.Expense.objects.filter(date__year=sel_year).values('date', 'amount'):
            exp_map_m[e['date'].month - 1] += float(e['amount'] or 0)

        # ════════════ REVENUE EXPORT ════════════
        if sel_type == 'revenue':
            BLUE_FILL   = _fill("1D4ED8")
            BLUE_LT     = _fill("EFF6FF")
            BLUE_DARK   = _fill("1E3A8A")

            # Sheet 1 – daily revenue
            ws1 = wb.active
            ws1.title = f"ລາຍຮັບ ຕໍ່ວັນ {sel_year}-{sel_month:02d}"
            ws1.sheet_view.showGridLines = False
            _title_row(ws1, f"ລາຍຮັບ ຕໍ່ວັນ  |  ເດືອນ {sel_month:02d}/{sel_year}", 5, "1D4ED8")
            _header_row(ws1,
                ["ລຳດັບ","ວັນທີ","ວັນ","ຈຳນວນອໍເດີ","ລາຍຮັບ (ກີບ)"],
                [BLUE_DARK, BLUE_DARK, BLUE_DARK, BLUE_FILL, BLUE_FILL])
            tot_rev = 0; tot_cnt = 0
            for i in range(days_in_month):
                d    = _dt.date(sel_year, sel_month, i + 1)
                rev  = rev_map_d.get(d, 0.0)
                cnt  = cnt_map_d.get(d, 0)
                tot_rev += rev; tot_cnt += cnt
                r    = i + 4
                rf   = ODD_FILL if i % 2 == 0 else EVEN_FILL
                for ci, (val, aln) in enumerate(zip(
                    [i+1, str(d), LAO_DAYS[d.weekday()], cnt or None, rev or None],
                    [_center, _center, _center, _center, _right]
                ), 1):
                    c = ws1.cell(row=r, column=ci, value=val)
                    c.font = _dfont(color="1D4ED8" if ci == 5 and rev > 0 else "374151",
                                    bold=(ci == 5 and rev > 0))
                    c.fill = BLUE_LT if (ci == 5 and rev > 0) else rf
                    c.alignment = aln; c.border = _border
                ws1.row_dimensions[r].height = 22
            tr = days_in_month + 4
            for ci, (v, a) in enumerate(zip(
                ["","ລວມທັງໝົດ","",tot_cnt,tot_rev],
                [_center,_center,_center,_center,_right]
            ), 1):
                c = ws1.cell(row=tr, column=ci, value=v)
                c.font = _hfont(size=12); c.border = _border
                c.fill = BLUE_FILL if ci == 5 else BLUE_DARK
                c.alignment = a
            ws1.row_dimensions[tr].height = 28
            for col, w in zip("ABCDE", [8,14,12,16,20]):
                ws1.column_dimensions[col].width = w

            # Sheet 2 – monthly revenue
            ws2 = wb.create_sheet(f"ລາຍຮັບ ຕໍ່ເດືອນ {sel_year}")
            ws2.sheet_view.showGridLines = False
            _title_row(ws2, f"ລາຍຮັບ ຕໍ່ເດືອນ  |  ປີ {sel_year}", 4, "1D4ED8")
            _header_row(ws2,
                ["ລຳດັບ","ເດືອນ (ລາວ)","ຈຳນວນອໍເດີ","ລາຍຮັບ (ກີບ)"],
                [BLUE_DARK, BLUE_DARK, BLUE_FILL, BLUE_FILL])
            ytot_rev = 0; ytot_cnt = 0
            for i in range(12):
                rev = rev_map_m[i]; cnt = cnt_map_m[i]
                ytot_rev += rev; ytot_cnt += cnt
                r = i + 4; rf = ODD_FILL if i % 2 == 0 else EVEN_FILL
                for ci, (v, a) in enumerate(zip(
                    [i+1, LAO_MONTHS[i], cnt or None, rev or None],
                    [_center, _left, _center, _right]
                ), 1):
                    c = ws2.cell(row=r, column=ci, value=v)
                    c.font = _dfont(color="1D4ED8" if ci == 4 and rev > 0 else "374151",
                                    bold=(ci == 4 and rev > 0))
                    c.fill = BLUE_LT if (ci == 4 and rev > 0) else rf
                    c.alignment = a; c.border = _border
                ws2.row_dimensions[r].height = 24
            tr2 = 16
            for ci, (v, a) in enumerate(zip(
                ["","ລວມທັງໝົດ",ytot_cnt,ytot_rev],
                [_center,_center,_center,_right]
            ), 1):
                c = ws2.cell(row=tr2, column=ci, value=v)
                c.font = _hfont(size=12); c.border = _border
                c.fill = BLUE_FILL if ci == 4 else BLUE_DARK; c.alignment = a
            ws2.row_dimensions[tr2].height = 28
            for col, w in zip("ABCD", [8,20,16,22]):
                ws2.column_dimensions[col].width = w
            filename = f"EX_ລາຍຮັບ_{sel_year}_{sel_month:02d}.xlsx"

        # ════════════ EXPENSE EXPORT ════════════
        elif sel_type == 'expense':
            RED_FILL = _fill("B91C1C")
            RED_LT2  = _fill("FEF2F2")
            RED_DARK = _fill("7F1D1D")

            # Sheet 1 – expense transactions for selected month
            ws1 = wb.active
            ws1.title = f"ລາຍຈ່າຍ {sel_year}-{sel_month:02d}"
            ws1.sheet_view.showGridLines = False
            _title_row(ws1, f"ລາຍຈ່າຍ ລາຍລະອຽດ  |  ເດືອນ {sel_month:02d}/{sel_year}", 5, "B91C1C")
            _header_row(ws1,
                ["ລຳດັບ","ວັນທີ","ໝວດໝູ່","ລາຍລະອຽດ","ຈຳນວນ (ກີບ)"],
                [RED_DARK, RED_DARK, RED_FILL, RED_DARK, RED_FILL])
            exps = list(models.Expense.objects.filter(
                date__gte=month_start_d, date__lte=month_end_d
            ).order_by('date', 'id').values('id','date','category','description','amount'))
            tot_exp = 0
            for i, e in enumerate(exps):
                amt = float(e['amount'] or 0); tot_exp += amt
                r = i + 4; rf = ODD_FILL if i % 2 == 0 else EVEN_FILL
                for ci, (v, a) in enumerate(zip(
                    [i+1, str(e['date']), e['category'], e['description'] or '—', amt or None],
                    [_center, _center, _center, _left, _right]
                ), 1):
                    c = ws1.cell(row=r, column=ci, value=v)
                    c.font = _dfont(color="B91C1C" if ci == 5 and amt > 0 else "374151",
                                    bold=(ci == 5 and amt > 0))
                    c.fill = RED_LT2 if (ci == 5 and amt > 0) else rf
                    c.alignment = a; c.border = _border
                ws1.row_dimensions[r].height = 22
            tr = len(exps) + 4
            for ci, (v, a) in enumerate(zip(
                ["",f"ລວມ {len(exps)} ລາຍການ","","",tot_exp],
                [_center,_left,_center,_center,_right]
            ), 1):
                c = ws1.cell(row=tr, column=ci, value=v)
                c.font = _hfont(size=12); c.border = _border
                c.fill = RED_FILL if ci == 5 else RED_DARK; c.alignment = a
            ws1.row_dimensions[tr].height = 28
            for col, w in zip("ABCDE", [8,14,18,32,20]):
                ws1.column_dimensions[col].width = w

            # Sheet 2 – monthly expense summary
            ws2 = wb.create_sheet(f"ລາຍຈ່າຍ ຕໍ່ເດືອນ {sel_year}")
            ws2.sheet_view.showGridLines = False
            _title_row(ws2, f"ລາຍຈ່າຍ ຕໍ່ເດືອນ  |  ປີ {sel_year}", 3, "B91C1C")
            _header_row(ws2,
                ["ລຳດັບ","ເດືອນ (ລາວ)","ລາຍຈ່າຍ (ກີບ)"],
                [RED_DARK, RED_DARK, RED_FILL])
            ytot_exp = 0
            for i in range(12):
                exp = exp_map_m[i]; ytot_exp += exp
                r = i + 4; rf = ODD_FILL if i % 2 == 0 else EVEN_FILL
                for ci, (v, a) in enumerate(zip(
                    [i+1, LAO_MONTHS[i], exp or None],
                    [_center, _left, _right]
                ), 1):
                    c = ws2.cell(row=r, column=ci, value=v)
                    c.font = _dfont(color="B91C1C" if ci == 3 and exp > 0 else "374151",
                                    bold=(ci == 3 and exp > 0))
                    c.fill = RED_LT2 if (ci == 3 and exp > 0) else rf
                    c.alignment = a; c.border = _border
                ws2.row_dimensions[r].height = 24
            tr2 = 16
            for ci, (v, a) in enumerate(zip(
                ["","ລວມທັງໝົດ",ytot_exp],
                [_center,_center,_right]
            ), 1):
                c = ws2.cell(row=tr2, column=ci, value=v)
                c.font = _hfont(size=12); c.border = _border
                c.fill = RED_FILL if ci == 3 else RED_DARK; c.alignment = a
            ws2.row_dimensions[tr2].height = 28
            for col, w in zip("ABC", [8,22,22]):
                ws2.column_dimensions[col].width = w
            filename = f"EX_ລາຍຈ່າຍ_{sel_year}_{sel_month:02d}.xlsx"

        # ════════════ PROFIT EXPORT ════════════
        elif sel_type == 'profit':
            GRN_FILL = _fill("065F46")
            GRN_LT2  = _fill("D1FAE5")
            GRN_DARK = _fill("022C22")

            # Sheet 1 – daily profit
            ws1 = wb.active
            ws1.title = f"ກຳໄລ ຕໍ່ວັນ {sel_year}-{sel_month:02d}"
            ws1.sheet_view.showGridLines = False
            _title_row(ws1, f"ກຳໄລ ຕໍ່ວັນ  |  ເດືອນ {sel_month:02d}/{sel_year}", 6, "065F46")
            _header_row(ws1,
                ["ລຳດັບ","ວັນທີ","ວັນ","ລາຍຮັບ (ກີບ)","ລາຍຈ່າຍ (ກີບ)","ກຳໄລ (ກີບ)"],
                [GRN_DARK,GRN_DARK,GRN_DARK,_fill("1D4ED8"),_fill("B91C1C"),GRN_FILL])
            tot_rev = tot_exp = tot_prof = 0.0
            for i in range(days_in_month):
                d    = _dt.date(sel_year, sel_month, i + 1)
                rev  = rev_map_d.get(d, 0.0)
                exp  = exp_map_d.get(d, 0.0)
                prof = rev - exp
                tot_rev += rev; tot_exp += exp; tot_prof += prof
                r = i + 4; rf = ODD_FILL if i % 2 == 0 else EVEN_FILL
                has = rev > 0 or exp > 0
                for ci, (v, a) in enumerate(zip(
                    [i+1, str(d), LAO_DAYS[d.weekday()], rev or None, exp or None,
                     prof if has else None],
                    [_center,_center,_center,_right,_right,_right]
                ), 1):
                    c = ws1.cell(row=r, column=ci, value=v)
                    if ci == 6 and has:
                        c.font = _dfont(bold=True, color=PROFIT_GREEN if prof >= 0 else PROFIT_RED)
                        c.fill = GREEN_LT if prof >= 0 else RED_LT
                    elif ci == 4 and rev > 0:
                        c.font = _dfont(color="1D4ED8"); c.fill = rf
                    elif ci == 5 and exp > 0:
                        c.font = _dfont(color="B91C1C"); c.fill = rf
                    else:
                        c.font = _dfont(); c.fill = rf
                    c.alignment = a; c.border = _border
                ws1.row_dimensions[r].height = 22
            tr = days_in_month + 4
            for ci, (v, a) in enumerate(zip(
                ["","ລວມທັງໝົດ","",tot_rev,tot_exp,tot_prof],
                [_center,_center,_center,_right,_right,_right]
            ), 1):
                c = ws1.cell(row=tr, column=ci, value=v)
                c.font = _hfont(size=12); c.border = _border
                c.fill = (GRN_FILL if ci==6 else (_fill("B91C1C") if ci==5
                         else (_fill("1D4ED8") if ci==4 else GRN_DARK)))
                c.alignment = a
            ws1.row_dimensions[tr].height = 28
            for col, w in zip("ABCDEF", [8,14,12,20,20,20]):
                ws1.column_dimensions[col].width = w

            # Sheet 2 – monthly profit
            ws2 = wb.create_sheet(f"ກຳໄລ ຕໍ່ເດືອນ {sel_year}")
            ws2.sheet_view.showGridLines = False
            _title_row(ws2, f"ກຳໄລ ຕໍ່ເດືອນ  |  ປີ {sel_year}", 5, "065F46")
            _header_row(ws2,
                ["ລຳດັບ","ເດືອນ (ລາວ)","ລາຍຮັບ (ກີບ)","ລາຍຈ່າຍ (ກີບ)","ກຳໄລ (ກີບ)"],
                [GRN_DARK,GRN_DARK,_fill("1D4ED8"),_fill("B91C1C"),GRN_FILL])
            ytot_rev = ytot_exp = ytot_prof = 0.0
            for i in range(12):
                rev  = rev_map_m[i]; exp = exp_map_m[i]; prof = rev - exp
                pct  = round(prof / rev * 100, 1) if rev > 0 else 0.0
                ytot_rev += rev; ytot_exp += exp; ytot_prof += prof
                r = i + 4; rf = ODD_FILL if i % 2 == 0 else EVEN_FILL
                has = rev > 0 or exp > 0
                for ci, (v, a) in enumerate(zip(
                    [i+1, LAO_MONTHS[i], rev or None, exp or None, prof if has else None],
                    [_center,_left,_right,_right,_right]
                ), 1):
                    c = ws2.cell(row=r, column=ci, value=v)
                    if ci == 5 and has:
                        c.font = _dfont(bold=True, color=PROFIT_GREEN if prof >= 0 else PROFIT_RED)
                        c.fill = GREEN_LT if prof >= 0 else RED_LT
                    elif ci == 3 and rev > 0:
                        c.font = _dfont(color="1D4ED8"); c.fill = rf
                    elif ci == 4 and exp > 0:
                        c.font = _dfont(color="B91C1C"); c.fill = rf
                    else:
                        c.font = _dfont(); c.fill = rf
                    c.alignment = a; c.border = _border
                ws2.row_dimensions[r].height = 24
            tr2 = 16
            for ci, (v, a) in enumerate(zip(
                ["","ລວມທັງໝົດ",ytot_rev,ytot_exp,ytot_prof],
                [_center,_center,_right,_right,_right]
            ), 1):
                c = ws2.cell(row=tr2, column=ci, value=v)
                c.font = _hfont(size=12); c.border = _border
                c.fill = (GRN_FILL if ci==5 else (_fill("B91C1C") if ci==4
                         else (_fill("1D4ED8") if ci==3 else GRN_DARK)))
                c.alignment = a
            ws2.row_dimensions[tr2].height = 28
            for col, w in zip("ABCDE", [8,20,22,22,22]):
                ws2.column_dimensions[col].width = w
            filename = f"EX_ກຳໄລ_{sel_year}_{sel_month:02d}.xlsx"

    else:
        # default (type=all) — use the wb already built above (daily + monthly sheets)
        filename = f"EX_ສະຫຼຸບ_{sel_year}_{sel_month:02d}.xlsx"

    # ── Response ──
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response